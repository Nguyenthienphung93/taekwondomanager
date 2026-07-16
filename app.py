from datetime import datetime, date, timedelta, timezone
import calendar
import hashlib
import json
import os
import shutil
import subprocess
import tarfile
import sys
from io import BytesIO
from PIL import Image, ImageOps, ImageFilter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file
from supabase_client import supabase
from utils import auto_generate_license, is_active, calc_tuition, format_money, build_month_codes
import re
import unicodedata
from pathlib import Path
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "change-this-secret")
BASE_DIR = Path(__file__).resolve().parent
STUDENT_PHOTO_DIR = BASE_DIR / "static" / "student_photos"
STUDENT_PHOTO_DIR.mkdir(parents=True, exist_ok=True)

ALLOWED_PHOTO_EXTENSIONS = {"jpg", "jpeg", "png", "webp"}
SUPABASE_URL = os.environ.get(
    "SUPABASE_URL",
    "https://hbmqqvrhjnhxgrxvlthm.supabase.co"
)

STUDENT_PHOTO_BUCKET = "student-photos"
CLUB_ASSET_BUCKET = "system-assets"


DEFAULT_APP_SETTINGS = {
    "header": {
        "club_small_title": "CÂU LẠC BỘ TAEKWONDO",
        "club_name": "HOA HƯỚNG DƯƠNG",
        "logo_url": "https://hbmqqvrhjnhxgrxvlthm.supabase.co/storage/v1/object/public/system-assets/logo.png"
    },
    "fees": {
        "tuition_fee": 500000,
        "exam_fee": 300000,
        "dan_fees": {
            "1 Đẳng": 950000,
            "2 Đẳng": 650000,
            "3 Đẳng": 800000,
            "4 Đẳng": 1250000,
            "5 Đẳng": 1500000,
            "6 Đẳng": 2000000
        }
    },
    "exam": {
        "exam_number_prefix": "Cấp_"
    },
    "class_options": {
        "classrooms": ["2 - 4 - 6", "3 - 5 - 7", "T7 - CN", "Hẹn hò"],
        "timeclasses": ["Ca 1", "Ca 2"],
        "clubs": ["Hoa Hướng Dương Q1(CLB_00019)", "Hoa Hướng Dương Q10(CLB_00104)"]
    },

    "club_info": {
        "intro_title": "CLB Taekwondo Sunflower - Hoa Hướng Dương Diên Hồng",
        "intro_subtitle": "Sunflower Taekwondo Club",
        "intro_content": (
            "CLB Taekwondo Sunflower, còn được biết đến với tên CLB Taekwondo Hoa Hướng Dương, "
            "là môi trường tập luyện Taekwondo dành cho võ sinh nhiều lứa tuổi. Hiện nay cơ sở Q10 "
            "hoạt động với tên CLB Hoa Hướng Dương - Diên Hồng, định hướng đào tạo võ sinh theo tinh thần "
            "kỷ luật, tự tin, lễ phép và phát triển thể chất toàn diện.\n\n"
            "CLB chú trọng nền tảng kỹ thuật căn bản, rèn luyện thể lực, tác phong võ đạo, đồng thời tạo điều kiện "
            "cho võ sinh tham gia thi nâng cấp đai, biểu diễn, giao lưu và các hoạt động phong trào phù hợp."
        ),

        "head_coach": {
            "name": "Nguyễn Thiên Phụng",
            "role": "HLV Trưởng",
            "phone": "",
            "qualification": "",
            "photo_url": "",
            "description": (
                "Nguyễn Thiên Phụng là VĐV Taekwondo Poomsae Việt Nam, nhiều năm thi đấu đỉnh cao "
                "và đạt nhiều thành tích nổi bật trong nước cũng như quốc tế. Hiện tại phụ trách định hướng "
                "chuyên môn chung, xây dựng chương trình huấn luyện, theo dõi chất lượng đào tạo và phát triển "
                "phong trào của CLB Hoa Hướng Dương - Diên Hồng."
            )
        },

        "registrar": {
            "name": "Đang cập nhật",
            "role": "Người ghi danh",
            "phone": "",
            "photo_url": "",
            "description": (
                "Phụ trách tư vấn, tiếp nhận thông tin đăng ký, hỗ trợ phụ huynh và võ sinh "
                "trong quá trình ghi danh, đóng học phí và theo dõi lịch học tại CLB."
            )
        },

        "coaches": [
            {
                "group": "246",
                "group_title": "HLV lớp 2-4-6",
                "group_subtitle": "Lớp buổi chiều",
                "name": "Nguyễn Hoàng Long",
                "role": "Phụ Trách HLV",
                "phone": "",
                "qualification": "",
                "photo_url": ""
            },
            {
                "group": "246",
                "group_title": "HLV lớp 2-4-6",
                "group_subtitle": "Lớp buổi chiều",
                "name": "Nguyễn Duy Thông",
                "role": "Trợ giảng",
                "phone": "",
                "qualification": "",
                "photo_url": ""
            },
            {
                "group": "357",
                "group_title": "HLV lớp 3-5-7",
                "group_subtitle": "Lớp chính trong tuần",
                "name": "Nông Thạch Khiêm",
                "role": "Phụ Trách HLV",
                "phone": "",
                "qualification": "",
                "photo_url": ""
            },
            {
                "group": "357",
                "group_title": "HLV lớp 3-5-7",
                "group_subtitle": "Lớp chính trong tuần",
                "name": "Nguyễn Trung Hiếu",
                "role": "Phụ Trách HLV",
                "phone": "",
                "qualification": "",
                "photo_url": ""
            },
            {
                "group": "357",
                "group_title": "HLV lớp 3-5-7",
                "group_subtitle": "Lớp chính trong tuần",
                "name": "Trần Ngọc Hà",
                "role": "Trợ giảng",
                "phone": "",
                "qualification": "",
                "photo_url": ""
            },
            {
                "group": "weekend",
                "group_title": "HLV lớp sáng & cuối tuần",
                "group_subtitle": "Thứ 7, Chủ nhật và lớp sáng",
                "name": "Nguyễn Lê Cường",
                "role": "Phụ Trách HLV",
                "phone": "",
                "qualification": "",
                "photo_url": ""
            }
        ],

        "regular_schedules": [
            {"title": "Lớp 2-4-6", "time": "18:00 - 19:30 và 19:30 - 21:00"},
            {"title": "Lớp 3-5-7", "time": "18:00 - 19:30 và 19:30 - 21:00"},
            {"title": "Lớp Thứ 7 và Chủ nhật", "time": "8:30 - 10:30"}
        ],

        "summer_schedules": [
            {"title": "Lớp hè", "time": "Từ đầu tháng 6 đến cuối tháng 8"},
            {"title": "Sáng 2-4-6 hoặc 3-5-7", "time": "8:00 - 9:30 và 9:30 - 11:00"}
        ],

        "fees_info": {
            "monthly_fee": "500.000đ",
            "three_month_discount": "Giảm 10%",
            "six_month_bonus": "Tặng 1 tháng miễn phí",
            "family_discount": "Giảm thêm 10% nếu đóng chung",
            "uniform_fee": "420.000đ / bộ",
            "exam_fee": "300.000đ / lần",
            "exam_notes": [
                "3 tháng CLB sẽ tổ chức thi nâng cấp đai 1 lần.",
                "Võ sinh cần đảm bảo thời gian tập luyện và chuyên môn, được HLV phụ trách xác nhận trước khi thi.",
                "Võ sinh mới thi lần đầu đóng thêm 20.000đ một lần duy nhất để Liên đoàn Taekwondo Việt Nam cấp mã, phục vụ hoạt động quản lý chung và đảm bảo quyền lợi trong quá trình tập luyện."
            ]
        }
    },

    "profile": {
        "owner": "Nguyễn Thiên Phụng",
        "phone_zalo": "0989 03 04 93",
        "email": "nhoctotokute93@gmail.com",
        "role": "VĐV Taekwondo Poomsae Việt Nam",
        "system_name": "Hệ thống quản lý hội viên CLB Taekwondo"
    },
}


def merge_settings(defaults, saved):
    result = {}

    for key, value in defaults.items():
        if isinstance(value, dict):
            result[key] = merge_settings(value, (saved or {}).get(key, {}))
        else:
            result[key] = (saved or {}).get(key, value)

    return result


def load_app_settings():
    """
    Đọc setup hệ thống từ Supabase.
    Nếu Supabase chưa có dữ liệu thì tự tạo từ DEFAULT_APP_SETTINGS.
    """
    try:
        rows = supabase.table(APP_SETTINGS_TABLE) \
            .select("key,value") \
            .execute().data or []

        if not rows:
            save_app_settings(DEFAULT_APP_SETTINGS)
            return DEFAULT_APP_SETTINGS.copy()

        saved = {}

        for row in rows:
            key = str(row.get("key") or "").strip()
            value = row.get("value") or {}

            if key:
                saved[key] = value

        return merge_settings(DEFAULT_APP_SETTINGS, saved)

    except Exception as e:
        print("[LOAD APP SETTINGS SUPABASE ERROR]", e)

        # Dự phòng nếu Supabase lỗi thì vẫn cho app chạy bằng mặc định
        return DEFAULT_APP_SETTINGS.copy()


def save_app_settings(settings):
    """
    Lưu setup hệ thống lên Supabase.
    Mỗi nhóm setup là 1 dòng:
    header, fees, exam, class_options, club_info, profile...
    """
    try:
        now_iso = datetime.now(timezone.utc).isoformat()

        for key, value in settings.items():
            payload = {
                "key": key,
                "value": value,
                "updated_at": now_iso,
            }

            existing = supabase.table(APP_SETTINGS_TABLE) \
                .select("key") \
                .eq("key", key) \
                .limit(1) \
                .execute().data or []

            if existing:
                supabase.table(APP_SETTINGS_TABLE) \
                    .update(payload) \
                    .eq("key", key) \
                    .execute()
            else:
                supabase.table(APP_SETTINGS_TABLE) \
                    .insert(payload) \
                    .execute()

    except Exception as e:
        print("[SAVE APP SETTINGS SUPABASE ERROR]", e)
        raise e


def money_to_int_web(v):
    return int(str(v or "0")
        .replace("đ", "")
        .replace("Đ", "")
        .replace(".", "")
        .replace(",", "")
        .replace(" ", "")
        .strip() or 0)


def get_app_setting(path, default=None):
    data = load_app_settings()
    cur = data

    for part in str(path).split("."):
        if not isinstance(cur, dict):
            return default
        cur = cur.get(part)

    return default if cur is None else cur


def get_app_setting_int(path, default=0):
    try:
        return int(get_app_setting(path, default) or default)
    except:
        return default

def clean_transfer_note_no_accent(text):
    text = str(text or "").strip()

    # Bỏ dấu tiếng Việt
    text = remove_accents(text)

    # Chỉ giữ ký tự an toàn cho nội dung chuyển khoản
    text = re.sub(r"[^A-Za-z0-9\s_\-\.]", "", text)

    # Gom khoảng trắng
    text = re.sub(r"\s+", " ", text).strip()

    return text


def build_transfer_note(template, student):
    template = str(template or "").strip()

    if not template:
        template = "{student_name}"

    student = student or {}

    note = (
        template
        .replace("{student_name}", str(student.get("name") or "").strip())
        .replace("{student_license}", str(student.get("license") or "").strip())
        .replace("{ma_hv}", str(student.get("license") or "").strip())
    ).strip()

    return clean_transfer_note_no_accent(note)


def build_vietqr_url(payment_info, student=None, amount=None):
    payment_info = payment_info or {}

    bank_code = str(payment_info.get("bank_code") or "").strip()
    account_number = str(payment_info.get("account_number") or "").strip()
    account_name = str(payment_info.get("account_name") or "").strip()
    note = build_transfer_note(payment_info.get("transfer_note"), student or {})

    if not bank_code or not account_number:
        return ""

    params = []

    if amount:
        try:
            amount_int = int(amount or 0)
            if amount_int > 0:
                params.append(f"amount={amount_int}")
        except:
            pass

    if note:
        params.append(f"addInfo={note}")

    if account_name:
        params.append(f"accountName={account_name}")

    query = "&".join(params)

    if query:
        return f"https://img.vietqr.io/image/{bank_code}-{account_number}-compact2.png?{query}"

    return f"https://img.vietqr.io/image/{bank_code}-{account_number}-compact2.png"

# =========================
# ADMIN LOGIN - BỘ QUẢN LÝ
# =========================
ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "nhoctotokute93")
ADMIN_PASSWORD_HASH = generate_password_hash(
    os.environ.get("ADMIN_PASSWORD", "Nguyenthienphung#93")
)

ADMIN_PUBLIC_ENDPOINTS = {
    "admin_login",
    "static",
    "index",

    # Các route học viên để không bị khóa bởi admin login
    "student_login",
    "student_logout",
    "student_portal_home",
    "student_portal_notifications",
    "student_portal_notification_detail",
    "student_portal_info",
    "student_portal_club_info",
    "student_portal_fees",
    "student_portal_exams",
    "student_portal_activities",
    "student_portal_settings",

    "coach_login",
    "coach_logout",
    "coach_exam",
    "coach_dan",
    "coach_exam_update_status",
}

ADMIN_PUBLIC_PATH_PREFIXES = (
    "/static/",
    "/student-login",
    "/student-logout",
    "/student-portal",
    "/coach-login",
    "/coach-logout",
    "/coach/",
)

def get_default_payment_settings():
    return {
        "id": "club_payment",
        "account_name": "",
        "account_number": "",
        "bank_code": "ACB",
        "bank_name": "ACB",
        "transfer_note": "{student_name}",
    }


def get_payment_settings():
    try:
        rows = supabase.table(PAYMENT_SETTINGS_TABLE) \
            .select("*") \
            .eq("id", "club_payment") \
            .limit(1) \
            .execute().data or []

        if rows:
            data = rows[0]
        else:
            data = get_default_payment_settings()

            supabase.table(PAYMENT_SETTINGS_TABLE) \
                .insert(data) \
                .execute()

        default_data = get_default_payment_settings()

        for key, value in default_data.items():
            if not data.get(key):
                data[key] = value

        return data

    except Exception as e:
        print("[GET PAYMENT SETTINGS ERROR]", e)
        return get_default_payment_settings()


def save_payment_settings(form):
    bank_code = str(form.get("bank_code") or "").strip()

    bank_map = {
        "VCB": "Vietcombank",
        "TCB": "Techcombank",
        "MB": "MB Bank",
        "ACB": "ACB",
        "BIDV": "BIDV",
        "VTB": "VietinBank",
        "VPB": "VPBank",
        "TPB": "TPBank",
        "VIB": "VIB",
        "OCB": "OCB",
        "STB": "Sacombank",
        "HDB": "HDBank",
        "SHB": "SHB",
        "EIB": "Eximbank",
        "MSB": "MSB",
        "BAB": "Bac A Bank",
        "SEAB": "SeABank",
        "LPB": "LPBank",
        "VAB": "VietABank",
        "ABB": "ABBank",
        "NAB": "Nam A Bank",
        "PGB": "PGBank",
        "PVCB": "PVcomBank",
    }

    payload = {
        "id": "club_payment",
        "account_name": str(form.get("account_name") or "").strip(),
        "account_number": str(form.get("account_number") or "").strip(),
        "bank_code": bank_code,
        "bank_name": bank_map.get(bank_code, bank_code),
        "transfer_note": str(form.get("transfer_note") or "{student_name}").strip() or "{student_name}",
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }

    existing = supabase.table(PAYMENT_SETTINGS_TABLE) \
        .select("id") \
        .eq("id", "club_payment") \
        .limit(1) \
        .execute().data or []

    if existing:
        supabase.table(PAYMENT_SETTINGS_TABLE) \
            .update(payload) \
            .eq("id", "club_payment") \
            .execute()
    else:
        supabase.table(PAYMENT_SETTINGS_TABLE) \
            .insert(payload) \
            .execute()

    return payload


def is_admin_logged_in():
    return session.get("admin_logged_in") is True


def is_safe_next_url(next_url):
    next_url = str(next_url or "").strip()
    return next_url.startswith("/") and not next_url.startswith("//")

def back_to_current_page(default_endpoint):
    """
    Quay lại đúng trang vừa thao tác.
    Dùng cho xóa/sửa để không bị nhảy về trang 1.
    """
    next_url = request.form.get("next") or request.referrer or url_for(default_endpoint)
    next_url = str(next_url or "").strip()

    # Chỉ cho redirect nội bộ, tránh redirect ra link ngoài
    if next_url.startswith(request.host_url):
        next_url = next_url.replace(request.host_url.rstrip("/"), "", 1)

    if not is_safe_next_url(next_url):
        next_url = url_for(default_endpoint)

    return redirect(next_url)

@app.before_request
def require_admin_login_for_management_pages():
    path = request.path or "/"

    # Cho phép các endpoint public
    if request.endpoint in ADMIN_PUBLIC_ENDPOINTS:
        return None

    # Cho phép static và bộ học viên
    if any(path.startswith(prefix) for prefix in ADMIN_PUBLIC_PATH_PREFIXES):
        return None

    # Cho phép trang đăng nhập admin
    if path == "/admin-login":
        return None

    # Nếu admin đã đăng nhập thì cho qua
    if is_admin_logged_in():
        return None

    # Chưa đăng nhập thì chuyển về login admin
    return redirect(url_for("admin_login", next=path))


@app.route("/admin-login", methods=["GET", "POST"])
def admin_login():
    if is_admin_logged_in():
        next_url = request.args.get("next") or url_for("students")

        if not is_safe_next_url(next_url):
            next_url = url_for("students")

        return redirect(next_url)

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        remember = request.form.get("remember") == "on"
        next_url = request.form.get("next") or url_for("students")

        if not is_safe_next_url(next_url):
            next_url = url_for("students")

        if username == ADMIN_USERNAME and check_password_hash(ADMIN_PASSWORD_HASH, password):
            # Không dùng session.clear(), vì sẽ làm out Student và Coach
            session["admin_logged_in"] = True
            session["admin_username"] = username
            session.permanent = remember

            flash("Đăng nhập admin thành công", "success")
            return redirect(next_url)

        flash("ID hoặc mật khẩu admin chưa đúng")
        return redirect(url_for("admin_login", next=next_url))

    next_url = request.args.get("next") or url_for("students")

    if not is_safe_next_url(next_url):
        next_url = url_for("students")

    return render_template("admin_login.html", next_url=next_url)


@app.get("/admin-logout")
def admin_logout():
    session.pop("admin_logged_in", None)
    session.pop("admin_username", None)

    session.pop("restore_unlocked", None)
    session.pop("restore_unlocked_until", None)

    flash("Đã đăng xuất admin")
    return redirect(url_for("admin_login"))

# =========================================================
# BACKUP & RESTORE CENTER - GIAI ĐOẠN 1
# - Mật khẩu riêng cho khu vực phục hồi
# - Khóa 15 phút sau 5 lần sai
# - Phiên xác thực có hiệu lực 10 phút
# - Hiển thị 3 bản backup mới nhất từ Google Drive
# =========================================================

BACKUP_REMOTE_DIR = os.environ.get(
    "BACKUP_REMOTE_DIR",
    "gdrive_new:VPS-Backup-PhungTKD"
)

BACKUP_FILE_PATTERN = os.environ.get(
    "BACKUP_FILE_PATTERN",
    "phungtkdsystem_full_*.tar.gz"
)

RESTORE_AUTH_MINUTES = 10
RESTORE_MAX_FAILED_ATTEMPTS = 5
RESTORE_LOCK_MINUTES = 15


def parse_iso_datetime_web(value):
    """
    Chuyển chuỗi thời gian Supabase/rclone thành datetime có timezone.
    """
    raw = str(value or "").strip()

    if not raw:
        return None

    try:
        return datetime.fromisoformat(
            raw.replace("Z", "+00:00")
        )
    except Exception:
        return None


def datetime_to_vietnam_web(value):
    """
    Chuyển datetime về giờ Việt Nam.
    """
    dt = value

    if isinstance(value, str):
        dt = parse_iso_datetime_web(value)

    if not dt:
        return None

    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)

    vietnam_tz = timezone(timedelta(hours=7))
    return dt.astimezone(vietnam_tz)


def format_file_size_web(size_bytes):
    """
    Đổi byte thành KB, MB, GB.
    """
    try:
        size = int(size_bytes or 0)
    except Exception:
        size = 0

    units = ["B", "KB", "MB", "GB", "TB"]
    value = float(size)
    unit_index = 0

    while value >= 1024 and unit_index < len(units) - 1:
        value /= 1024
        unit_index += 1

    if unit_index == 0:
        return f"{int(value)} {units[unit_index]}"

    return f"{value:.2f} {units[unit_index]}"

def sha256_file(file_path):
    """
    Tính mã SHA-256 của một file.
    Dùng để kiểm tra chính xác file backup
    có khác file hiện tại hay không.
    """
    file_path = Path(file_path)

    if not file_path.is_file():
        raise FileNotFoundError(
            f"Không tìm thấy file để tính checksum: {file_path}"
        )

    digest = hashlib.sha256()

    with file_path.open("rb") as file:
        for chunk in iter(
            lambda: file.read(1024 * 1024),
            b"",
        ):
            digest.update(chunk)

    return digest.hexdigest()

def get_restore_security_row():
    """
    Lấy dòng cấu hình bảo mật phục hồi.
    Hệ thống chỉ sử dụng dòng đầu tiên.
    """
    try:
        rows = (
            supabase.table(RESTORE_SECURITY_TABLE)
            .select("*")
            .order("id")
            .limit(1)
            .execute()
            .data
            or []
        )

        return rows[0] if rows else None

    except Exception as e:
        print("[GET RESTORE SECURITY ERROR]", repr(e))
        return None


def restore_password_has_been_created():
    row = get_restore_security_row()

    return bool(
        row
        and str(row.get("password_hash") or "").strip()
    )


def is_restore_area_unlocked():
    """
    Kiểm tra phiên xác thực khu vực phục hồi còn hiệu lực hay không.
    """
    if session.get("restore_unlocked") is not True:
        return False

    expires_raw = session.get("restore_unlocked_until")
    expires_at = parse_iso_datetime_web(expires_raw)

    if not expires_at:
        session.pop("restore_unlocked", None)
        session.pop("restore_unlocked_until", None)
        return False

    now_utc = datetime.now(timezone.utc)

    if expires_at <= now_utc:
        session.pop("restore_unlocked", None)
        session.pop("restore_unlocked_until", None)
        return False

    return True


def unlock_restore_area():
    expires_at = (
        datetime.now(timezone.utc)
        + timedelta(minutes=RESTORE_AUTH_MINUTES)
    )

    session["restore_unlocked"] = True
    session["restore_unlocked_until"] = expires_at.isoformat()


def lock_restore_area():
    session.pop("restore_unlocked", None)
    session.pop("restore_unlocked_until", None)


def get_restore_lock_status(security_row):
    """
    Trả về:
    {
        "locked": True/False,
        "locked_until": datetime hoặc None,
        "remaining_seconds": số giây
    }
    """
    locked_until = parse_iso_datetime_web(
        (security_row or {}).get("locked_until")
    )

    if not locked_until:
        return {
            "locked": False,
            "locked_until": None,
            "remaining_seconds": 0,
        }

    now_utc = datetime.now(timezone.utc)

    if locked_until <= now_utc:
        return {
            "locked": False,
            "locked_until": None,
            "remaining_seconds": 0,
        }

    remaining_seconds = int(
        (locked_until - now_utc).total_seconds()
    )

    return {
        "locked": True,
        "locked_until": locked_until,
        "remaining_seconds": max(remaining_seconds, 0),
    }


def reset_restore_failed_attempts(security_id):
    try:
        (
            supabase.table(RESTORE_SECURITY_TABLE)
            .update({
                "failed_attempts": 0,
                "locked_until": None,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            })
            .eq("id", security_id)
            .execute()
        )
    except Exception as e:
        print("[RESET RESTORE ATTEMPTS ERROR]", repr(e))


def register_restore_failed_attempt(security_row):
    """
    Tăng số lần sai.
    Đủ 5 lần sẽ khóa 15 phút.
    """
    security_id = security_row.get("id")

    try:
        failed_attempts = int(
            security_row.get("failed_attempts") or 0
        )
    except Exception:
        failed_attempts = 0

    failed_attempts += 1
    locked_until = None

    if failed_attempts >= RESTORE_MAX_FAILED_ATTEMPTS:
        locked_until = (
            datetime.now(timezone.utc)
            + timedelta(minutes=RESTORE_LOCK_MINUTES)
        ).isoformat()

    payload = {
        "failed_attempts": failed_attempts,
        "locked_until": locked_until,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }

    try:
        (
            supabase.table(RESTORE_SECURITY_TABLE)
            .update(payload)
            .eq("id", security_id)
            .execute()
        )
    except Exception as e:
        print("[REGISTER RESTORE FAILURE ERROR]", repr(e))

    return {
        "failed_attempts": failed_attempts,
        "locked_until": locked_until,
    }


def log_restore_history(
    action_type,
    status,
    backup_file=None,
    selected_items=None,
    restore_mode=None,
    error_message=None,
):
    """
    Ghi lịch sử truy cập/thao tác khu vực phục hồi.
    """
    payload = {
        "backup_file": backup_file,
        "action_type": str(action_type or "").strip(),
        "selected_items": selected_items or {},
        "restore_mode": restore_mode,
        "status": str(status or "pending").strip(),
        "error_message": error_message,
        "admin_name": session.get(
            "admin_username",
            ADMIN_USERNAME
        ),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }

    if status in ["success", "failed", "blocked"]:
        payload["completed_at"] = (
            datetime.now(timezone.utc).isoformat()
        )

    try:
        (
            supabase.table(RESTORE_HISTORY_TABLE)
            .insert(payload)
            .execute()
        )
    except Exception as e:
        print("[RESTORE HISTORY LOG ERROR]", repr(e))


def list_google_drive_backups(limit=3):
    """
    Liệt kê backup bằng rclone lsjson.

    Chỉ đọc file có tên:
    phungtkdsystem_full_*.tar.gz
    """
    command = [
        "rclone",
        "lsjson",
        BACKUP_REMOTE_DIR,
        "--files-only",
        "--include",
        BACKUP_FILE_PATTERN,
    ]

    try:
        result = subprocess.run(
            command,
            capture_output=True,
            text=True,
            timeout=60,
            check=False,
        )

        if result.returncode != 0:
            error_text = (
                result.stderr
                or result.stdout
                or "Không đọc được Google Drive."
            ).strip()

            raise RuntimeError(error_text)

        items = json.loads(result.stdout or "[]")

        backups = []

        for item in items:
            file_name = str(
                item.get("Name")
                or item.get("Path")
                or ""
            ).strip()

            if not file_name:
                continue

            mod_time_raw = str(
                item.get("ModTime") or ""
            ).strip()

            mod_time = parse_iso_datetime_web(mod_time_raw)
            mod_time_vn = datetime_to_vietnam_web(mod_time)

            size_bytes = int(item.get("Size") or 0)

            backups.append({
                "name": file_name,
                "path": str(
                    item.get("Path") or file_name
                ).strip(),
                "size_bytes": size_bytes,
                "size_text": format_file_size_web(size_bytes),
                "mod_time_raw": mod_time_raw,
                "mod_time": mod_time,
                "mod_time_vn": mod_time_vn,
                "mod_time_text": (
                    mod_time_vn.strftime("%d/%m/%Y %H:%M:%S")
                    if mod_time_vn
                    else "Không xác định"
                ),
                "mime_type": str(
                    item.get("MimeType")
                    or "application/gzip"
                ),
            })

        backups.sort(
            key=lambda x: (
                x.get("mod_time")
                or datetime.min.replace(tzinfo=timezone.utc)
            ),
            reverse=True,
        )

        return backups[:max(int(limit or 3), 1)], ""

    except subprocess.TimeoutExpired:
        return [], "Google Drive phản hồi quá lâu. Ken thử tải lại trang."

    except FileNotFoundError:
        return [], "VPS chưa tìm thấy lệnh rclone."

    except Exception as e:
        print("[LIST GOOGLE DRIVE BACKUPS ERROR]", repr(e))
        return [], str(e)


@app.route(
    "/backup-restore/security",
    methods=["GET", "POST"]
)
def backup_restore_security():
    """
    Trang tạo mật khẩu lần đầu hoặc nhập mật khẩu để mở khóa.
    """
    if not is_admin_logged_in():
        return redirect(
            url_for(
                "admin_login",
                next=url_for("backup_restore_security")
            )
        )

    security_row = get_restore_security_row()
    password_created = bool(
        security_row
        and str(security_row.get("password_hash") or "").strip()
    )

    # Nếu đã mở khóa thì đi thẳng vào trung tâm backup.
    if password_created and is_restore_area_unlocked():
        return redirect(url_for("backup_restore"))

    lock_status = get_restore_lock_status(security_row)

    if request.method == "POST":
        action = str(
            request.form.get("action") or ""
        ).strip()

        # =========================================
        # TẠO MẬT KHẨU LẦN ĐẦU
        # =========================================
        if not password_created:
            password = request.form.get("password", "")
            confirm_password = request.form.get(
                "confirm_password",
                ""
            )

            if len(password) < 8:
                flash(
                    "Mật khẩu phục hồi phải có ít nhất 8 ký tự.",
                    "danger"
                )
                return redirect(
                    url_for("backup_restore_security")
                )

            if password != confirm_password:
                flash(
                    "Hai lần nhập mật khẩu chưa giống nhau.",
                    "danger"
                )
                return redirect(
                    url_for("backup_restore_security")
                )

            payload = {
                "password_hash": generate_password_hash(
                    password
                ),
                "failed_attempts": 0,
                "locked_until": None,
                "updated_at": datetime.now(
                    timezone.utc
                ).isoformat(),
            }

            try:
                if security_row:
                    (
                        supabase.table(
                            RESTORE_SECURITY_TABLE
                        )
                        .update(payload)
                        .eq("id", security_row.get("id"))
                        .execute()
                    )
                else:
                    (
                        supabase.table(
                            RESTORE_SECURITY_TABLE
                        )
                        .insert(payload)
                        .execute()
                    )

                unlock_restore_area()

                log_restore_history(
                    action_type="create_restore_password",
                    status="success",
                )

                flash(
                    "Đã tạo mật khẩu bảo mật phục hồi.",
                    "success"
                )

                return redirect(
                    url_for("backup_restore")
                )

            except Exception as e:
                print(
                    "[CREATE RESTORE PASSWORD ERROR]",
                    repr(e)
                )

                flash(
                    f"Không tạo được mật khẩu phục hồi: {e}",
                    "danger"
                )

                return redirect(
                    url_for("backup_restore_security")
                )

        # =========================================
        # KIỂM TRA ĐANG BỊ KHÓA
        # =========================================
        if lock_status["locked"]:
            log_restore_history(
                action_type="unlock_restore_area",
                status="blocked",
                error_message="Khu vực phục hồi đang bị khóa.",
            )

            flash(
                "Khu vực phục hồi đang bị khóa tạm thời.",
                "danger"
            )

            return redirect(
                url_for("backup_restore_security")
            )

        # =========================================
        # ĐĂNG NHẬP KHU VỰC PHỤC HỒI
        # =========================================
        password = request.form.get("password", "")
        password_hash = str(
            security_row.get("password_hash") or ""
        )

        if check_password_hash(password_hash, password):
            reset_restore_failed_attempts(
                security_row.get("id")
            )

            unlock_restore_area()

            log_restore_history(
                action_type="unlock_restore_area",
                status="success",
            )

            flash(
                "Đã mở khóa khu vực sao lưu và phục hồi.",
                "success"
            )

            next_url = request.form.get("next") or url_for(
                "backup_restore"
            )

            if not is_safe_next_url(next_url):
                next_url = url_for("backup_restore")

            return redirect(next_url)

        failure_result = register_restore_failed_attempt(
            security_row
        )

        remaining = max(
            RESTORE_MAX_FAILED_ATTEMPTS
            - failure_result["failed_attempts"],
            0,
        )

        log_restore_history(
            action_type="unlock_restore_area",
            status="failed",
            error_message="Nhập sai mật khẩu phục hồi.",
        )

        if failure_result["locked_until"]:
            flash(
                "Ken đã nhập sai 5 lần. "
                "Khu vực phục hồi bị khóa 15 phút.",
                "danger"
            )
        else:
            flash(
                f"Mật khẩu chưa đúng. "
                f"Còn {remaining} lần thử.",
                "danger"
            )

        return redirect(
            url_for("backup_restore_security")
        )

    locked_until_vn = datetime_to_vietnam_web(
        lock_status.get("locked_until")
    )

    return render_template(
        "backup_restore_login.html",
        password_created=password_created,
        restore_locked=lock_status["locked"],
        remaining_seconds=lock_status["remaining_seconds"],
        locked_until_text=(
            locked_until_vn.strftime("%d/%m/%Y %H:%M:%S")
            if locked_until_vn
            else ""
        ),
        next_url=request.args.get("next")
        or url_for("backup_restore"),
    )


@app.get("/backup-restore")
def backup_restore():
    """
    Trang danh sách 3 bản backup mới nhất.
    Giai đoạn 1 chưa thực hiện phục hồi.
    """
    if not restore_password_has_been_created():
        return redirect(
            url_for("backup_restore_security")
        )

    if not is_restore_area_unlocked():
        return redirect(
            url_for(
                "backup_restore_security",
                next=request.path
            )
        )

    backups, backup_error = list_google_drive_backups(
        limit=3
    )

    unlocked_until = parse_iso_datetime_web(
        session.get("restore_unlocked_until")
    )

    unlocked_until_vn = datetime_to_vietnam_web(
        unlocked_until
    )

    try:
        history_rows = (
            supabase.table(RESTORE_HISTORY_TABLE)
            .select(
                "id,action_type,status,admin_name,"
                "created_at,error_message"
            )
            .order("created_at", desc=True)
            .limit(10)
            .execute()
            .data
            or []
        )
    except Exception as e:
        print("[GET RESTORE HISTORY ERROR]", repr(e))
        history_rows = []

    return render_template(
        "backup_restore.html",
        backups=backups,
        backup_error=backup_error,
        backup_remote_dir=BACKUP_REMOTE_DIR,
        restore_session_expires=(
            unlocked_until_vn.strftime("%H:%M:%S")
            if unlocked_until_vn
            else ""
        ),
        history_rows=history_rows,
    )


@app.post("/backup-restore/lock")
def backup_restore_lock():
    """
    Khóa khu vực phục hồi thủ công.
    """
    lock_restore_area()

    log_restore_history(
        action_type="lock_restore_area",
        status="success",
    )

    flash(
        "Đã khóa khu vực sao lưu và phục hồi.",
        "success"
    )

    return redirect(
        url_for("backup_restore_security")
    )


@app.get("/backup-restore/refresh")
def backup_restore_refresh():
    """
    Tải lại danh sách backup nhưng không kéo dài thời hạn phiên.
    """
    if not is_restore_area_unlocked():
        return redirect(
            url_for(
                "backup_restore_security",
                next=url_for("backup_restore")
            )
        )

    return redirect(url_for("backup_restore"))

# =========================================================
# BACKUP & RESTORE - GIAI ĐOẠN 3
# PHẦN 1: ĐỌC MANIFEST VÀ SO SÁNH BẢNG STUDENT
# =========================================================

BACKUP_LOCAL_TEST_DIR = Path(
    os.environ.get(
        "BACKUP_LOCAL_TEST_DIR",
        str(BASE_DIR / "backup_manifest_supabase_test")
    )
)

# =========================================================
# CẤU HÌNH SO SÁNH TỪNG BẢNG
# =========================================================

BACKUP_TABLE_COMPARE_CONFIGS = {
    "student": {
        "key_columns": ["license"],
        "display_columns": ["name", "license"],
        "ignored_fields": [
            "created_at",
            "updated_at",
        ],
    },

    "hocphi": {
        "key_columns": ["id"],
        "display_columns": [
            "ho_ten",
            "ma_hv",
            "thang_dong_phi",
        ],
        "ignored_fields": [
            "updated_at",
        ],
    },

    "ketqua": {
        "key_columns": ["id"],
        "display_columns": [
            "ho_ten",
            "ma_hv",
            "ky_thi",
            "cap_dai_thi",
        ],
        "ignored_fields": [
            "updated_at",
        ],
    },

    "hoatdong": {
        "key_columns": ["id"],
        "display_columns": [
            "ho_ten",
            "ma_hv",
            "hoat_dong",
        ],
        "ignored_fields": [
            "updated_at",
        ],
    },

    "activity_events": {
        "key_columns": ["id"],
        "display_columns": [
            "event_name",
            "name",
            "title",
            "location",
        ],
        "ignored_fields": [
            "updated_at",
        ],
    },

    "notifications": {
        "key_columns": ["id"],
        "display_columns": [
            "title",
            "target_name",
            "target_license",
        ],
        "ignored_fields": [
            "updated_at",
        ],
    },

    "notification_reads": {
        "key_columns": ["id"],
        "display_columns": [
            "student_license",
            "notification_id",
        ],
        "ignored_fields": [
            "updated_at",
        ],
    },

    "coaches": {
        "key_columns": ["id"],
        "display_columns": [
            "name",
            "license",
            "role",
        ],
        "ignored_fields": [
            "updated_at",
        ],
    },

    "exam_infos": {
        # Nếu bảng exam_infos có id thì dùng id.
        # Nếu dữ liệu cũ không có id, hàm bên dưới sẽ dự phòng bằng ky_thi.
        "key_columns": ["id"],
        "fallback_key_columns": ["ky_thi"],
        "display_columns": [
            "ky_thi",
            "venue",
            "exam_date",
        ],
        "ignored_fields": [
            "created_at",
            "updated_at",
        ],
    },

    "app_settings": {
        "key_columns": ["key"],
        "display_columns": ["key"],
        "ignored_fields": [
            "updated_at",
        ],
    },

    "payment_settings": {
        "key_columns": ["id"],
        "display_columns": [
            "bank_name",
            "account_name",
            "account_number",
        ],
        "ignored_fields": [
            "updated_at",
        ],
    },
}

def normalize_compare_value_web(value):
    """
    Chuẩn hóa dữ liệu trước khi so sánh.

    Mục tiêu:
    - None và chuỗi rỗng được xem gần giống nhau.
    - Dict/list được chuẩn hóa ổn định.
    - Chuỗi được bỏ khoảng trắng đầu cuối.
    """
    if value is None:
        return ""

    if isinstance(value, bool):
        return value

    if isinstance(value, (int, float)):
        return value

    if isinstance(value, dict):
        return {
            str(key): normalize_compare_value_web(item)
            for key, item in sorted(
                value.items(),
                key=lambda pair: str(pair[0])
            )
        }

    if isinstance(value, list):
        return [
            normalize_compare_value_web(item)
            for item in value
        ]

    return str(value).strip()


def load_json_file_web(file_path):
    """
    Đọc file JSON an toàn.
    """
    file_path = Path(file_path)

    if not file_path.exists():
        raise FileNotFoundError(
            f"Không tìm thấy file: {file_path}"
        )

    with file_path.open(
        "r",
        encoding="utf-8"
    ) as json_file:
        return json.load(json_file)


def extract_rows_from_export_json_web(raw_data):
    """
    Chấp nhận nhiều cấu trúc JSON:

    1. Danh sách trực tiếp:
       [
           {...},
           {...}
       ]

    2. Có khóa rows:
       {
           "rows": [...]
       }

    3. Có khóa data:
       {
           "data": [...]
       }

    4. Có khóa records:
       {
           "records": [...]
       }
    """
    if isinstance(raw_data, list):
        return [
            row
            for row in raw_data
            if isinstance(row, dict)
        ]

    if isinstance(raw_data, dict):
        for key in [
            "rows",
            "data",
            "records",
            "items"
        ]:
            rows = raw_data.get(key)

            if isinstance(rows, list):
                return [
                    row
                    for row in rows
                    if isinstance(row, dict)
                ]

    return []


def safe_backup_relative_path_web(
    backup_root,
    relative_path
):
    """
    Chống đường dẫn kiểu:
    ../../file

    Chỉ cho phép đọc file nằm bên trong thư mục backup.
    """
    backup_root = Path(backup_root).resolve()
    target_path = (
        backup_root
        / str(relative_path or "").lstrip("/\\")
    ).resolve()

    try:
        target_path.relative_to(backup_root)
    except ValueError:
        raise ValueError(
            "Đường dẫn file backup không an toàn."
        )

    return target_path


def get_backup_manifest_by_id_web(manifest_id):
    try:
        rows = (
            supabase.table(BACKUP_MANIFEST_TABLE)
            .select("*")
            .eq("id", manifest_id)
            .limit(1)
            .execute()
            .data
            or []
        )

        return rows[0] if rows else {}

    except Exception as e:
        print(
            "[GET BACKUP MANIFEST ERROR]",
            repr(e)
        )
        return {}


def get_backup_manifest_items_web(manifest_id):
    try:
        return (
            supabase.table(
                BACKUP_MANIFEST_ITEM_TABLE
            )
            .select("*")
            .eq("manifest_id", manifest_id)
            .order("item_type")
            .order("item_name")
            .execute()
            .data
            or []
        )

    except Exception as e:
        print(
            "[GET BACKUP MANIFEST ITEMS ERROR]",
            repr(e)
        )
        return []


def get_latest_backup_manifests_web(limit=10):
    """
    Đọc danh sách manifest từ Supabase.

    Phần này hoạt động cả local và VPS.
    """
    try:
        return (
            supabase.table(BACKUP_MANIFEST_TABLE)
            .select("*")
            .order("created_at", desc=True)
            .limit(limit)
            .execute()
            .data
            or []
        )

    except Exception as e:
        print(
            "[GET LATEST BACKUP MANIFESTS ERROR]",
            repr(e)
        )
        return []


def get_manifest_table_item_web(
    manifest_id,
    table_name
):
    try:
        rows = (
            supabase.table(
                BACKUP_MANIFEST_ITEM_TABLE
            )
            .select("*")
            .eq("manifest_id", manifest_id)
            .eq("item_type", "table")
            .eq("item_name", table_name)
            .limit(1)
            .execute()
            .data
            or []
        )

        return rows[0] if rows else {}

    except Exception as e:
        print(
            "[GET MANIFEST TABLE ITEM ERROR]",
            repr(e)
        )
        return {}


def get_backup_local_root_web(manifest):
    """
    Giai đoạn đang làm trên local:

    - Manifest thử dùng thư mục:
      backup_manifest_supabase_test

    Sau này trên VPS hàm này sẽ nhận thư mục giải nén
    của file Google Drive.
    """
    backup_name = str(
        (manifest or {}).get("backup_name") or ""
    ).strip()

    custom_root = str(
        (manifest or {}).get("local_extract_path")
        or ""
    ).strip()

    if custom_root:
        return Path(custom_root)

    # Manifest test đang dùng trên local
    if "manifest_test" in backup_name:
        return BACKUP_LOCAL_TEST_DIR

    # Tạm thời vẫn dùng thư mục test nếu file tồn tại.
    if BACKUP_LOCAL_TEST_DIR.exists():
        return BACKUP_LOCAL_TEST_DIR

    return BACKUP_LOCAL_TEST_DIR


def fetch_all_supabase_rows_web(
    table_name,
    page_size=1000
):
    """
    Đọc toàn bộ dữ liệu một bảng theo từng trang.

    Không chỉ lấy 1.000 dòng đầu tiên.
    """
    rows = []
    start = 0

    while True:
        end = start + page_size - 1

        batch = (
            supabase.table(table_name)
            .select("*")
            .range(start, end)
            .execute()
            .data
            or []
        )

        rows.extend(batch)

        if len(batch) < page_size:
            break

        start += page_size

    return rows


def build_row_map_web(
    rows,
    primary_key
):
    """
    Đưa danh sách dữ liệu thành map theo khóa chính.

    Ví dụ student:
    {
        "V23-001": {...},
        "V24-002": {...}
    }
    """
    result = {}
    duplicate_keys = []

    for row in rows:
        if not isinstance(row, dict):
            continue

        key_value = str(
            row.get(primary_key) or ""
        ).strip()

        if not key_value:
            continue

        if key_value in result:
            duplicate_keys.append(key_value)

        result[key_value] = row

    return result, duplicate_keys


def compare_record_fields_web(
    backup_row,
    current_row,
    ignored_fields=None
):
    """
    So sánh từng cột của hai bản ghi.
    """
    ignored_fields = set(
        ignored_fields or []
    )

    all_fields = sorted(
        set(backup_row.keys())
        | set(current_row.keys())
    )

    changes = []

    for field in all_fields:
        if field in ignored_fields:
            continue

        backup_value = normalize_compare_value_web(
            backup_row.get(field)
        )

        current_value = normalize_compare_value_web(
            current_row.get(field)
        )

        if backup_value != current_value:
            changes.append({
                "field": field,
                "backup_value": backup_row.get(field),
                "current_value": current_row.get(field),
                "backup_text": json.dumps(
                    backup_row.get(field),
                    ensure_ascii=False,
                    default=str
                ) if isinstance(
                    backup_row.get(field),
                    (dict, list)
                ) else str(
                    backup_row.get(field)
                    if backup_row.get(field) is not None
                    else ""
                ),
                "current_text": json.dumps(
                    current_row.get(field),
                    ensure_ascii=False,
                    default=str
                ) if isinstance(
                    current_row.get(field),
                    (dict, list)
                ) else str(
                    current_row.get(field)
                    if current_row.get(field) is not None
                    else ""
                ),
            })

    return changes


def compare_student_backup_web(
    backup_rows,
    current_rows
):
    """
    So sánh bảng student theo khóa license.

    Nhóm kết quả:
    - add: Có trong backup nhưng hiện tại không có.
    - update: Có cả hai nhưng khác dữ liệu.
    - unchanged: Có cả hai và giống nhau.
    - current_only: Hiện tại có nhưng backup không có.
    """
    primary_key = "license"

    backup_map, backup_duplicates = (
        build_row_map_web(
            backup_rows,
            primary_key
        )
    )

    current_map, current_duplicates = (
        build_row_map_web(
            current_rows,
            primary_key
        )
    )

    add_rows = []
    update_rows = []
    unchanged_rows = []
    current_only_rows = []

    # Những trường không cần dùng để quyết định dữ liệu thay đổi.
    ignored_fields = {
        "updated_at",
        "created_at",
    }

    for license_code, backup_row in backup_map.items():
        current_row = current_map.get(
            license_code
        )

        if not current_row:
            add_rows.append({
                "license": license_code,
                "name": backup_row.get("name", ""),
                "backup_row": backup_row,
            })
            continue

        changes = compare_record_fields_web(
            backup_row,
            current_row,
            ignored_fields=ignored_fields
        )

        if changes:
            update_rows.append({
                "license": license_code,
                "name": (
                    backup_row.get("name")
                    or current_row.get("name")
                    or ""
                ),
                "changes": changes,
                "change_count": len(changes),
                "backup_row": backup_row,
                "current_row": current_row,
            })
        else:
            unchanged_rows.append({
                "license": license_code,
                "name": (
                    backup_row.get("name")
                    or current_row.get("name")
                    or ""
                ),
            })

    for license_code, current_row in current_map.items():
        if license_code not in backup_map:
            current_only_rows.append({
                "license": license_code,
                "name": current_row.get("name", ""),
                "current_row": current_row,
            })

    def student_sort_key(item):
        return (
            remove_accents(
                str(item.get("name") or "")
            ).lower(),
            str(item.get("license") or "")
        )

    add_rows.sort(key=student_sort_key)
    update_rows.sort(key=student_sort_key)
    unchanged_rows.sort(key=student_sort_key)
    current_only_rows.sort(key=student_sort_key)

    return {
        "primary_key": primary_key,
        "backup_count": len(backup_map),
        "current_count": len(current_map),
        "add_count": len(add_rows),
        "update_count": len(update_rows),
        "unchanged_count": len(unchanged_rows),
        "current_only_count": len(
            current_only_rows
        ),
        "add_rows": add_rows,
        "update_rows": update_rows,
        "unchanged_rows": unchanged_rows,
        "current_only_rows": current_only_rows,
        "backup_duplicates": backup_duplicates,
        "current_duplicates": current_duplicates,
    }


# =========================================================
# BACKUP & RESTORE - CẤU HÌNH PHỤC HỒI DÙNG CHUNG
# Thay toàn bộ phần Giai đoạn 4 cũ bằng khối này.
# =========================================================

BACKUP_RESTORE_TABLE_CONFIGS = {
    "student": {
        "title": "Hội viên",
        "row_label": "học viên",
        "add_title": "Có trong backup nhưng hiện tại chưa có",
        "update_title": "Thông tin học viên khác với hiện tại",
        "allow_restore": True,
        "protected_fields": {
            "license", "id", "created_at", "updated_at",
        },
        "field_labels": {
            "name": "Họ tên",
            "birthdate": "Ngày sinh",
            "gender": "Giới tính",
            "classroom": "Lớp",
            "timeclass": "Ca học",
            "clup": "Câu lạc bộ",
            "phonenumber": "Số điện thoại",
            "address": "Địa chỉ",
            "belt": "Cấp đai",
            "family": "Gia đình",
            "active": "Trạng thái",
        },
    },
    "hocphi": {
        "title": "Học phí",
        "row_label": "phiếu học phí",
        "add_title": "Phiếu học phí có trong backup nhưng hiện tại chưa có",
        "update_title": "Dữ liệu phiếu học phí khác với hiện tại",
        "allow_restore": True,
        "protected_fields": {"id", "created_at", "updated_at"},
        "field_labels": {
            "thoi_gian": "Thời gian",
            "ma_hv": "Mã hội viên",
            "ho_ten": "Họ tên",
            "tong_tien": "Tổng tiền",
            "thang_dong_phi": "Tháng đóng phí",
            "ghi_chu": "Ghi chú",
            "ma_thang": "Mã tháng",
            "ma_quy": "Mã quý",
            "chuyen_khoan": "Phương thức thanh toán",
        },
    },
    "ketqua": {
        "title": "Kết quả thi cấp đẳng",
        "row_label": "kết quả thi",
        "add_title": "Kết quả thi có trong backup nhưng hiện tại chưa có",
        "update_title": "Kết quả thi khác với hiện tại",
        "allow_restore": True,
        "protected_fields": {"id", "created_at", "updated_at"},
        "field_labels": {
            "ky_thi": "Kỳ thi",
            "ma_hv": "Mã hội viên",
            "ho_ten": "Họ tên",
            "cap_dai_thi": "Cấp đai thi",
            "so_thi": "Số thi",
            "ket_qua": "Kết quả",
            "ghi_chu": "Ghi chú",
        },
    },
    "hoatdong": {
        "title": "Hoạt động",
        "row_label": "hoạt động",
        "add_title": "Hoạt động có trong backup nhưng hiện tại chưa có",
        "update_title": "Dữ liệu hoạt động khác với hiện tại",
        "allow_restore": True,
        "protected_fields": {"id", "created_at", "updated_at"},
        "field_labels": {
            "ma_hv": "Mã hội viên",
            "ho_ten": "Họ tên",
            "hoat_dong": "Tên hoạt động",
            "thoi_gian": "Thời gian",
            "dia_diem": "Địa điểm",
            "noi_dung": "Nội dung",
            "ket_qua": "Kết quả",
        },
    },
    "activity_events": {
        "title": "Sự kiện hoạt động",
        "row_label": "sự kiện",
        "add_title": "Sự kiện có trong backup nhưng hiện tại chưa có",
        "update_title": "Thông tin sự kiện khác với hiện tại",
        "allow_restore": True,
        "protected_fields": {"id", "created_at", "updated_at"},
        "field_labels": {},
    },
    "notifications": {
        "title": "Thông báo",
        "row_label": "thông báo",
        "add_title": "Thông báo có trong backup nhưng hiện tại chưa có",
        "update_title": "Nội dung thông báo khác với hiện tại",
        "allow_restore": True,
        "protected_fields": {"id", "created_at", "updated_at"},
        "field_labels": {},
    },
    "notification_reads": {
        "title": "Lịch sử đọc thông báo",
        "row_label": "lượt đọc",
        "add_title": "Lượt đọc có trong backup nhưng hiện tại chưa có",
        "update_title": "Dữ liệu lượt đọc khác với hiện tại",
        "allow_restore": True,
        "protected_fields": {
            "id", "notification_id", "created_at", "updated_at",
        },
        "field_labels": {},
    },
    "coaches": {
        "title": "Huấn luyện viên",
        "row_label": "huấn luyện viên",
        "add_title": "Huấn luyện viên có trong backup nhưng hiện tại chưa có",
        "update_title": "Thông tin huấn luyện viên khác với hiện tại",
        "allow_restore": True,
        "protected_fields": {"id", "created_at", "updated_at"},
        "field_labels": {},
    },
    "exam_infos": {
        "title": "Thông tin kỳ thi",
        "row_label": "kỳ thi",
        "add_title": "Kỳ thi có trong backup nhưng hiện tại chưa có",
        "update_title": "Thông tin kỳ thi khác với hiện tại",
        "allow_restore": True,
        "protected_fields": {"id", "ky_thi", "created_at", "updated_at"},
        "field_labels": {},
    },
    "app_settings": {
        "title": "Cài đặt ứng dụng",
        "row_label": "nhóm cài đặt",
        "add_title": "Nhóm cài đặt có trong backup nhưng hiện tại chưa có",
        "update_title": "Cài đặt ứng dụng khác với hiện tại",
        "allow_restore": True,
        "protected_fields": {"key", "created_at", "updated_at"},
        "field_labels": {},
    },
    "payment_settings": {
        "title": "Cài đặt thanh toán",
        "row_label": "cấu hình thanh toán",
        "add_title": "Cấu hình có trong backup nhưng hiện tại chưa có",
        "update_title": "Cài đặt thanh toán khác với hiện tại",
        "allow_restore": True,
        "protected_fields": {"id", "created_at", "updated_at"},
        "field_labels": {
            "account_name": "Tên chủ tài khoản",
            "account_number": "Số tài khoản",
            "bank_code": "Mã ngân hàng",
            "bank_name": "Tên ngân hàng",
            "transfer_note": "Nội dung chuyển khoản",
        },
    },
}

BACKUP_RESTORE_BLOCKED_TABLES = {
    "backup_manifests",
    "backup_manifest_items",
    "restore_history",
    "restore_security",
}


def get_backup_restore_table_meta_web(table_name):
    table_name = str(table_name or "").strip()
    config = BACKUP_RESTORE_TABLE_CONFIGS.get(table_name)

    if not config:
        return {
            "title": table_name,
            "row_label": "dòng",
            "add_title": "Có trong backup nhưng hiện tại chưa có",
            "update_title": "Dữ liệu khác với hiện tại",
            "allow_restore": False,
            "protected_fields": set(),
            "field_labels": {},
        }

    result = dict(config)
    result["protected_fields"] = set(
        result.get("protected_fields") or set()
    )
    result["field_labels"] = dict(
        result.get("field_labels") or {}
    )
    return result


def load_generic_backup_compare_context_web(manifest_id, table_name):
    table_name = str(table_name or "").strip()

    if table_name not in BACKUP_RESTORE_TABLE_CONFIGS:
        raise ValueError(
            f"Bảng {table_name} chưa được phép phục hồi."
        )

    manifest = get_backup_manifest_by_id_web(manifest_id)

    if not manifest:
        raise ValueError("Không tìm thấy manifest backup.")

    table_item = get_manifest_table_item_web(
        manifest_id,
        table_name,
    )

    if not table_item:
        raise ValueError(
            f"Manifest này không có bảng {table_name}."
        )

    backup_root = get_backup_local_root_web(manifest)
    backup_path = str(
        table_item.get("backup_path")
        or f"database_exports/{table_name}.json"
    ).strip()

    json_path = safe_backup_relative_path_web(
        backup_root,
        backup_path,
    )

    raw_backup_data = load_json_file_web(json_path)
    backup_rows = extract_rows_from_export_json_web(
        raw_backup_data
    )
    current_rows = fetch_all_supabase_rows_web(
        table_name
    )
    compare_result = compare_generic_table_backup_web(
        table_name,
        backup_rows,
        current_rows,
    )

    return {
        "manifest": manifest,
        "table_item": table_item,
        "backup_path": backup_path,
        "backup_rows": backup_rows,
        "current_rows": current_rows,
        "compare_result": compare_result,
    }


def get_restore_key_filter_web(table_name, backup_row):
    compare_config = get_table_compare_config_web(
        table_name
    ) or {}

    key_columns = get_row_key_columns_web(
        backup_row,
        compare_config,
    )

    filters = {}

    for column in key_columns:
        value = backup_row.get(column)

        if value in [None, ""]:
            raise ValueError(
                f"Dòng backup thiếu khóa {column}."
            )

        filters[column] = value

    if not filters:
        raise ValueError(
            "Không xác định được khóa đối chiếu."
        )

    return filters


def apply_restore_filters_web(query, filters):
    for column, value in (filters or {}).items():
        query = query.eq(column, value)

    return query

def is_empty_restore_value_web(value):
    """
    Không cho dữ liệu trống trong backup ghi đè dữ liệu hiện tại.
    """

    if value is None:
        return True

    if isinstance(value, str):
        return not value.strip()

    if isinstance(value, (list, dict, tuple, set)):
        return len(value) == 0

    return False

def build_generic_restore_payload_web(
    table_name,
    backup_row,
    for_update=False,
):
    meta = get_backup_restore_table_meta_web(
        table_name
    )

    protected = set(
        meta.get("protected_fields") or set()
    )

    payload = {}

    for field, value in (backup_row or {}).items():
        if field in {"created_at", "updated_at"}:
            continue

        if for_update and field in protected:
            continue

        payload[field] = value

    return payload


@app.post(
    "/backup-restore/manifest/"
    "<manifest_id>/restore/"
    "<table_name>/add-missing"
)
def backup_restore_table_add_missing(
    manifest_id,
    table_name,
):
    if not is_restore_area_unlocked():
        return redirect(
            url_for(
                "backup_restore_security",
                next=request.path,
            )
        )

    table_name = str(table_name or "").strip()
    meta = get_backup_restore_table_meta_web(
        table_name
    )

    if (
        table_name in BACKUP_RESTORE_BLOCKED_TABLES
        or not meta.get("allow_restore")
    ):
        flash(
            f"Bảng {table_name} không được phép phục hồi.",
            "danger",
        )
        return redirect(
            url_for(
                "backup_restore_manifest_detail",
                manifest_id=manifest_id,
            )
        )

    selected_keys = []

    for value in request.form.getlist("selected_keys"):
        row_key = str(value or "").strip()
        if row_key and row_key not in selected_keys:
            selected_keys.append(row_key)

    if not selected_keys:
        flash("Ken chưa chọn dòng cần phục hồi.", "danger")
        return redirect(
            url_for(
                "backup_restore_compare_table",
                manifest_id=manifest_id,
                table_name=table_name,
            )
        )

    try:
        context = load_generic_backup_compare_context_web(
            manifest_id,
            table_name,
        )
        manifest = context["manifest"]
        compare_result = context["compare_result"]

        if (
            compare_result.get("duplicate_backup_keys")
            or compare_result.get("duplicate_current_keys")
        ):
            raise ValueError(
                "Khóa đối chiếu bị trùng nên không thể phục hồi."
            )

        add_map = {
            str(item.get("key") or "").strip():
            item.get("backup_row") or {}
            for item in compare_result.get("add_rows") or []
        }

        inserted = []
        skipped = []
        failed = []

        for row_key in selected_keys:
            backup_row = add_map.get(row_key)

            if not backup_row:
                skipped.append({
                    "key": row_key,
                    "reason": "Dòng đã tồn tại hoặc không còn thiếu.",
                })
                continue

            try:
                filters = get_restore_key_filter_web(
                    table_name,
                    backup_row,
                )

                existing_query = (
                    supabase.table(table_name)
                    .select("*")
                )
                existing_query = apply_restore_filters_web(
                    existing_query,
                    filters,
                )
                existing = (
                    existing_query.limit(1)
                    .execute().data
                    or []
                )

                if existing:
                    skipped.append({
                        "key": row_key,
                        "reason": "Dòng đã tồn tại trên Supabase.",
                    })
                    continue

                payload = build_generic_restore_payload_web(
                    table_name,
                    backup_row,
                    for_update=False,
                )

                supabase.table(table_name).insert(
                    payload
                ).execute()

                inserted.append({
                    "key": row_key,
                    "filters": filters,
                })

            except Exception as row_error:
                failed.append({
                    "key": row_key,
                    "error": str(row_error),
                })

        log_restore_history(
            action_type=f"restore_{table_name}_add_missing",
            status="success" if not failed else "failed",
            backup_file=manifest.get("backup_name"),
            selected_items={
                "manifest_id": manifest_id,
                "table": table_name,
                "requested_keys": selected_keys,
                "inserted": inserted,
                "skipped": skipped,
                "failed": failed,
                "safety": {
                    "updates": 0,
                    "deletes": 0,
                    "mode": "add_missing_only",
                },
            },
            restore_mode="add_missing_only",
            error_message=(
                f"Có {len(failed)} dòng lỗi."
                if failed else None
            ),
        )

        if inserted:
            flash(
                f"Đã thêm {len(inserted)} "
                f"{meta['row_label']} còn thiếu. "
                f"Không sửa và không xóa dữ liệu hiện tại.",
                "success",
            )

        if skipped:
            flash(
                f"Đã bỏ qua {len(skipped)} dòng không còn thiếu.",
                "warning",
            )

        if failed:
            flash(
                f"Có {len(failed)} dòng phục hồi lỗi.",
                "danger",
            )

        if not inserted and not failed:
            flash(
                "Không còn dữ liệu cần thêm tại thời điểm thực hiện.",
                "warning",
            )

    except Exception as e:
        print(
            "[GENERIC RESTORE ADD MISSING ERROR]",
            table_name,
            repr(e),
        )
        flash(
            f"Không thể phục hồi bảng {table_name}: {e}",
            "danger",
        )

    return redirect(
        url_for(
            "backup_restore_compare_table",
            manifest_id=manifest_id,
            table_name=table_name,
        )
    )


@app.post(
    "/backup-restore/manifest/"
    "<manifest_id>/restore/"
    "<table_name>/selected-fields"
)
def backup_restore_table_selected_fields(
    manifest_id,
    table_name,
):
    if not is_restore_area_unlocked():
        return redirect(
            url_for(
                "backup_restore_security",
                next=request.path,
            )
        )

    table_name = str(table_name or "").strip()
    meta = get_backup_restore_table_meta_web(
        table_name
    )

    if (
        table_name in BACKUP_RESTORE_BLOCKED_TABLES
        or not meta.get("allow_restore")
    ):
        flash(
            f"Bảng {table_name} không được phép phục hồi.",
            "danger",
        )
        return redirect(
            url_for(
                "backup_restore_manifest_detail",
                manifest_id=manifest_id,
            )
        )

    requested_pairs = []

    for raw in request.form.getlist("selected_fields"):
        raw = str(raw or "").strip()

        if "|||" not in raw:
            continue

        row_key, field = raw.split("|||", 1)
        pair = (row_key.strip(), field.strip())

        if (
            pair[0]
            and pair[1]
            and pair not in requested_pairs
        ):
            requested_pairs.append(pair)

    if not requested_pairs:
        flash(
            "Ken chưa chọn trường cần phục hồi.",
            "danger",
        )
        return redirect(
            url_for(
                "backup_restore_compare_table",
                manifest_id=manifest_id,
                table_name=table_name,
            )
        )

    protected_fields = set(
        meta.get("protected_fields") or set()
    )

    invalid_fields = sorted({
        field
        for _, field in requested_pairs
        if field in protected_fields
    })

    if invalid_fields:
        flash(
            "Không được phục hồi trường khóa/kỹ thuật: "
            + ", ".join(invalid_fields),
            "danger",
        )
        return redirect(
            url_for(
                "backup_restore_compare_table",
                manifest_id=manifest_id,
                table_name=table_name,
            )
        )

    try:
        context = load_generic_backup_compare_context_web(
            manifest_id,
            table_name,
        )
        manifest = context["manifest"]
        compare_result = context["compare_result"]

        if (
            compare_result.get("duplicate_backup_keys")
            or compare_result.get("duplicate_current_keys")
        ):
            raise ValueError(
                "Khóa đối chiếu bị trùng nên không thể phục hồi."
            )

        update_map = {
            str(item.get("key") or "").strip(): item
            for item in compare_result.get("update_rows") or []
        }

        grouped_fields = {}

        for row_key, field in requested_pairs:
            grouped_fields.setdefault(row_key, [])
            if field not in grouped_fields[row_key]:
                grouped_fields[row_key].append(field)

        updated = []
        skipped = []
        failed = []

        for row_key, fields in grouped_fields.items():
            item = update_map.get(row_key)

            if not item:
                skipped.append({
                    "key": row_key,
                    "reason": "Dòng không còn thuộc nhóm thay đổi.",
                })
                continue

            backup_row = item.get("backup_row") or {}
            current_row = item.get("current_row") or {}
            changed_fields = {
                str(change.get("field") or "").strip()
                for change in item.get("changes") or []
            }

            payload = {}
            change_log = []

            for field in fields:
                if field not in changed_fields:
                    skipped.append({
                        "key": row_key,
                        "field": field,
                        "reason": "Trường không còn khác dữ liệu.",
                    })
                    continue

                backup_value = backup_row.get(field)

                if is_empty_restore_value_web(backup_value):
                    skipped.append({
                        "key": row_key,
                        "field": field,
                        "reason": "Backup đang trống nên không ghi đè.",
                    })
                    continue

                payload[field] = backup_value
                change_log.append({
                    "field": field,
                    "old_value": current_row.get(field),
                    "new_value": backup_value,
                })

            if not payload:
                continue

            try:
                filters = get_restore_key_filter_web(
                    table_name,
                    backup_row,
                )

                update_query = (
                    supabase.table(table_name)
                    .update(payload)
                )
                update_query = apply_restore_filters_web(
                    update_query,
                    filters,
                )
                update_query.execute()

                if (
                    table_name == "student"
                    and set(payload.keys()) & {
                        "name", "birthdate", "gender",
                        "classroom", "timeclass", "clup",
                        "phonenumber", "address",
                    }
                ):
                    sync_student_profile_to_related_tables(
                        str(backup_row.get("license") or "").strip()
                    )

                updated.append({
                    "key": row_key,
                    "filters": filters,
                    "fields": change_log,
                })

            except Exception as row_error:
                failed.append({
                    "key": row_key,
                    "fields": list(payload.keys()),
                    "error": str(row_error),
                })

        updated_field_count = sum(
            len(item.get("fields") or [])
            for item in updated
        )

        log_restore_history(
            action_type=f"restore_{table_name}_selected_fields",
            status="success" if not failed else "failed",
            backup_file=manifest.get("backup_name"),
            selected_items={
                "manifest_id": manifest_id,
                "table": table_name,
                "requested": [
                    {"key": key, "field": field}
                    for key, field in requested_pairs
                ],
                "updated": updated,
                "skipped": skipped,
                "failed": failed,
                "safety": {
                    "inserts": 0,
                    "deletes": 0,
                    "blank_overwrite": False,
                    "mode": "selected_non_empty_fields_only",
                },
            },
            restore_mode="selected_non_empty_fields_only",
            error_message=(
                f"Có {len(failed)} dòng lỗi."
                if failed else None
            ),
        )

        if updated:
            flash(
                f"Đã phục hồi {updated_field_count} trường "
                f"cho {len(updated)} {meta['row_label']}.",
                "success",
            )

        if skipped:
            flash(
                f"Đã bỏ qua {len(skipped)} mục không còn phù hợp "
                f"hoặc backup đang trống.",
                "warning",
            )

        if failed:
            flash(
                f"Có {len(failed)} dòng cập nhật lỗi.",
                "danger",
            )

        if not updated and not failed:
            flash(
                "Không có trường nào được phục hồi.",
                "warning",
            )

    except Exception as e:
        print(
            "[GENERIC RESTORE SELECTED FIELDS ERROR]",
            table_name,
            repr(e),
        )
        flash(
            f"Không thể phục hồi trường bảng {table_name}: {e}",
            "danger",
        )

    return redirect(
        url_for(
            "backup_restore_compare_table",
            manifest_id=manifest_id,
            table_name=table_name,
        )
    )

# =========================================================
# SO SÁNH DÙNG CHUNG CHO TOÀN BỘ BẢNG
# =========================================================

def get_table_compare_config_web(table_name):
    table_name = str(table_name or "").strip()

    return BACKUP_TABLE_COMPARE_CONFIGS.get(
        table_name
    )


def stringify_compare_value_web(value):
    """
    Đổi dữ liệu thành chuỗi để hiển thị trên giao diện.
    """
    if value is None:
        return ""

    if isinstance(value, (dict, list)):
        try:
            return json.dumps(
                value,
                ensure_ascii=False,
                sort_keys=True,
                default=str,
            )
        except Exception:
            return str(value)

    return str(value)


def get_row_key_columns_web(row, config):
    """
    Lấy bộ cột khóa phù hợp cho một dòng.

    Ví dụ:
    - student dùng license.
    - app_settings dùng key.
    - exam_infos ưu tiên id, nếu id trống thì dùng ky_thi.
    """
    config = config or {}

    key_columns = list(
        config.get("key_columns") or []
    )

    primary_values = [
        str(row.get(column) or "").strip()
        for column in key_columns
    ]

    if (
        key_columns
        and all(primary_values)
    ):
        return key_columns

    fallback_columns = list(
        config.get("fallback_key_columns") or []
    )

    fallback_values = [
        str(row.get(column) or "").strip()
        for column in fallback_columns
    ]

    if (
        fallback_columns
        and all(fallback_values)
    ):
        return fallback_columns

    return key_columns


def build_generic_row_key_web(row, config):
    """
    Tạo khóa đối chiếu của một bản ghi.

    Với khóa nhiều cột:
    ma_hv + ky_thi
    sẽ thành:
    ma_hv||ky_thi
    """
    if not isinstance(row, dict):
        return ""

    key_columns = get_row_key_columns_web(
        row,
        config
    )

    if not key_columns:
        return ""

    values = []

    for column in key_columns:
        value = normalize_compare_value_web(
            row.get(column)
        )

        if isinstance(value, (dict, list)):
            value = json.dumps(
                value,
                ensure_ascii=False,
                sort_keys=True,
                default=str,
            )

        value = str(value).strip()

        if not value:
            return ""

        values.append(value)

    return " || ".join(values)


def build_generic_display_name_web(
    row,
    config
):
    """
    Lấy thông tin nhận diện dễ đọc cho một dòng.

    Ví dụ student:
    Nguyễn Văn A • V23-001
    """
    if not isinstance(row, dict):
        return ""

    display_columns = list(
        (config or {}).get(
            "display_columns"
        ) or []
    )

    values = []

    for column in display_columns:
        value = row.get(column)

        if value is None:
            continue

        value = str(value).strip()

        if not value:
            continue

        if value not in values:
            values.append(value)

    return " • ".join(values[:4])


def build_generic_row_map_web(
    rows,
    config
):
    """
    Chuyển danh sách bản ghi thành dict theo khóa đối chiếu.
    Đồng thời ghi nhận các khóa bị trùng.
    """
    row_map = {}
    duplicate_keys = []
    missing_key_rows = []

    for row_index, row in enumerate(
        rows or [],
        start=1
    ):
        if not isinstance(row, dict):
            continue

        row_key = build_generic_row_key_web(
            row,
            config
        )

        if not row_key:
            missing_key_rows.append({
                "row_index": row_index,
                "row": row,
            })
            continue

        if row_key in row_map:
            if row_key not in duplicate_keys:
                duplicate_keys.append(row_key)

        row_map[row_key] = row

    return {
        "map": row_map,
        "duplicate_keys": duplicate_keys,
        "missing_key_rows": missing_key_rows,
    }


def compare_generic_table_backup_web(
    table_name,
    backup_rows,
    current_rows,
):
    """
    So sánh một bảng bất kỳ theo cấu hình.

    Kết quả:
    - add_rows:
      Có trong backup nhưng không có ở hiện tại.

    - update_rows:
      Có ở cả hai nhưng khác dữ liệu.

    - unchanged_rows:
      Có ở cả hai và giống nhau.

    - current_only_rows:
      Có ở hiện tại nhưng không có trong backup.
    """
    config = get_table_compare_config_web(
        table_name
    )

    if not config:
        raise ValueError(
            f"Bảng {table_name} chưa khai báo "
            f"khóa đối chiếu."
        )

    backup_result = build_generic_row_map_web(
        backup_rows,
        config
    )

    current_result = build_generic_row_map_web(
        current_rows,
        config
    )

    backup_map = backup_result["map"]
    current_map = current_result["map"]

    ignored_fields = set(
        config.get("ignored_fields") or []
    )

    add_rows = []
    update_rows = []
    unchanged_rows = []
    current_only_rows = []

    # =========================================
    # CÓ TRONG BACKUP
    # =========================================
    for row_key, backup_row in backup_map.items():
        current_row = current_map.get(
            row_key
        )

        display_name = (
            build_generic_display_name_web(
                backup_row,
                config
            )
        )

        if not current_row:
            add_rows.append({
                "key": row_key,
                "display_name": display_name,
                "backup_row": backup_row,
            })
            continue

        changes = compare_record_fields_web(
            backup_row,
            current_row,
            ignored_fields=ignored_fields,
        )

        if changes:
            normalized_changes = []

            for change in changes:
                normalized_changes.append({
                    "field": change.get(
                        "field",
                        ""
                    ),
                    "backup_value": change.get(
                        "backup_value"
                    ),
                    "current_value": change.get(
                        "current_value"
                    ),
                    "backup_text": (
                        change.get("backup_text")
                        or stringify_compare_value_web(
                            change.get(
                                "backup_value"
                            )
                        )
                    ),
                    "current_text": (
                        change.get("current_text")
                        or stringify_compare_value_web(
                            change.get(
                                "current_value"
                            )
                        )
                    ),
                })

            update_rows.append({
                "key": row_key,
                "display_name": (
                    display_name
                    or build_generic_display_name_web(
                        current_row,
                        config
                    )
                ),
                "change_count": len(
                    normalized_changes
                ),
                "changes": normalized_changes,
                "backup_row": backup_row,
                "current_row": current_row,
            })

        else:
            unchanged_rows.append({
                "key": row_key,
                "display_name": (
                    display_name
                    or build_generic_display_name_web(
                        current_row,
                        config
                    )
                ),
            })

    # =========================================
    # CHỈ CÓ Ở HIỆN TẠI
    # =========================================
    for row_key, current_row in current_map.items():
        if row_key in backup_map:
            continue

        current_only_rows.append({
            "key": row_key,
            "display_name": (
                build_generic_display_name_web(
                    current_row,
                    config
                )
            ),
            "current_row": current_row,
        })

    def generic_sort_key(item):
        return (
            remove_accents(
                str(
                    item.get("display_name")
                    or ""
                )
            ).lower(),
            str(item.get("key") or ""),
        )

    add_rows.sort(key=generic_sort_key)
    update_rows.sort(key=generic_sort_key)
    unchanged_rows.sort(key=generic_sort_key)
    current_only_rows.sort(
        key=generic_sort_key
    )

    effective_key_columns = list(
        config.get("key_columns") or []
    )

    primary_key_text = " + ".join(
        effective_key_columns
    )

    fallback_columns = list(
        config.get(
            "fallback_key_columns"
        ) or []
    )

    if fallback_columns:
        primary_key_text += (
            " — dự phòng: "
            + " + ".join(fallback_columns)
        )

    return {
        "table_name": table_name,
        "primary_key": primary_key_text,

        "backup_count": len(backup_map),
        "current_count": len(current_map),

        "add_count": len(add_rows),
        "update_count": len(update_rows),
        "unchanged_count": len(
            unchanged_rows
        ),
        "current_only_count": len(
            current_only_rows
        ),

        "add_rows": add_rows,
        "update_rows": update_rows,
        "unchanged_rows": unchanged_rows,
        "current_only_rows": current_only_rows,

        "duplicate_backup_keys": (
            backup_result["duplicate_keys"]
        ),
        "duplicate_current_keys": (
            current_result["duplicate_keys"]
        ),

        "missing_backup_key_count": len(
            backup_result["missing_key_rows"]
        ),
        "missing_current_key_count": len(
            current_result["missing_key_rows"]
        ),
    }

@app.get("/backup-restore/manifests")
def backup_restore_manifests():
    """
    Route tương thích với đường dẫn cũ.

    Hệ thống hiện đọc manifest.json trực tiếp từ file backup
    trên Google Drive, nên không còn dùng danh sách manifest
    lưu trong Supabase.
    """
    if not restore_password_has_been_created():
        return redirect(
            url_for("backup_restore_security")
        )

    if not is_restore_area_unlocked():
        return redirect(
            url_for(
                "backup_restore_security",
                next=url_for("backup_restore")
            )
        )

    return redirect(
        url_for("backup_restore")
    )


@app.get(
    "/backup-restore/manifest/<manifest_id>"
)
def backup_restore_manifest_detail(manifest_id):
    """
    Trang chi tiết một manifest.
    """
    if not is_restore_area_unlocked():
        return redirect(
            url_for(
                "backup_restore_security",
                next=request.path
            )
        )

    manifest = get_backup_manifest_by_id_web(
        manifest_id
    )

    if not manifest:
        flash(
            "Không tìm thấy manifest backup.",
            "danger"
        )
        return redirect(
            url_for("backup_restore")
        )

    items = get_backup_manifest_items_web(
        manifest_id
    )

    table_items = [
        item
        for item in items
        if item.get("item_type") == "table"
    ]

    bucket_items = [
        item
        for item in items
        if item.get("item_type") == "bucket"
    ]

    app_items = [
        item
        for item in items
        if item.get("item_type") == "app"
    ]

    comparable_tables = [
        item.get("item_name")
        for item in table_items
        if item.get("item_name")
        in BACKUP_TABLE_COMPARE_CONFIGS
    ]

    return render_template(
        "backup_restore_detail.html",
        manifest=manifest,
        items=items,
        table_items=table_items,
        bucket_items=bucket_items,
        app_items=app_items,

        table_compare_configs=(
            BACKUP_TABLE_COMPARE_CONFIGS
        ),
        comparable_tables=comparable_tables,
    )



@app.get(
    "/backup-restore/manifest/"
    "<manifest_id>/compare/<table_name>"
)
def backup_restore_compare_table(
    manifest_id,
    table_name
):
    """
    So sánh một bảng bất kỳ trong backup
    với bảng hiện tại trên Supabase.
    """
    if not is_restore_area_unlocked():
        return redirect(
            url_for(
                "backup_restore_security",
                next=request.path,
            )
        )

    table_name = str(
        table_name or ""
    ).strip()

    config = get_table_compare_config_web(
        table_name
    )

    # Chặn người dùng tự sửa URL để đọc bảng ngoài danh sách.
    if not config:
        flash(
            f"Bảng {table_name} chưa được phép so sánh.",
            "danger"
        )

        return redirect(
            url_for(
                "backup_restore_manifest_detail",
                manifest_id=manifest_id,
            )
        )

    manifest = get_backup_manifest_by_id_web(
        manifest_id
    )

    if not manifest:
        flash(
            "Không tìm thấy manifest backup.",
            "danger"
        )

        return redirect(
            url_for(
                "backup_restore_manifests"
            )
        )

    table_item = get_manifest_table_item_web(
        manifest_id,
        table_name,
    )

    if not table_item:
        flash(
            f"Manifest này không có bảng {table_name}.",
            "danger"
        )

        return redirect(
            url_for(
                "backup_restore_manifest_detail",
                manifest_id=manifest_id,
            )
        )

    backup_root = get_backup_local_root_web(
        manifest
    )

    backup_path = str(
        table_item.get("backup_path")
        or f"database_exports/{table_name}.json"
    ).strip()

    try:
        json_path = safe_backup_relative_path_web(
            backup_root,
            backup_path,
        )

        raw_backup_data = load_json_file_web(
            json_path
        )

        backup_rows = (
            extract_rows_from_export_json_web(
                raw_backup_data
            )
        )

        current_rows = (
            fetch_all_supabase_rows_web(
                table_name
            )
        )

        compare_result = (
            compare_generic_table_backup_web(
                table_name,
                backup_rows,
                current_rows,
            )
        )

        log_restore_history(
            action_type="compare_backup_table",
            status="success",
            backup_file=manifest.get(
                "backup_name"
            ),
            selected_items={
                "manifest_id": manifest_id,
                "table": table_name,
                "backup_path": backup_path,
            },
            restore_mode="preview_only",
        )

        table_meta = get_backup_restore_table_meta_web(
            table_name
        )

        return render_template(
            "backup_restore_compare_table.html",
            manifest=manifest,
            table_item=table_item,
            table_name=table_name,
            table_meta=table_meta,
            backup_root=str(backup_root),
            json_path=str(json_path),
            compare_result=compare_result,
        )

    except Exception as e:
        print(
            "[COMPARE GENERIC BACKUP ERROR]",
            table_name,
            repr(e),
        )

        log_restore_history(
            action_type="compare_backup_table",
            status="failed",
            backup_file=manifest.get(
                "backup_name"
            ),
            selected_items={
                "manifest_id": manifest_id,
                "table": table_name,
                "backup_path": backup_path,
            },
            restore_mode="preview_only",
            error_message=str(e),
        )

        flash(
            f"Không so sánh được bảng "
            f"{table_name}: {e}",
            "danger"
        )

        return redirect(
            url_for(
                "backup_restore_manifest_detail",
                manifest_id=manifest_id,
            )
        )

# =========================================================
# BACKUP & RESTORE - STORAGE
# Chế độ an toàn: chỉ phục hồi file còn thiếu.
# =========================================================

BACKUP_RESTORE_STORAGE_BUCKETS = {
    "student-photos",
    "system-assets",
}


def normalize_storage_object_path_web(value):
    return str(value or "").replace("\\", "/").strip("/")


def storage_item_is_folder_web(item):
    if not isinstance(item, dict):
        return False

    if item.get("id"):
        return False

    metadata = item.get("metadata")

    if isinstance(metadata, dict):
        if metadata.get("size") not in [None, ""]:
            return False

    return True


def list_storage_directory_web(
    bucket_name,
    folder_path="",
):
    folder_path = normalize_storage_object_path_web(
        folder_path
    )

    all_items = []
    offset = 0
    page_size = 1000

    while True:
        options = {
            "limit": page_size,
            "offset": offset,
            "sortBy": {
                "column": "name",
                "order": "asc",
            },
        }

        batch = (
            supabase.storage
            .from_(bucket_name)
            .list(folder_path, options)
            or []
        )

        all_items.extend(batch)

        if len(batch) < page_size:
            break

        offset += page_size

    return all_items


def list_all_storage_files_web(bucket_name):
    files = []
    folders_to_scan = [""]

    while folders_to_scan:
        current_folder = folders_to_scan.pop(0)

        for item in list_storage_directory_web(
            bucket_name,
            current_folder,
        ):
            item_name = str(
                item.get("name") or ""
            ).strip()

            if not item_name:
                continue

            object_path = normalize_storage_object_path_web(
                f"{current_folder}/{item_name}"
            )

            if storage_item_is_folder_web(item):
                folders_to_scan.append(object_path)
                continue

            metadata = item.get("metadata") or {}

            try:
                size_bytes = int(
                    metadata.get("size")
                    or item.get("size")
                    or 0
                )
            except Exception:
                size_bytes = 0

            files.append({
                "path": object_path,
                "size": size_bytes,
                "size_text": format_file_size_web(
                    size_bytes
                ),
                "mime_type": (
                    metadata.get("mimetype")
                    or metadata.get("contentType")
                    or ""
                ),
            })

    files.sort(
        key=lambda item: item["path"].lower()
    )

    return files


def get_manifest_bucket_item_web(
    manifest_id,
    bucket_name,
):
    try:
        rows = (
            supabase.table(
                BACKUP_MANIFEST_ITEM_TABLE
            )
            .select("*")
            .eq("manifest_id", manifest_id)
            .eq("item_type", "bucket")
            .eq("item_name", bucket_name)
            .limit(1)
            .execute()
            .data
            or []
        )

        return rows[0] if rows else {}

    except Exception as e:
        print(
            "[GET MANIFEST BUCKET ITEM ERROR]",
            repr(e),
        )
        return {}


def load_backup_bucket_index_web(
    manifest,
    bucket_item,
    bucket_name,
):
    backup_root = get_backup_local_root_web(
        manifest
    )

    bucket_backup_path = str(
        (bucket_item or {}).get("backup_path")
        or f"storage_exports/{bucket_name}"
    ).strip()

    bucket_directory = safe_backup_relative_path_web(
        backup_root,
        bucket_backup_path,
    )

    index_file = (
        bucket_directory
        / "_bucket_manifest.json"
    )

    raw_index = load_json_file_web(
        index_file
    )

    file_entries = []

    for item in (
        raw_index.get("files")
        if isinstance(raw_index, dict)
        else []
    ) or []:
        if not isinstance(item, dict):
            continue

        object_path = normalize_storage_object_path_web(
            item.get("path")
        )

        if not object_path:
            continue

        # File kỹ thuật của bản backup không phải object Storage.
        if object_path == "_bucket_manifest.json":
            continue

        try:
            size_bytes = int(
                item.get("size") or 0
            )
        except Exception:
            size_bytes = 0

        file_entries.append({
            "path": object_path,
            "backup_path": str(
                item.get("backup_path")
                or (
                    f"{bucket_backup_path}/"
                    f"{object_path}"
                )
            ).strip(),
            "size": size_bytes,
            "size_text": format_file_size_web(
                size_bytes
            ),
            "sha256": str(
                item.get("sha256") or ""
            ).strip(),
            "mime_type": str(
                item.get("mime_type") or ""
            ).strip(),
        })

    file_entries.sort(
        key=lambda item: item["path"].lower()
    )

    return {
        "backup_root": backup_root,
        "bucket_directory": bucket_directory,
        "index_file": index_file,
        "files": file_entries,
    }


def compare_storage_bucket_web(
    backup_files,
    current_files,
):
    backup_map = {
        item["path"]: item
        for item in backup_files or []
    }

    current_map = {
        item["path"]: item
        for item in current_files or []
    }

    missing_files = []
    existing_files = []
    current_only_files = []

    for object_path, backup_item in backup_map.items():
        if object_path in current_map:
            existing_files.append(
                backup_item
            )
        else:
            missing_files.append(
                backup_item
            )

    for object_path, current_item in current_map.items():
        if object_path not in backup_map:
            current_only_files.append(
                current_item
            )

    return {
        "backup_count": len(backup_map),
        "current_count": len(current_map),
        "missing_count": len(missing_files),
        "existing_count": len(existing_files),
        "current_only_count": len(
            current_only_files
        ),
        "missing_files": missing_files,
        "existing_files": existing_files,
        "current_only_files": current_only_files,
    }


def guess_storage_content_type_web(
    object_path,
    fallback="",
):
    fallback = str(fallback or "").strip()

    if fallback:
        return fallback

    extension = Path(
        str(object_path or "")
    ).suffix.lower()

    content_type_map = {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".webp": "image/webp",
        ".gif": "image/gif",
        ".svg": "image/svg+xml",
        ".pdf": "application/pdf",
        ".json": "application/json",
        ".txt": "text/plain",
    }

    return content_type_map.get(
        extension,
        "application/octet-stream",
    )


def storage_object_exists_web(
    bucket_name,
    object_path,
):
    object_path = normalize_storage_object_path_web(
        object_path
    )

    parent_folder = (
        object_path.rsplit("/", 1)[0]
        if "/" in object_path
        else ""
    )

    file_name = (
        object_path.rsplit("/", 1)[-1]
    )

    for item in list_storage_directory_web(
        bucket_name,
        parent_folder,
    ):
        if storage_item_is_folder_web(item):
            continue

        if str(item.get("name") or "").strip() == file_name:
            return True

    return False


@app.get(
    "/backup-restore/manifest/"
    "<manifest_id>/storage/<bucket_name>"
)
def backup_restore_compare_storage(
    manifest_id,
    bucket_name,
):
    if not is_restore_area_unlocked():
        return redirect(
            url_for(
                "backup_restore_security",
                next=request.path,
            )
        )

    bucket_name = str(
        bucket_name or ""
    ).strip()

    if bucket_name not in BACKUP_RESTORE_STORAGE_BUCKETS:
        flash(
            f"Bucket {bucket_name} chưa được phép phục hồi.",
            "danger",
        )
        return redirect(
            url_for(
                "backup_restore_manifest_detail",
                manifest_id=manifest_id,
            )
        )

    manifest = get_backup_manifest_by_id_web(
        manifest_id
    )

    bucket_item = get_manifest_bucket_item_web(
        manifest_id,
        bucket_name,
    )

    if not manifest or not bucket_item:
        flash(
            "Không tìm thấy manifest hoặc bucket backup.",
            "danger",
        )
        return redirect(
            url_for(
                "backup_restore_manifest_detail",
                manifest_id=manifest_id,
            )
        )

    try:
        backup_context = (
            load_backup_bucket_index_web(
                manifest,
                bucket_item,
                bucket_name,
            )
        )

        current_files = (
            list_all_storage_files_web(
                bucket_name
            )
        )

        compare_result = (
            compare_storage_bucket_web(
                backup_context["files"],
                current_files,
            )
        )

        log_restore_history(
            action_type="compare_backup_storage",
            status="success",
            backup_file=manifest.get(
                "backup_name"
            ),
            selected_items={
                "manifest_id": manifest_id,
                "bucket": bucket_name,
                "backup_count": (
                    compare_result[
                        "backup_count"
                    ]
                ),
                "missing_count": (
                    compare_result[
                        "missing_count"
                    ]
                ),
            },
            restore_mode="preview_only",
        )

        return render_template(
            "backup_restore_compare_storage.html",
            manifest=manifest,
            bucket_item=bucket_item,
            bucket_name=bucket_name,
            compare_result=compare_result,
        )

    except Exception as e:
        print(
            "[COMPARE STORAGE ERROR]",
            bucket_name,
            repr(e),
        )

        flash(
            f"Không so sánh được bucket "
            f"{bucket_name}: {e}",
            "danger",
        )

        return redirect(
            url_for(
                "backup_restore_manifest_detail",
                manifest_id=manifest_id,
            )
        )


@app.post(
    "/backup-restore/manifest/"
    "<manifest_id>/storage/"
    "<bucket_name>/restore-missing"
)
def backup_restore_storage_restore_missing(
    manifest_id,
    bucket_name,
):
    if not is_restore_area_unlocked():
        return redirect(
            url_for(
                "backup_restore_security",
                next=request.path,
            )
        )

    bucket_name = str(
        bucket_name or ""
    ).strip()

    if bucket_name not in BACKUP_RESTORE_STORAGE_BUCKETS:
        flash(
            f"Bucket {bucket_name} chưa được phép phục hồi.",
            "danger",
        )
        return redirect(
            url_for(
                "backup_restore_manifest_detail",
                manifest_id=manifest_id,
            )
        )

    selected_paths = []

    for raw_path in request.form.getlist(
        "selected_paths"
    ):
        object_path = normalize_storage_object_path_web(
            raw_path
        )

        if (
            object_path
            and object_path not in selected_paths
        ):
            selected_paths.append(
                object_path
            )

    if not selected_paths:
        flash(
            "Ken chưa chọn file Storage cần phục hồi.",
            "danger",
        )
        return redirect(
            url_for(
                "backup_restore_compare_storage",
                manifest_id=manifest_id,
                bucket_name=bucket_name,
            )
        )

    manifest = get_backup_manifest_by_id_web(
        manifest_id
    )

    bucket_item = get_manifest_bucket_item_web(
        manifest_id,
        bucket_name,
    )

    if not manifest or not bucket_item:
        flash(
            "Không tìm thấy manifest hoặc bucket backup.",
            "danger",
        )
        return redirect(
            url_for(
                "backup_restore_manifest_detail",
                manifest_id=manifest_id,
            )
        )

    restored = []
    skipped = []
    failed = []

    try:
        backup_context = (
            load_backup_bucket_index_web(
                manifest,
                bucket_item,
                bucket_name,
            )
        )

        backup_map = {
            item["path"]: item
            for item in backup_context["files"]
        }

        for object_path in selected_paths:
            backup_item = backup_map.get(
                object_path
            )

            if not backup_item:
                skipped.append({
                    "path": object_path,
                    "reason": (
                        "File không còn tồn tại "
                        "trong chỉ mục backup."
                    ),
                })
                continue

            try:
                # Kiểm tra lại ngay trước khi upload.
                if storage_object_exists_web(
                    bucket_name,
                    object_path,
                ):
                    skipped.append({
                        "path": object_path,
                        "reason": (
                            "File đã tồn tại trên Supabase."
                        ),
                    })
                    continue

                local_file = safe_backup_relative_path_web(
                    backup_context["backup_root"],
                    backup_item["backup_path"],
                )

                if not local_file.is_file():
                    raise FileNotFoundError(
                        f"Không tìm thấy file backup: "
                        f"{local_file}"
                    )

                expected_sha256 = str(
                    backup_item.get("sha256") or ""
                ).strip()

                if expected_sha256:
                    actual_sha256 = sha256_file(
                        local_file
                    )

                    if actual_sha256 != expected_sha256:
                        raise ValueError(
                            "Checksum file backup không khớp."
                        )

                content = local_file.read_bytes()
                content_type = (
                    guess_storage_content_type_web(
                        object_path,
                        backup_item.get(
                            "mime_type"
                        ),
                    )
                )

                (
                    supabase.storage
                    .from_(bucket_name)
                    .upload(
                        object_path,
                        content,
                        {
                            "content-type": content_type,
                            "upsert": "false",
                        },
                    )
                )

                restored.append({
                    "path": object_path,
                    "size": len(content),
                    "sha256": expected_sha256,
                })

            except Exception as file_error:
                failed.append({
                    "path": object_path,
                    "error": str(file_error),
                })

        log_restore_history(
            action_type=(
                "restore_storage_missing_files"
            ),
            status=(
                "success"
                if not failed
                else "failed"
            ),
            backup_file=manifest.get(
                "backup_name"
            ),
            selected_items={
                "manifest_id": manifest_id,
                "bucket": bucket_name,
                "requested_paths": selected_paths,
                "restored": restored,
                "skipped": skipped,
                "failed": failed,
                "safety": {
                    "overwrite": False,
                    "delete": False,
                    "mode": (
                        "restore_missing_files_only"
                    ),
                },
            },
            restore_mode=(
                "restore_missing_files_only"
            ),
            error_message=(
                f"Có {len(failed)} file lỗi."
                if failed
                else None
            ),
        )

        if restored:
            flash(
                f"Đã phục hồi {len(restored)} "
                f"file còn thiếu vào bucket "
                f"{bucket_name}.",
                "success",
            )

        if skipped:
            flash(
                f"Đã bỏ qua {len(skipped)} file "
                f"đã tồn tại hoặc không còn hợp lệ.",
                "warning",
            )

        if failed:
            flash(
                f"Có {len(failed)} file phục hồi lỗi.",
                "danger",
            )

        if not restored and not failed:
            flash(
                "Không còn file cần phục hồi.",
                "warning",
            )

    except Exception as e:
        print(
            "[RESTORE STORAGE ERROR]",
            bucket_name,
            repr(e),
        )

        flash(
            f"Không thể phục hồi Storage: {e}",
            "danger",
        )

    return redirect(
        url_for(
            "backup_restore_compare_storage",
            manifest_id=manifest_id,
            bucket_name=bucket_name,
        )
    )

# =========================================================
# BACKUP & RESTORE - MÃ NGUỒN ỨNG DỤNG
# So sánh theo đường dẫn, dung lượng và SHA-256.
# =========================================================

APP_RESTORE_BLOCKED_PATH_PARTS = {
    ".git",
    ".env",
    "venv",
    ".venv",
    "__pycache__",
    "node_modules",
    "restore_rollback",
}

APP_RESTORE_BLOCKED_FILE_NAMES = {
    ".env",
    "rclone.conf",
    "_app_manifest.json",
}


def normalize_app_relative_path_web(value):
    value = str(value or "").replace("\\", "/").strip("/")

    if not value:
        return ""

    pure_path = Path(value)

    if pure_path.is_absolute():
        raise ValueError("Đường dẫn mã nguồn tuyệt đối không hợp lệ.")

    if ".." in pure_path.parts:
        raise ValueError("Đường dẫn mã nguồn không an toàn.")

    return pure_path.as_posix()


def is_app_restore_path_allowed_web(relative_path):
    relative_path = normalize_app_relative_path_web(
        relative_path
    )

    if not relative_path:
        return False

    parts = Path(relative_path).parts

    if any(
        part in APP_RESTORE_BLOCKED_PATH_PARTS
        for part in parts
    ):
        return False

    if Path(relative_path).name in APP_RESTORE_BLOCKED_FILE_NAMES:
        return False

    return True


def get_manifest_app_item_web(
    manifest_id,
    app_name,
):
    try:
        rows = (
            supabase.table(
                BACKUP_MANIFEST_ITEM_TABLE
            )
            .select("*")
            .eq("manifest_id", manifest_id)
            .eq("item_type", "app")
            .eq("item_name", app_name)
            .limit(1)
            .execute()
            .data
            or []
        )

        return rows[0] if rows else {}

    except Exception as e:
        print("[GET MANIFEST APP ITEM ERROR]", repr(e))
        return {}


def load_backup_app_index_web(
    manifest,
    app_item,
):
    backup_root = get_backup_local_root_web(
        manifest
    )

    app_backup_path = str(
        (app_item or {}).get("backup_path")
        or "app_exports/application"
    ).strip()

    app_directory = safe_backup_relative_path_web(
        backup_root,
        app_backup_path,
    )

    index_file = app_directory / "_app_manifest.json"
    raw_index = load_json_file_web(index_file)

    file_entries = []

    for item in (
        raw_index.get("files")
        if isinstance(raw_index, dict)
        else []
    ) or []:
        if not isinstance(item, dict):
            continue

        relative_path = normalize_app_relative_path_web(
            item.get("path")
        )

        if (
            not relative_path
            or not is_app_restore_path_allowed_web(
                relative_path
            )
        ):
            continue

        try:
            file_size = int(
                item.get("size") or 0
            )
        except Exception:
            file_size = 0

        file_entries.append({
            "path": relative_path,
            "backup_path": str(
                item.get("backup_path")
                or (
                    f"{app_backup_path}/"
                    f"{relative_path}"
                )
            ).strip(),
            "size": file_size,
            "sha256": str(
                item.get("sha256") or ""
            ).strip(),
        })

    file_entries.sort(
        key=lambda item: item["path"].lower()
    )

    return {
        "backup_root": backup_root,
        "app_directory": app_directory,
        "index_file": index_file,
        "files": file_entries,
    }


def list_current_app_files_web():
    file_entries = []

    for file_path in BASE_DIR.rglob("*"):
        if not file_path.is_file():
            continue

        try:
            relative_path = file_path.relative_to(
                BASE_DIR
            ).as_posix()
        except Exception:
            continue

        if not is_app_restore_path_allowed_web(
            relative_path
        ):
            continue

        file_entries.append({
            "path": relative_path,
            "size": file_path.stat().st_size,
            "sha256": sha256_file(file_path),
        })

    file_entries.sort(
        key=lambda item: item["path"].lower()
    )

    return file_entries


def compare_app_files_web(
    backup_files,
    current_files,
):
    backup_map = {
        item["path"]: item
        for item in backup_files or []
    }

    current_map = {
        item["path"]: item
        for item in current_files or []
    }

    missing_files = []
    changed_files = []
    same_files = []
    current_only_files = []

    for relative_path, backup_item in backup_map.items():
        current_item = current_map.get(
            relative_path
        )

        base_info = {
            "path": relative_path,
            "backup_size": int(
                backup_item.get("size") or 0
            ),
            "backup_size_text": format_file_size_web(
                backup_item.get("size") or 0
            ),
            "backup_sha256": backup_item.get(
                "sha256"
            ) or "",
            "backup_path": backup_item.get(
                "backup_path"
            ) or "",
        }

        if not current_item:
            missing_files.append(base_info)
            continue

        current_size = int(
            current_item.get("size") or 0
        )
        current_sha256 = str(
            current_item.get("sha256") or ""
        )

        base_info.update({
            "current_size": current_size,
            "current_size_text": format_file_size_web(
                current_size
            ),
            "current_sha256": current_sha256,
        })

        if (
            str(base_info["backup_sha256"])
            == current_sha256
        ):
            same_files.append(base_info)
        else:
            changed_files.append(base_info)

    for relative_path, current_item in current_map.items():
        if relative_path in backup_map:
            continue

        current_size = int(
            current_item.get("size") or 0
        )

        current_only_files.append({
            "path": relative_path,
            "current_size": current_size,
            "current_size_text": format_file_size_web(
                current_size
            ),
            "current_sha256": current_item.get(
                "sha256"
            ) or "",
        })

    return {
        "backup_count": len(backup_map),
        "current_count": len(current_map),
        "missing_count": len(missing_files),
        "changed_count": len(changed_files),
        "same_count": len(same_files),
        "current_only_count": len(
            current_only_files
        ),
        "missing_files": missing_files,
        "changed_files": changed_files,
        "same_files": same_files,
        "current_only_files": current_only_files,
    }


def create_app_rollback_directory_web():
    timestamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S_%f"
    )

    rollback_root = (
        BASE_DIR
        / "restore_rollback"
        / timestamp
    )

    rollback_root.mkdir(
        parents=True,
        exist_ok=True,
    )

    return rollback_root


def copy_file_to_rollback_web(
    current_file,
    relative_path,
    rollback_root,
):
    if not current_file.is_file():
        return None

    rollback_file = (
        rollback_root
        / relative_path
    )

    rollback_file.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    shutil.copy2(
        current_file,
        rollback_file,
    )

    return rollback_file


def validate_restored_python_file_web(file_path):
    if file_path.suffix.lower() != ".py":
        return

    result = subprocess.run(
        [
            sys.executable,
            "-m",
            "py_compile",
            str(file_path),
        ],
        capture_output=True,
        text=True,
        timeout=60,
        check=False,
    )

    if result.returncode != 0:
        error_text = (
            result.stderr
            or result.stdout
            or "Python compile failed."
        ).strip()

        raise ValueError(error_text)


def restore_one_app_file_web(
    backup_root,
    backup_item,
    rollback_root,
    allow_overwrite,
):
    relative_path = normalize_app_relative_path_web(
        backup_item.get("path")
    )

    if not is_app_restore_path_allowed_web(
        relative_path
    ):
        raise ValueError(
            f"File bị chặn phục hồi: {relative_path}"
        )

    backup_file = safe_backup_relative_path_web(
        backup_root,
        backup_item.get("backup_path"),
    )

    if not backup_file.is_file():
        raise FileNotFoundError(
            f"Không tìm thấy file backup: {backup_file}"
        )

    expected_sha256 = str(
        backup_item.get("sha256") or ""
    ).strip()

    if expected_sha256:
        actual_sha256 = sha256_file(
            backup_file
        )

        if actual_sha256 != expected_sha256:
            raise ValueError(
                "Checksum file backup không khớp."
            )

    current_file = (
        BASE_DIR
        / relative_path
    ).resolve()

    try:
        current_file.relative_to(
            BASE_DIR.resolve()
        )
    except ValueError:
        raise ValueError(
            "Đường dẫn đích nằm ngoài thư mục ứng dụng."
        )

    existed_before = current_file.exists()

    if existed_before and not allow_overwrite:
        return {
            "status": "skipped",
            "reason": "File hiện tại đã tồn tại.",
        }

    rollback_file = None

    if existed_before:
        rollback_file = copy_file_to_rollback_web(
            current_file,
            relative_path,
            rollback_root,
        )

    current_file.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    shutil.copy2(
        backup_file,
        current_file,
    )

    try:
        validate_restored_python_file_web(
            current_file
        )

    except Exception:
        if rollback_file and rollback_file.is_file():
            shutil.copy2(
                rollback_file,
                current_file,
            )
        elif current_file.exists():
            current_file.unlink()

        raise

    return {
        "status": "restored",
        "path": relative_path,
        "rollback_path": (
            str(rollback_file)
            if rollback_file
            else None
        ),
        "overwritten": existed_before,
    }


@app.get(
    "/backup-restore/manifest/"
    "<manifest_id>/app/<app_name>"
)
def backup_restore_compare_app(
    manifest_id,
    app_name,
):
    if not is_restore_area_unlocked():
        return redirect(
            url_for(
                "backup_restore_security",
                next=request.path,
            )
        )

    app_name = str(
        app_name or ""
    ).strip()

    manifest = get_backup_manifest_by_id_web(
        manifest_id
    )

    app_item = get_manifest_app_item_web(
        manifest_id,
        app_name,
    )

    if not manifest or not app_item:
        flash(
            "Không tìm thấy manifest hoặc mã nguồn backup.",
            "danger",
        )
        return redirect(
            url_for(
                "backup_restore_manifest_detail",
                manifest_id=manifest_id,
            )
        )

    try:
        backup_context = (
            load_backup_app_index_web(
                manifest,
                app_item,
            )
        )

        current_files = (
            list_current_app_files_web()
        )

        compare_result = compare_app_files_web(
            backup_context["files"],
            current_files,
        )

        log_restore_history(
            action_type="compare_backup_app",
            status="success",
            backup_file=manifest.get(
                "backup_name"
            ),
            selected_items={
                "manifest_id": manifest_id,
                "app_name": app_name,
                "missing_count": compare_result[
                    "missing_count"
                ],
                "changed_count": compare_result[
                    "changed_count"
                ],
            },
            restore_mode="preview_only",
        )

        return render_template(
            "backup_restore_compare_app.html",
            manifest=manifest,
            app_item=app_item,
            app_name=app_name,
            current_app_root=str(BASE_DIR),
            compare_result=compare_result,
        )

    except Exception as e:
        print(
            "[COMPARE APP ERROR]",
            repr(e),
        )

        flash(
            f"Không so sánh được mã nguồn: {e}",
            "danger",
        )

        return redirect(
            url_for(
                "backup_restore_manifest_detail",
                manifest_id=manifest_id,
            )
        )


def restore_selected_app_files_route_web(
    manifest_id,
    app_name,
    allow_overwrite,
):
    if not is_restore_area_unlocked():
        return redirect(
            url_for(
                "backup_restore_security",
                next=request.path,
            )
        )

    selected_paths = []

    for raw_path in request.form.getlist(
        "selected_paths"
    ):
        try:
            relative_path = normalize_app_relative_path_web(
                raw_path
            )
        except Exception:
            continue

        if (
            relative_path
            and relative_path not in selected_paths
        ):
            selected_paths.append(
                relative_path
            )

    if not selected_paths:
        flash(
            "Ken chưa chọn file mã nguồn cần phục hồi.",
            "danger",
        )
        return redirect(
            url_for(
                "backup_restore_compare_app",
                manifest_id=manifest_id,
                app_name=app_name,
            )
        )

    manifest = get_backup_manifest_by_id_web(
        manifest_id
    )

    app_item = get_manifest_app_item_web(
        manifest_id,
        app_name,
    )

    if not manifest or not app_item:
        flash(
            "Không tìm thấy manifest hoặc mã nguồn backup.",
            "danger",
        )
        return redirect(
            url_for(
                "backup_restore_manifest_detail",
                manifest_id=manifest_id,
            )
        )

    restored = []
    skipped = []
    failed = []

    try:
        backup_context = (
            load_backup_app_index_web(
                manifest,
                app_item,
            )
        )

        backup_map = {
            item["path"]: item
            for item in backup_context["files"]
        }

        rollback_root = (
            create_app_rollback_directory_web()
        )

        for relative_path in selected_paths:
            backup_item = backup_map.get(
                relative_path
            )

            if not backup_item:
                skipped.append({
                    "path": relative_path,
                    "reason": "File không còn trong backup.",
                })
                continue

            try:
                result = restore_one_app_file_web(
                    backup_context["backup_root"],
                    backup_item,
                    rollback_root,
                    allow_overwrite=allow_overwrite,
                )

                if result.get("status") == "restored":
                    restored.append(result)
                else:
                    skipped.append({
                        "path": relative_path,
                        "reason": result.get(
                            "reason"
                        ) or "Đã bỏ qua.",
                    })

            except Exception as file_error:
                failed.append({
                    "path": relative_path,
                    "error": str(file_error),
                })

        log_restore_history(
            action_type=(
                "restore_app_changed_files"
                if allow_overwrite
                else "restore_app_missing_files"
            ),
            status=(
                "success"
                if not failed
                else "failed"
            ),
            backup_file=manifest.get(
                "backup_name"
            ),
            selected_items={
                "manifest_id": manifest_id,
                "app_name": app_name,
                "requested_paths": selected_paths,
                "restored": restored,
                "skipped": skipped,
                "failed": failed,
                "rollback_root": str(
                    rollback_root
                ),
                "safety": {
                    "delete": False,
                    "overwrite": allow_overwrite,
                    "python_compile_check": True,
                },
            },
            restore_mode=(
                "restore_changed_with_rollback"
                if allow_overwrite
                else "restore_missing_only"
            ),
            error_message=(
                f"Có {len(failed)} file lỗi."
                if failed
                else None
            ),
        )

        if restored:
            flash(
                f"Đã phục hồi {len(restored)} file mã nguồn. "
                f"Bản dự phòng nằm tại {rollback_root}.",
                "success",
            )

        if skipped:
            flash(
                f"Đã bỏ qua {len(skipped)} file.",
                "warning",
            )

        if failed:
            flash(
                f"Có {len(failed)} file phục hồi lỗi.",
                "danger",
            )

        if not restored and not failed:
            flash(
                "Không có file nào được phục hồi.",
                "warning",
            )

    except Exception as e:
        print(
            "[RESTORE APP ERROR]",
            repr(e),
        )

        flash(
            f"Không thể phục hồi mã nguồn: {e}",
            "danger",
        )

    return redirect(
        url_for(
            "backup_restore_compare_app",
            manifest_id=manifest_id,
            app_name=app_name,
        )
    )


@app.post(
    "/backup-restore/manifest/"
    "<manifest_id>/app/"
    "<app_name>/restore-missing"
)
def backup_restore_app_restore_missing(
    manifest_id,
    app_name,
):
    return restore_selected_app_files_route_web(
        manifest_id,
        app_name,
        allow_overwrite=False,
    )


@app.post(
    "/backup-restore/manifest/"
    "<manifest_id>/app/"
    "<app_name>/restore-changed"
)
def backup_restore_app_restore_changed(
    manifest_id,
    app_name,
):
    return restore_selected_app_files_route_web(
        manifest_id,
        app_name,
        allow_overwrite=True,
    )

# =========================================================
# BACKUP & RESTORE - MỘT NGUỒN DUY NHẤT
# Đọc manifest.json trực tiếp trong file tar.gz trên Drive.
#
# manifest_id trong các route cũ hiện được dùng như backup_name.
# Ví dụ:
# phungtkdsystem_full_2026-07-16_12-30.tar.gz
# =========================================================

BACKUP_RESTORE_CACHE_DIR = Path(
    os.environ.get(
        "BACKUP_RESTORE_CACHE_DIR",
        "/var/cache/phungtkdsystem/backup_restore",
    )
)

BACKUP_RCLONE_BIN = os.environ.get(
    "RCLONE_BIN",
    "/usr/bin/rclone",
)

BACKUP_RCLONE_CONFIG = os.environ.get(
    "RCLONE_CONFIG",
    "/root/.config/rclone/rclone.conf",
)


def normalize_backup_archive_name_web(value):
    """
    Chỉ chấp nhận đúng tên file backup của hệ thống.
    Chặn dấu /, .. và các tên file tùy ý.
    """
    backup_name = str(value or "").strip()

    if not re.fullmatch(
        r"phungtkdsystem_full_"
        r"\d{4}-\d{2}-\d{2}_"
        r"\d{2}-\d{2}\.tar\.gz",
        backup_name,
    ):
        raise ValueError(
            "Tên file backup không hợp lệ."
        )

    return backup_name


def get_backup_cache_directory_web(backup_name):
    """
    Mỗi file backup có một thư mục cache riêng.
    Cache nằm ngoài thư mục mã nguồn để không bị backup lồng.
    """
    backup_name = normalize_backup_archive_name_web(
        backup_name
    )

    cache_key = backup_name[:-7]

    cache_directory = (
        BACKUP_RESTORE_CACHE_DIR
        / cache_key
    )

    return cache_directory


def get_backup_cache_archive_web(backup_name):
    backup_name = normalize_backup_archive_name_web(
        backup_name
    )

    cache_directory = (
        get_backup_cache_directory_web(
            backup_name
        )
    )

    return cache_directory / backup_name


def ensure_tar_member_is_safe_web(
    extract_root,
    member,
):
    """
    Chặn:
    - ../../file
    - đường dẫn tuyệt đối
    - symbolic link
    - hard link
    - file trỏ ra ngoài thư mục cache
    """
    extract_root = Path(extract_root).resolve()

    member_name = str(
        member.name or ""
    ).replace("\\", "/")

    if not member_name:
        raise ValueError(
            "File backup có thành phần không hợp lệ."
        )

    pure_path = Path(member_name)

    if pure_path.is_absolute():
        raise ValueError(
            "File backup chứa đường dẫn tuyệt đối."
        )

    if ".." in pure_path.parts:
        raise ValueError(
            "File backup chứa đường dẫn không an toàn."
        )

    if member.issym() or member.islnk():
        raise ValueError(
            "File backup chứa liên kết không được phép."
        )

    target_path = (
        extract_root
        / member_name
    ).resolve()

    try:
        target_path.relative_to(
            extract_root
        )
    except ValueError:
        raise ValueError(
            "File backup có dữ liệu nằm ngoài thư mục giải nén."
        )


def safely_extract_backup_archive_web(
    archive_path,
    extract_root,
):
    """
    Giải nén archive sau khi kiểm tra toàn bộ đường dẫn.
    """
    archive_path = Path(archive_path)
    extract_root = Path(extract_root)

    if not archive_path.is_file():
        raise FileNotFoundError(
            f"Không tìm thấy file backup: {archive_path}"
        )

    extract_root.mkdir(
        parents=True,
        exist_ok=True,
    )

    with tarfile.open(
        archive_path,
        mode="r:gz",
    ) as archive:
        members = archive.getmembers()

        if not members:
            raise ValueError(
                "File backup không có dữ liệu."
            )

        for member in members:
            ensure_tar_member_is_safe_web(
                extract_root,
                member,
            )

        archive.extractall(
            path=extract_root,
            members=members,
        )


def download_backup_from_drive_web(
    backup_name,
    destination_file,
):
    """
    Tải đúng một file backup từ Google Drive bằng rclone.
    """
    backup_name = normalize_backup_archive_name_web(
        backup_name
    )

    destination_file = Path(
        destination_file
    )

    destination_file.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    remote_file = (
        f"{BACKUP_REMOTE_DIR}/"
        f"{backup_name}"
    )

    command = [
        BACKUP_RCLONE_BIN,
        "copyto",
        remote_file,
        str(destination_file),
        "--config",
        BACKUP_RCLONE_CONFIG,
    ]

    result = subprocess.run(
        command,
        capture_output=True,
        text=True,
        timeout=300,
        check=False,
    )

    if result.returncode != 0:
        error_text = (
            result.stderr
            or result.stdout
            or "Không tải được file backup từ Google Drive."
        ).strip()

        raise RuntimeError(error_text)

    if not destination_file.is_file():
        raise FileNotFoundError(
            "rclone chạy xong nhưng không tìm thấy file đã tải."
        )

    if destination_file.stat().st_size <= 0:
        raise ValueError(
            "File backup tải về có dung lượng bằng 0."
        )


def prepare_backup_archive_web(
    backup_name,
):
    """
    Bảo đảm file đã được tải, giải nén và có manifest.json.

    Nếu cache đã hoàn chỉnh thì dùng lại, không tải lại.
    """
    backup_name = normalize_backup_archive_name_web(
        backup_name
    )

    cache_directory = (
        get_backup_cache_directory_web(
            backup_name
        )
    )

    archive_file = (
        get_backup_cache_archive_web(
            backup_name
        )
    )

    extract_root = (
        cache_directory
        / "extracted"
    )

    manifest_file = (
        extract_root
        / "manifest.json"
    )

    ready_marker = (
        cache_directory
        / ".ready"
    )

    if (
        ready_marker.is_file()
        and manifest_file.is_file()
    ):
        return {
            "backup_name": backup_name,
            "cache_directory": cache_directory,
            "archive_file": archive_file,
            "extract_root": extract_root,
            "manifest_file": manifest_file,
        }

    BACKUP_RESTORE_CACHE_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    temporary_archive = (
        cache_directory
        / f"{backup_name}.downloading"
    )

    if cache_directory.exists():
        shutil.rmtree(
            cache_directory
        )

    cache_directory.mkdir(
        parents=True,
        exist_ok=True,
    )

    try:
        download_backup_from_drive_web(
            backup_name,
            temporary_archive,
        )

        temporary_archive.replace(
            archive_file
        )

        safely_extract_backup_archive_web(
            archive_file,
            extract_root,
        )

        if not manifest_file.is_file():
            raise FileNotFoundError(
                "File backup không có manifest.json."
            )

        required_paths = [
            extract_root / "database_exports",
            extract_root / "storage_exports",
            extract_root / "app_exports",
        ]

        missing_paths = [
            path.name
            for path in required_paths
            if not path.exists()
        ]

        if missing_paths:
            raise FileNotFoundError(
                "File backup thiếu thành phần: "
                + ", ".join(missing_paths)
            )

        ready_marker.write_text(
            datetime.now(
                timezone.utc
            ).isoformat(),
            encoding="utf-8",
        )

        return {
            "backup_name": backup_name,
            "cache_directory": cache_directory,
            "archive_file": archive_file,
            "extract_root": extract_root,
            "manifest_file": manifest_file,
        }

    except Exception:
        if cache_directory.exists():
            shutil.rmtree(
                cache_directory,
                ignore_errors=True,
            )

        raise


def load_backup_manifest_from_archive_web(
    backup_name,
):
    """
    Đọc manifest.json trong đúng file backup đã chọn.
    """
    context = prepare_backup_archive_web(
        backup_name
    )

    manifest = load_json_file_web(
        context["manifest_file"]
    )

    if not isinstance(manifest, dict):
        raise ValueError(
            "manifest.json không đúng cấu trúc."
        )

    manifest_backup_name = str(
        manifest.get("backup_name") or ""
    ).strip()

    if manifest_backup_name != context["backup_name"]:
        raise ValueError(
            "Tên trong manifest không khớp tên file backup."
        )

    result = dict(manifest)

    # Giữ tương thích với HTML và các route cũ.
    # Trước đây id là UUID Supabase.
    # Bây giờ id chính là tên file backup.
    result["id"] = context["backup_name"]
    result["local_extract_path"] = str(
        context["extract_root"]
    )
    result["local_archive_path"] = str(
        context["archive_file"]
    )

    if not result.get("file_size"):
        result["file_size"] = (
            context["archive_file"]
            .stat()
            .st_size
        )

    return result


def get_backup_manifest_by_id_web(
    manifest_id,
):
    """
    Ghi đè hàm cũ:
    Không còn đọc backup_manifests trên Supabase.

    manifest_id hiện chính là tên file tar.gz.
    """
    try:
        return load_backup_manifest_from_archive_web(
            manifest_id
        )

    except Exception as error:
        print(
            "[READ MANIFEST FROM ARCHIVE ERROR]",
            repr(error),
        )

        return {}


def get_backup_manifest_items_web(
    manifest_id,
):
    """
    Gộp ba nhóm trong manifest.json thành danh sách item.
    """
    manifest = get_backup_manifest_by_id_web(
        manifest_id
    )

    if not manifest:
        return []

    items = []

    for group_name in [
        "tables",
        "storage",
        "application",
    ]:
        group_items = manifest.get(
            group_name
        ) or []

        if not isinstance(
            group_items,
            list,
        ):
            continue

        for item in group_items:
            if isinstance(item, dict):
                items.append(
                    dict(item)
                )

    items.sort(
        key=lambda item: (
            str(
                item.get("item_type")
                or ""
            ),
            str(
                item.get("item_name")
                or ""
            ),
        )
    )

    return items


def get_manifest_table_item_web(
    manifest_id,
    table_name,
):
    table_name = str(
        table_name or ""
    ).strip()

    manifest = get_backup_manifest_by_id_web(
        manifest_id
    )

    for item in (
        manifest.get("tables")
        if manifest
        else []
    ) or []:
        if (
            isinstance(item, dict)
            and item.get("item_type") == "table"
            and str(
                item.get("item_name") or ""
            ).strip() == table_name
        ):
            return dict(item)

    return {}


def get_manifest_bucket_item_web(
    manifest_id,
    bucket_name,
):
    bucket_name = str(
        bucket_name or ""
    ).strip()

    manifest = get_backup_manifest_by_id_web(
        manifest_id
    )

    for item in (
        manifest.get("storage")
        if manifest
        else []
    ) or []:
        if (
            isinstance(item, dict)
            and item.get("item_type") == "bucket"
            and str(
                item.get("item_name") or ""
            ).strip() == bucket_name
        ):
            return dict(item)

    return {}


def get_manifest_app_item_web(
    manifest_id,
    app_name,
):
    app_name = str(
        app_name or ""
    ).strip()

    manifest = get_backup_manifest_by_id_web(
        manifest_id
    )

    for item in (
        manifest.get("application")
        if manifest
        else []
    ) or []:
        if (
            isinstance(item, dict)
            and item.get("item_type") == "app"
            and str(
                item.get("item_name") or ""
            ).strip() == app_name
        ):
            return dict(item)

    return {}


def get_backup_local_root_web(
    manifest,
):
    """
    Trả về thư mục đã giải nén của chính file backup đang chọn.
    """
    local_extract_path = str(
        (manifest or {}).get(
            "local_extract_path"
        ) or ""
    ).strip()

    if not local_extract_path:
        raise ValueError(
            "Bản backup chưa được tải và giải nén."
        )

    extract_root = Path(
        local_extract_path
    ).resolve()

    cache_root = (
        BACKUP_RESTORE_CACHE_DIR
        .resolve()
    )

    try:
        extract_root.relative_to(
            cache_root
        )
    except ValueError:
        raise ValueError(
            "Thư mục giải nén backup không an toàn."
        )

    manifest_file = (
        extract_root
        / "manifest.json"
    )

    if not manifest_file.is_file():
        raise FileNotFoundError(
            "Không tìm thấy manifest.json trong cache."
        )

    return extract_root

@app.context_processor
def inject_app_settings():
    settings = load_app_settings()

    fees = settings.get("fees", {})
    exam = settings.get("exam", {})
    class_options = settings.get("class_options", DEFAULT_APP_SETTINGS["class_options"])

    return {
        "app_settings": settings,
        "admin_logged_in": is_admin_logged_in(),
        "admin_username": session.get("admin_username", ADMIN_USERNAME),
        "tuition_fee": int(fees.get("tuition_fee", 500000) or 500000),
        "exam_fee": int(fees.get("exam_fee", 300000) or 300000),
        "dan_fees": fees.get("dan_fees", {}),
        "exam_number_prefix": exam.get("exam_number_prefix", "Cấp_"),
        "class_options": class_options
    }

@app.template_filter('money')
def money_filter(v):
    return format_money(v or 0)

@app.template_filter("datetime_vn")
def datetime_vn_filter(value):
    raw = str(value or "").strip()

    if not raw:
        return "—"

    try:
        dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))

        return dt.strftime("%d/%m/%Y %H:%M")

    except Exception:
        return raw

@app.get('/')
def index():
    return redirect(url_for('student_login'))

def remove_accents(text):
    text = str(text or "").replace("đ", "d").replace("Đ", "D")
    text = unicodedata.normalize("NFKD", text)
    return "".join(c for c in text if not unicodedata.combining(c))


def normalize_birthdate_web(raw):
    s = re.sub(r"\D", "", str(raw or ""))

    if len(s) == 8:
        day, month, year = s[:2], s[2:4], s[4:]
    elif len(s) == 6:
        day, month, yy = s[:2], s[2:4], s[4:]
        year = "19" + yy if int(yy) >= 30 else "20" + yy
    else:
        return ""

    try:
        d = datetime(int(year), int(month), int(day))
        return d.strftime("%d/%m/%Y")
    except:
        return ""


def auto_hv_code_web(name, birthdate):
    name = remove_accents(name).lower().strip()
    parts = [p for p in name.split() if p]

    if not parts:
        return ""

    ten = parts[-1]
    initials = "".join(p[0] for p in parts[:-1])

    birth = normalize_birthdate_web(birthdate)
    if not birth:
        return ""

    dd, mm, yyyy = birth.split("/")
    return f"HV_{ten}{initials}_{dd}{mm}{yyyy[-2:]}"

@app.get('/students')
def students():
    q = request.args.get('q', '').strip()
    new_license = request.args.get('new', '').strip()

    try:
        query = supabase.table(STUDENT_TABLE).select('*')

        if q:
            query = query.or_(f'license.ilike.%{q}%,name.ilike.%{q}%,phonenumber.ilike.%{q}%')

        rows = query.execute().data or []

    except Exception as e:
        print("[SUPABASE ERROR /students]", e)
        flash("Không kết nối được Supabase. Ken kiểm tra lại mạng, DNS hoặc SUPABASE_URL trong file .env.")
        rows = []

    def sort_student(r):
        license_code = str(r.get('license') or '').strip()

        # Hội viên vừa thêm luôn nằm đầu danh sách
        new_rank = 0 if new_license and license_code == new_license else 1

        # Sau đó vẫn ưu tiên HV đang hoạt động
        active_rank = 0 if is_active(r.get('active')) else 1

        # Sau đó sắp xếp theo id/idx mới nhất
        new_id = r.get('id') or r.get('idx') or 0
        try:
            new_id = int(new_id)
        except:
            new_id = 0

        return (new_rank, active_rank, -new_id)

    rows.sort(key=sort_student)

    return render_template(
        'students.html',
        rows=rows,
        q=q,
        new_license=new_license
    )

@app.post('/students/add')
def students_add():
    form = request.form

    birthdate = normalize_birthdate_web(form.get("birthdate", ""))

    if not birthdate:
        flash("Ngày sinh không hợp lệ")
        return redirect(url_for("students"))

    license_code = form.get("license", "").strip()

    if not license_code:
        license_code = auto_hv_code_web(form.get("name", ""), birthdate)

    if not license_code:
        flash("Không tạo được Mã HV. Ken kiểm tra lại Họ tên và Ngày sinh.")
        return redirect(url_for("students"))

    # =========================
    # CHẶN TRÙNG MÃ HỘI VIÊN
    # =========================
    duplicated = supabase.table(STUDENT_TABLE) \
        .select("license,name,birthdate") \
        .eq("license", license_code) \
        .limit(1) \
        .execute().data or []

    if duplicated:
        old = duplicated[0]
        flash(
            f"Mã HV {license_code} đã tồn tại cho học viên "
            f"{old.get('name', '')} - {old.get('birthdate', '')}. "
            f"Ken không thể thêm trùng mã."
        )
        return redirect(url_for("students"))
    payload = {
        'license': license_code,
        'name': form.get('name','').strip(),
        'birthdate': birthdate,
        'gender': form.get('gender','').strip(),
        'classroom': form.get('classroom','').strip(),
        'timeclass': form.get('timeclass','').strip(),
        'clup': form.get('clup','').strip(),
        'phonenumber': form.get('phonenumber','').strip(),
        'address': form.get('address', '').strip(),
        'belt': form.get('belt','Cấp 10').strip() or 'Cấp 10',
        'family': form.get('family','Không'),
        'active': form.get('active','Có'),
        'telegram_id': form.get('telegram_id','').strip(),
    }
    supabase.table(STUDENT_TABLE).insert(payload).execute()
    flash(f'Đã thêm học viên: {license_code}')

    return redirect(url_for('students', new=license_code))

@app.post('/students/<license_code>/delete')
def students_delete(license_code):
    supabase.table(STUDENT_TABLE).delete().eq('license', license_code).execute()
    flash(f'Đã xoá học viên: {license_code}', "success")
    return back_to_current_page("students")

@app.post('/students/<license_code>/update')
def students_update(license_code):
    form = request.form

    birthdate = normalize_birthdate_web(form.get("birthdate", ""))

    payload = {
        'name': form.get('name', '').strip(),
        'birthdate': birthdate,
        'gender': form.get('gender', '').strip(),
        'classroom': form.get('classroom', '').strip(),
        'timeclass': form.get('timeclass', '').strip(),
        'clup': form.get('clup', '').strip(),
        'phonenumber': form.get('phonenumber', '').strip(),
        'address': form.get('address', '').strip(),
        'belt': form.get('belt', 'Cấp 10').strip(),
        'family': form.get('family', 'Không'),
        'active': form.get('active', 'Có'),
        'telegram_id': form.get('telegram_id', '').strip(),
    }

    supabase.table(STUDENT_TABLE).update(payload).eq('license', license_code).execute()
    flash(f"Đã cập nhật học viên: {license_code}", "success")
    return back_to_current_page("students")

@app.get('/fees')
def fees():
    q = request.args.get('q','').strip()
    selected_license = request.args.get('ma_hv', '').strip()
    query = supabase.table(HOCPHI_TABLE).select('*').order('thoi_gian', desc=True).limit(300)

    if q:
        query = query.or_(f'ma_hv.ilike.%{q}%,ho_ten.ilike.%{q}%')

    rows = query.execute().data or []

    students = supabase.table(STUDENT_TABLE) \
        .select('license,name,birthdate,gender,classroom,timeclass,phonenumber,belt,family') \
        .order('name') \
        .execute().data or []

    now = datetime.now()

    return render_template(
        'fees.html',
        rows=rows,
        students=students,
        q=q,
        current_month=now.month,
        current_year=now.year,
        selected_license=selected_license
    )

@app.post('/fees/add')
def fees_add():
    f = request.form
    ma_hv = f.get('ma_hv', '').strip()

    sv = supabase.table(STUDENT_TABLE) \
        .select('*') \
        .eq('license', ma_hv) \
        .limit(1) \
        .execute().data

    if not sv:
        flash('Không tìm thấy Mã HV')
        return redirect(url_for('fees'))

    sv = sv[0]

    def money_to_int(v):
        return int(
            str(v or '0')
            .replace('đ', '')
            .replace('Đ', '')
            .replace('.', '')
            .replace(',', '')
            .replace(' ', '')
            .strip() or 0
        )

    # =========================
    # THÁNG HỌC PHÍ
    # =========================
    months, years = [], []

    for i in range(1, 13):
        m = f.get(f'month{i}', '').strip()
        y = f.get(f'year{i}', '').strip()

        if m and y:
            months.append(int(m))
            years.append(int(y))

    month_label, month_codes = build_month_codes(months, years)

    # =========================
    # THI CẤP / THI ĐẲNG
    # =========================
    current_year = str(datetime.now().year)

    exam_quarter = f.get("exam_quarter", "").strip().upper()
    exam_amount = money_to_int(f.get("exam_amount"))

    dan_quarter = f.get("dan_quarter", "").strip().upper()
    dan_level = f.get("dan_level", "").strip()
    dan_amount = money_to_int(f.get("dan_amount"))

    grand_total = money_to_int(f.get('grand_total'))

    discount_type = f.get('discount_type', '').strip()
    discount_value = money_to_int(f.get('discount_value'))

    payment_method = 'CK' if f.get('chuyen_khoan') else 'TM'

    # =========================
    # CỘT ĐÓNG PHÍ
    # =========================
    fee_parts = []

    if month_label:
        fee_parts.append(f"Học phí {month_label}")

    ma_quy = ""

    # Thi cấp: Q1/Q2/Q3/Q4 + năm
    if exam_quarter:
        ma_quy = f"{exam_quarter}{current_year}"

        quarter_number = exam_quarter.replace("Q", "")
        fee_parts.append(f"Thi cấp quý {quarter_number}-{current_year}")

    # Thi đẳng: L1/L2/MN/MT/MB/QG - năm
    # Nếu có thi đẳng thì ưu tiên ma_quy thi đẳng
    if dan_quarter and dan_level:
        ma_quy = f"{dan_quarter}-{current_year}"

        dan_label_map = {
            "L1": "Lần 1",
            "L2": "Lần 2",
            "MN": "KV miền nam",
            "MT": "KV miền trung",
            "MB": "KV miền bắc",
            "QG": "Quốc Gia",
        }

        dan_quarter_label = dan_label_map.get(dan_quarter, dan_quarter)
        fee_parts.append(f"Thi đẳng {dan_level} - {dan_quarter_label}-{current_year}")

    dong_phi_text = " - ".join(fee_parts)

    # =========================
    # CỘT GHI CHÚ
    # =========================
    main_note = f.get('note', '').strip()
    auto_fee_note = f.get('auto_fee_note', '').strip()

    sub_notes = [payment_method]

    if auto_fee_note:
        sub_notes.append(auto_fee_note)

    # Dự phòng nếu frontend không gửi auto_fee_note
    if not auto_fee_note:
        if discount_type == "half_month":
            sub_notes.append("Đóng nửa tháng - giảm 50% học phí")

        elif discount_type == "percent" and discount_value > 0:
            sub_notes.append(f"Giảm giá {discount_value}%")

        elif discount_type == "money" and discount_value > 0:
            sub_notes.append(f"Giảm giá {format_money(discount_value)}")

        family_value = str(sv.get("family") or "").strip().lower()
        family_value = remove_accents(family_value)

        if family_value in ["co", "yes", "true", "1"]:
            sub_notes.append("Gia đình - giảm 10% học phí")

    sub_note_text = " - ".join(sub_notes)

    if main_note:
        note = f"{main_note} ({sub_note_text})"
    else:
        note = f"({sub_note_text})"

    payload = {
        'thoi_gian': datetime.now().isoformat(),
        'ma_hv': ma_hv,
        'ho_ten': sv.get('name', ''),
        'ngay_sinh': sv.get('birthdate', ''),
        'gioi_tinh': sv.get('gender', ''),
        'lop': sv.get('classroom', ''),
        'ca': sv.get('timeclass', ''),
        'tong_tien': grand_total,
        'thang_dong_phi': dong_phi_text,
        'ghi_chu': note,
        'ma_thang': month_codes,
        'ma_quy': ma_quy,
        'chuyen_khoan': payment_method,
    }

    supabase.table(HOCPHI_TABLE).insert(payload).execute()

    # Nếu HV đang tạm nghỉ đúng tháng vừa đóng, tự chuyển về bình thường
    clear_student_temp_leave_if_paid(ma_hv, month_codes)

    flash(f'Đã lưu học phí: {sv.get("name")} - {format_money(grand_total)}', "success")
    return back_to_current_page("fees")

@app.post("/fees/delete/<fee_id>")
def delete_fee(fee_id):
    try:
        supabase.table(HOCPHI_TABLE) \
            .delete() \
            .eq("id", fee_id) \
            .execute()

        flash("Đã xóa phiếu thu", "success")

    except Exception as e:
        flash(f"Lỗi xóa: {e}", "danger")

    return back_to_current_page("fees")


def get_student_paid_months_count(ma_hv):
    """
    Đếm số tháng học viên đã đóng học phí.
    Dựa theo bảng hocphi và cột ma_thang.
    VD ma_thang: 062026 hoặc 062026 - 072026 - 082026
    """
    try:
        ma_hv = str(ma_hv or "").strip()

        if not ma_hv:
            return 0

        rows = supabase.table(HOCPHI_TABLE) \
            .select("ma_hv,ma_thang") \
            .eq("ma_hv", ma_hv) \
            .execute().data or []

        paid_set = set()

        for r in rows:
            ma_thang = str(r.get("ma_thang") or "").strip()

            if not ma_thang:
                continue

            month_codes = [
                x.strip()
                for x in re.split(r"\s*-\s*", ma_thang)
                if x.strip()
            ]

            for code in month_codes:
                # Chỉ nhận mã tháng dạng 062026, 072026...
                if re.fullmatch(r"\d{6}", code):
                    paid_set.add(code)

        return len(paid_set)

    except Exception as e:
        print("[WELCOME POPUP PAID MONTHS ERROR]", e)
        return 0


@app.context_processor
def inject_student_welcome_popup():
    try:
        student_license = str(session.get("student_license") or "").strip()
        student_name = str(session.get("student_name") or "").strip()

        show_popup = session.pop("show_student_welcome_popup", False)

        months_count = 0
        if student_license:
            months_count = get_student_paid_months_count(student_license)

        return {
            "show_student_welcome_popup": show_popup,
            "student_name": student_name,
            "welcome_months_count": months_count,
        }

    except Exception as e:
        print("[INJECT STUDENT WELCOME POPUP ERROR]", e)

        return {
            "show_student_welcome_popup": False,
            "student_name": "",
            "welcome_months_count": 0,
        }

def build_tracking_rows_web(year, month):
    """
    Dùng chung cho:
    - Trang Theo dõi học phí
    - File Excel xuất theo bộ lọc

    Quy tắc:
    - Phiếu đóng 1 tháng: hiện tiền ở tháng đó.
    - Phiếu đóng nhiều tháng: chỉ hiện tiền ở tháng đầu tiên.
    - Các tháng sau vẫn có dòng nhưng số tiền hiển thị là 0đ.
    - Tiền thi cấp/đẳng không tính vào học phí.
    """

    code = f"{int(month):02d}{year}"

    rows_all = (
        supabase.table(HOCPHI_TABLE)
        .select("*")
        .execute()
        .data
        or []
    )

    rows = []
    total = 0
    cash = 0
    bank = 0

    for source_row in rows_all:
        r = dict(source_row)

        ma_thang = str(r.get("ma_thang") or "").strip()

        # Chỉ lấy phiếu có tháng đang lọc
        if code not in ma_thang:
            continue

        amount = int(r.get("tong_tien") or 0)

        month_codes = [
            x.strip()
            for x in re.split(r"\s*-\s*", ma_thang)
            if x.strip()
        ]

        first_month_code = month_codes[0] if month_codes else ""

        # =========================
        # TÁCH TIỀN THI CẤP/ĐẲNG
        # =========================
        exam_fee = 0

        has_exam_fee = bool(
            str(r.get("ma_quy") or "").strip()
        )

        if has_exam_fee:
            exam_fee = get_app_setting_int(
                "fees.exam_fee",
                300000
            )

        tuition_amount = amount - exam_fee

        if tuition_amount < 0:
            tuition_amount = 0

        # =========================
        # TIỀN HIỂN THỊ THEO THÁNG
        # =========================
        if len(month_codes) <= 1:
            display_amount = tuition_amount

        elif code == first_month_code:
            display_amount = tuition_amount

        else:
            display_amount = 0

        r["display_tong_tien"] = display_amount
        r["_month_codes"] = month_codes

        if display_amount > 0:
            total += display_amount

            payment_method = str(
                r.get("chuyen_khoan") or ""
            ).strip().upper()

            if payment_method == "TM":
                cash += display_amount

            elif payment_method == "CK":
                bank += display_amount

        rows.append(r)

    rows.sort(
        key=lambda x: str(x.get("thoi_gian") or ""),
        reverse=True
    )

    total_count = len([
        r for r in rows
        if int(r.get("display_tong_tien") or 0) > 0
    ])

    cash_count = len([
        r for r in rows
        if int(r.get("display_tong_tien") or 0) > 0
        and str(r.get("chuyen_khoan") or "").strip().upper() == "TM"
    ])

    bank_count = len([
        r for r in rows
        if int(r.get("display_tong_tien") or 0) > 0
        and str(r.get("chuyen_khoan") or "").strip().upper() == "CK"
    ])

    return {
        "rows": rows,
        "total": total,
        "cash": cash,
        "bank": bank,
        "total_count": total_count,
        "cash_count": cash_count,
        "bank_count": bank_count,
    }


def build_tracking_export_note_web(row):
    """
    Tạo ghi chú rút gọn cho file Excel.

    Ví dụ:
    - Gia đình
    - Đóng nửa tháng
    - Đóng 3 tháng
    - Đóng 6 tháng
    - Đóng nửa tháng - Gia đình
    """

    raw = str(row.get("ghi_chu") or "").strip()
    normalized = remove_accents(raw).lower()

    notes = []

    if "nua thang" in normalized:
        notes.append("Đóng nửa tháng")

    if "dong 3 thang" in normalized:
        notes.append("Đóng 3 tháng")

    if (
        "dong 6 thang" in normalized
        or "tang 1" in normalized
    ):
        notes.append("Đóng 6 tháng")

    if "gia dinh" in normalized:
        notes.append("Gia đình")

    # Dự phòng: dữ liệu cũ không ghi rõ gói tháng
    month_codes = row.get("_month_codes") or []
    month_count = len(month_codes)

    has_package_note = any(
        x in notes
        for x in [
            "Đóng nửa tháng",
            "Đóng 3 tháng",
            "Đóng 6 tháng",
        ]
    )

    if not has_package_note:
        if month_count == 3:
            notes.append("Đóng 3 tháng")

        elif month_count in [6, 7]:
            notes.append("Đóng 6 tháng")

    result = []

    for item in notes:
        if item not in result:
            result.append(item)

    return " - ".join(result)


def build_tracking_discount_text_web(
    base_fee,
    final_amount,
    raw_note
):
    """
    Hiển thị giảm giá trong Excel.

    Ví dụ:
    - 50.000 đ
    - 50.000 đ (10%)
    - 250.000 đ (50%)
    """

    base_fee = int(base_fee or 0)
    final_amount = int(final_amount or 0)

    discount_amount = max(
        base_fee - final_amount,
        0
    )

    raw_note = str(raw_note or "")
    normalized = remove_accents(raw_note).lower()

    percent_values = re.findall(
        r"(?:giam(?: gia)?|gia dinh)"
        r"[^0-9]{0,30}"
        r"(\d+(?:[.,]\d+)?)\s*%",
        normalized
    )

    percent_parts = []

    for value in percent_values:
        value = value.replace(",", ".")

        try:
            number = float(value)
            label = f"{number:g}%"

            if label not in percent_parts:
                percent_parts.append(label)

        except Exception:
            pass

    # Đóng nửa tháng tương ứng giảm 50%
    if (
        "nua thang" in normalized
        and "50%" not in percent_parts
    ):
        percent_parts.insert(0, "50%")

    if discount_amount <= 0 and not percent_parts:
        return "—"

    amount_text = (
        format_money(discount_amount)
        if discount_amount > 0
        else ""
    )

    if amount_text and percent_parts:
        return (
            f"{amount_text} "
            f"({' + '.join(percent_parts)})"
        )

    return amount_text or " + ".join(percent_parts)

@app.get('/tracking')
def tracking():
    now = datetime.now()

    year = request.args.get(
        "year",
        str(now.year)
    )

    month = request.args.get(
        "month",
        str(now.month)
    )

    tracking_data = build_tracking_rows_web(
        year,
        month
    )

    return render_template(
        "tracking.html",

        rows=tracking_data["rows"],

        year=year,
        month=month,
        current_year=now.year,

        total=tracking_data["total"],
        cash=tracking_data["cash"],
        bank=tracking_data["bank"],

        total_count=tracking_data["total_count"],
        cash_count=tracking_data["cash_count"],
        bank_count=tracking_data["bank_count"],
    )



@app.get("/tracking/export")
def tracking_export():
    now = datetime.now()

    year = request.args.get(
        "year",
        str(now.year)
    ).strip()

    month = request.args.get(
        "month",
        str(now.month)
    ).strip()

    try:
        month_int = int(month)
        year_int = int(year)

    except Exception:
        month_int = now.month
        year_int = now.year

        month = str(month_int)
        year = str(year_int)

    tracking_data = build_tracking_rows_web(
        year,
        month
    )

    # =========================
    # LOẠI DÒNG 0Đ
    # Các dòng này là tháng đã đóng trước
    # =========================
    source_rows = [
        r
        for r in tracking_data["rows"]
        if int(r.get("display_tong_tien") or 0) > 0
    ]

    # =========================
    # TRA LỚP TỪ BẢNG STUDENT
    # =========================
    student_ids = []

    for r in source_rows:
        license_code = str(
            r.get("ma_hv") or ""
        ).strip()

        if (
            license_code
            and license_code not in student_ids
        ):
            student_ids.append(license_code)

    students_map = {}

    if student_ids:
        student_rows = (
            supabase.table(STUDENT_TABLE)
            .select("license,name,classroom")
            .in_("license", student_ids)
            .execute()
            .data
            or []
        )

        students_map = {
            str(
                s.get("license") or ""
            ).strip(): s
            for s in student_rows
        }

    monthly_fee = get_app_setting_int(
        "fees.tuition_fee",
        500000
    )

    export_rows = []

    for r in source_rows:
        license_code = str(
            r.get("ma_hv") or ""
        ).strip()

        student = students_map.get(
            license_code,
            {}
        )

        final_amount = int(
            r.get("display_tong_tien") or 0
        )

        raw_note = str(
            r.get("ghi_chu") or ""
        )

        normalized_note = remove_accents(
            raw_note
        ).lower()

        month_codes = (
            r.get("_month_codes")
            or []
        )

        # =========================
        # TÍNH HỌC PHÍ GỐC
        # =========================
        if "nua thang" in normalized_note:
            billable_months = 1

        elif (
            "dong 6 thang" in normalized_note
            or "tang 1" in normalized_note
        ):
            billable_months = 6

        elif "dong 3 thang" in normalized_note:
            billable_months = 3

        elif len(month_codes) == 7:
            # Đóng 6 tháng, tặng 1 tháng
            billable_months = 6

        else:
            billable_months = max(
                len(month_codes),
                1
            )

        base_fee = (
            monthly_fee
            * billable_months
        )

        # Tránh trường hợp dữ liệu thực thu
        # lớn hơn học phí gốc
        if final_amount > base_fee:
            base_fee = final_amount

        export_rows.append({
            "name": (
                student.get("name")
                or r.get("ho_ten")
                or ""
            ),

            "classroom": (
                student.get("classroom")
                or r.get("lop")
                or ""
            ),

            "base_fee": base_fee,

            "discount": (
                build_tracking_discount_text_web(
                    base_fee,
                    final_amount,
                    raw_note
                )
            ),

            "final_amount": final_amount,

            "note": (
                build_tracking_export_note_web(r)
            ),
        })

    # =========================
    # TẠO FILE EXCEL
    # =========================
    wb = Workbook()
    ws = wb.active
    ws.title = "Theo doi hoc phi"

    headers = [
        "STT",
        "Họ và tên",
        "Lớp",
        "Học phí",
        "Giảm giá",
        "Thành tiền",
        "Ghi chú",
    ]

    thin_side = Side(
        style="thin",
        color="CBD5E1"
    )

    medium_side = Side(
        style="medium",
        color="64748B"
    )

    thin_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side
    )

    total_border = Border(
        left=medium_side,
        right=medium_side,
        top=medium_side,
        bottom=medium_side
    )

    center = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    left = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True
    )

    right = Alignment(
        horizontal="right",
        vertical="center"
    )

    # =========================
    # TIÊU ĐỀ CHÍNH
    # =========================
    ws.merge_cells("A1:G1")

    ws["A1"] = (
        f"DANH SÁCH HỌC PHÍ "
        f"THÁNG {month_int:02d}/{year_int}"
    )

    ws["A1"].font = Font(
        bold=True,
        size=18,
        color="FFFFFF"
    )

    ws["A1"].fill = PatternFill(
        "solid",
        fgColor="0F2A4A"
    )

    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 32

    # =========================
    # THÔNG TIN BỘ LỌC
    # =========================
    ws.merge_cells("A2:G2")

    ws["A2"] = (
        f"Bộ lọc: Tháng "
        f"{month_int:02d}/{year_int}"
        f" | Số học viên: "
        f"{len(export_rows)}"
    )

    ws["A2"].font = Font(
        size=13,
        italic=True,
        color="475569"
    )

    ws["A2"].alignment = center
    ws.row_dimensions[2].height = 24

    # =========================
    # HEADER BẢNG
    # =========================
    header_row = 4

    for col_idx, header in enumerate(
        headers,
        start=1
    ):
        cell = ws.cell(
            row=header_row,
            column=col_idx,
            value=header
        )

        cell.font = Font(
            bold=True,
            size=13,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            "solid",
            fgColor="1D4ED8"
        )

        cell.alignment = center
        cell.border = thin_border

    ws.row_dimensions[header_row].height = 28

    # =========================
    # DỮ LIỆU
    # =========================
    data_start = header_row + 1

    for stt, item in enumerate(
        export_rows,
        start=1
    ):
        row_idx = data_start + stt - 1

        values = [
            stt,
            item["name"],
            item["classroom"],
            item["base_fee"],
            item["discount"],
            item["final_amount"],
            item["note"] or "—",
        ]

        for col_idx, value in enumerate(
            values,
            start=1
        ):
            cell = ws.cell(
                row=row_idx,
                column=col_idx,
                value=value
            )

            cell.font = Font(size=13)
            cell.border = thin_border

            if col_idx in [1, 3]:
                cell.alignment = center

            elif col_idx in [4, 6]:
                cell.alignment = right

            else:
                cell.alignment = left

        # Học phí
        ws.cell(
            row=row_idx,
            column=4
        ).number_format = '#,##0 "đ"'

        # Thành tiền
        ws.cell(
            row=row_idx,
            column=6
        ).number_format = '#,##0 "đ"'

        ws.row_dimensions[row_idx].height = 23

    # =========================
    # TỔNG CỘNG
    # =========================
    total_row = (
        data_start
        + len(export_rows)
    )

    ws.merge_cells(
        start_row=total_row,
        start_column=1,
        end_row=total_row,
        end_column=5
    )

    total_label = ws.cell(
        row=total_row,
        column=1,
        value="TỔNG CỘNG"
    )

    total_label.font = Font(
        bold=True,
        size=16
    )

    total_label.fill = PatternFill(
        "solid",
        fgColor="FEF3C7"
    )

    total_label.alignment = center

    total_value = ws.cell(
        row=total_row,
        column=6,
        value=sum(
            item["final_amount"]
            for item in export_rows
        )
    )

    total_value.font = Font(
        bold=True,
        size=16,
        color="B91C1C"
    )

    total_value.fill = PatternFill(
        "solid",
        fgColor="FEF3C7"
    )

    total_value.number_format = '#,##0 "đ"'
    total_value.alignment = right

    total_note = ws.cell(
        row=total_row,
        column=7,
        value=""
    )

    total_note.fill = PatternFill(
        "solid",
        fgColor="FEF3C7"
    )

    for col_idx in range(1, 8):
        ws.cell(
            row=total_row,
            column=col_idx
        ).border = total_border

    ws.row_dimensions[total_row].height = 30

    # =========================
    # ĐỘ RỘNG CỘT
    # =========================
    widths = {
        "A": 8,
        "B": 28,
        "C": 18,
        "D": 17,
        "E": 22,
        "F": 18,
        "G": 32,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    # =========================
    # THIẾT LẬP EXCEL
    # =========================
    ws.freeze_panes = "A5"

    ws.auto_filter.ref = (
        f"A4:G{max(total_row - 1, 4)}"
    )

    ws.sheet_view.showGridLines = False

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    ws.print_title_rows = "1:4"

    # =========================================================
    # SHEET 2: TỔNG KẾT TỰ CẬP NHẬT BẰNG CÔNG THỨC
    # Khi sửa Lớp hoặc Thành tiền ở sheet 1,
    # sheet Tổng kết sẽ tự tính lại.
    # =========================================================
    ws_summary = wb.create_sheet("Tổng kết")

    # Các lớp và mức Thành tiền đang có khi xuất file
    summary_classes = sorted({
        str(item.get("classroom") or "Chưa xếp lớp").strip()
        for item in export_rows
        if str(item.get("classroom") or "").strip()
    })

    summary_amounts = sorted({
        int(item.get("final_amount") or 0)
        for item in export_rows
        if int(item.get("final_amount") or 0) > 0
    })


    # =========================
    # SẮP XẾP LỚP
    # =========================
    def tracking_class_sort_key_web(value):
        normalized = remove_accents(
            str(value or "")
        ).lower()

        normalized = re.sub(
            r"\s+",
            "",
            normalized
        )

        order_map = {
            "2-4-6": 1,
            "246": 1,

            "3-5-7": 2,
            "357": 2,

            "7-cn": 3,
            "t7-cn": 3,
            "t7cn": 3,
            "thu7-chunhat": 3,
            "thu7chunhat": 3,

            "henho": 4,
            "hen-ho": 4,
        }

        return (
            order_map.get(normalized, 99),
            normalized
        )


    summary_classes = sorted(
        summary_classes,
        key=tracking_class_sort_key_web
    )


    # =========================
    # SẮP XẾP MỨC TIỀN
    # Ưu tiên giống mẫu của Ken
    # =========================
    preferred_amount_order = {
        450000: 1,
        500000: 2,
        1350000: 3,
        1200000: 4,
        225000: 5,
        250000: 6,
    }

    summary_amounts = sorted(
        summary_amounts,
        key=lambda value: (
            preferred_amount_order.get(value, 99),
            value
        )
    )


    # =========================
    # STYLE
    # =========================
    summary_thin_side = Side(
        style="thin",
        color="D9E2F3"
    )

    summary_border = Border(
        left=summary_thin_side,
        right=summary_thin_side,
        top=summary_thin_side,
        bottom=summary_thin_side
    )

    summary_center = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    summary_left = Alignment(
        horizontal="left",
        vertical="center"
    )

    summary_right = Alignment(
        horizontal="right",
        vertical="center"
    )

    summary_class_fill = PatternFill(
        "solid",
        fgColor="D9EAF7"
    )

    summary_grand_fill = PatternFill(
        "solid",
        fgColor="BDD7EE"
    )


    # =========================
    # TIÊU ĐỀ
    # =========================
    ws_summary.merge_cells("A1:B1")

    ws_summary["A1"] = (
        f"TỔNG KẾT HỌC PHÍ "
        f"THÁNG {month_int:02d}/{year_int}"
    )

    ws_summary["A1"].font = Font(
        bold=True,
        size=18,
        color="FFFFFF"
    )

    ws_summary["A1"].fill = PatternFill(
        "solid",
        fgColor="0F2A4A"
    )

    ws_summary["A1"].alignment = summary_center
    ws_summary.row_dimensions[1].height = 32


    ws_summary.merge_cells("A2:B2")

    ws_summary["A2"] = (
        "Tự cập nhật khi chỉnh Lớp hoặc "
        "Thành tiền trong sheet Theo doi hoc phi"
    )

    ws_summary["A2"].font = Font(
        size=13,
        italic=True,
        color="475569"
    )

    ws_summary["A2"].alignment = summary_center
    ws_summary.row_dimensions[2].height = 24


    # =========================
    # HEADER
    # =========================
    ws_summary["A4"] = "Lớp / Thành tiền"
    ws_summary["B4"] = "Count of Họ và tên"

    for col_idx in range(1, 3):
        cell = ws_summary.cell(
            row=4,
            column=col_idx
        )

        cell.font = Font(
            bold=True,
            size=13
        )

        cell.fill = summary_class_fill
        cell.border = summary_border
        cell.alignment = summary_center

    ws_summary.row_dimensions[4].height = 26


    # =========================
    # PHẠM VI DỮ LIỆU SHEET 1
    # Sheet 1:
    # B = Họ và tên
    # C = Lớp
    # F = Thành tiền
    # =========================
    data_first_row = 5
    data_last_row = max(
        total_row - 1,
        data_first_row
    )


    # =========================
    # GHI CÁC NHÓM LỚP
    # =========================
    current_row = 5

    for classroom in summary_classes:
        class_row = current_row

        # Tên lớp
        ws_summary.cell(
            row=class_row,
            column=1,
            value=classroom
        )

        # Tổng số HV của lớp có Thành tiền > 0
        ws_summary.cell(
            row=class_row,
            column=2,
            value=(
                f'=COUNTIFS('
                f"'Theo doi hoc phi'!$C${data_first_row}:$C${data_last_row},"
                f'A{class_row},'
                f"'Theo doi hoc phi'!$F${data_first_row}:$F${data_last_row},"
                f'">0"'
                f')'
            )
        )

        for col_idx in range(1, 3):
            cell = ws_summary.cell(
                row=class_row,
                column=col_idx
            )

            cell.font = Font(
                bold=True,
                size=13
            )

            cell.fill = summary_class_fill
            cell.border = summary_border

        ws_summary.cell(
            row=class_row,
            column=1
        ).alignment = summary_left

        ws_summary.cell(
            row=class_row,
            column=2
        ).alignment = summary_right

        ws_summary.row_dimensions[
            class_row
        ].height = 23

        current_row += 1

        # =========================
        # CÁC MỨC THÀNH TIỀN
        # =========================
        for amount in summary_amounts:
            amount_row = current_row

            ws_summary.cell(
                row=amount_row,
                column=1,
                value=amount
            )

            ws_summary.cell(
                row=amount_row,
                column=1
            ).number_format = '#,##0 "đ"'

            # Đếm học viên theo lớp + thành tiền
            ws_summary.cell(
                row=amount_row,
                column=2,
                value=(
                    f'=COUNTIFS('
                    f"'Theo doi hoc phi'!$C${data_first_row}:$C${data_last_row},"
                    f'$A${class_row},'
                    f"'Theo doi hoc phi'!$F${data_first_row}:$F${data_last_row},"
                    f'A{amount_row}'
                    f')'
                )
            )

            for col_idx in range(1, 3):
                cell = ws_summary.cell(
                    row=amount_row,
                    column=col_idx
                )

                cell.font = Font(size=13)
                cell.border = summary_border

            ws_summary.cell(
                row=amount_row,
                column=1
            ).alignment = summary_right

            ws_summary.cell(
                row=amount_row,
                column=2
            ).alignment = summary_right

            # Tạo dấu +/- để thu gọn các mức tiền
            ws_summary.row_dimensions[
                amount_row
            ].outlineLevel = 1

            ws_summary.row_dimensions[
                amount_row
            ].height = 22

            current_row += 1


    # =========================
    # GRAND TOTAL
    # =========================
    grand_total_row = current_row

    ws_summary.cell(
        row=grand_total_row,
        column=1,
        value="Grand Total"
    )

    # Đếm tất cả các dòng có Thành tiền > 0
    ws_summary.cell(
        row=grand_total_row,
        column=2,
        value=(
            f'=COUNTIF('
            f"'Theo doi hoc phi'!$F${data_first_row}:$F${data_last_row},"
            f'">0"'
            f')'
        )
    )

    for col_idx in range(1, 3):
        cell = ws_summary.cell(
            row=grand_total_row,
            column=col_idx
        )

        cell.font = Font(
            bold=True,
            size=14
        )

        cell.fill = summary_grand_fill
        cell.border = summary_border

    ws_summary.cell(
        row=grand_total_row,
        column=1
    ).alignment = summary_left

    ws_summary.cell(
        row=grand_total_row,
        column=2
    ).alignment = summary_right

    ws_summary.row_dimensions[
        grand_total_row
    ].height = 25


    # =========================
    # KÍCH THƯỚC VÀ HIỂN THỊ
    # =========================
    ws_summary.column_dimensions["A"].width = 27
    ws_summary.column_dimensions["B"].width = 22

    ws_summary.freeze_panes = "A5"
    ws_summary.sheet_view.showGridLines = False

    ws_summary.sheet_properties.outlinePr.summaryBelow = False
    ws_summary.sheet_properties.outlinePr.showOutlineSymbols = True

    ws_summary.page_setup.orientation = "portrait"
    ws_summary.page_setup.fitToWidth = 1
    ws_summary.page_setup.fitToHeight = 0

    ws_summary.print_title_rows = "1:4"


    # =========================
    # BẮT EXCEL TỰ TÍNH LẠI
    # =========================
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    # =========================
    # XUẤT FILE
    # =========================
    output = BytesIO()



    wb.save(output)
    output.seek(0)

    filename = (
        f"theo_doi_hoc_phi_"
        f"{year_int}_"
        f"{month_int:02d}.xlsx"
    )

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )

def get_exam_info_web(ky_thi):
    ky_thi = str(ky_thi or "").strip()

    if not ky_thi:
        return {}

    try:
        rows = supabase.table(EXAM_INFO_TABLE) \
            .select("*") \
            .eq("ky_thi", ky_thi) \
            .limit(1) \
            .execute().data or []

        info = rows[0] if rows else {}

        judges_raw = info.get("judges_json") if info else ""

        if isinstance(judges_raw, str):
            try:
                info["judges"] = json.loads(judges_raw or "[]")
            except:
                info["judges"] = []
        elif isinstance(judges_raw, list):
            info["judges"] = judges_raw
        else:
            info["judges"] = []

        return info

    except Exception as e:
        print("[GET EXAM INFO ERROR]", e)
        return {}

def validate_exam_info_form(form):
    exam_date = str(form.get("exam_date") or "").strip()
    exam_time = str(form.get("exam_time") or "").strip()
    venue = str(form.get("venue") or "").strip()

    supervisor_name = str(form.get("supervisor_name") or "").strip()
    supervisor_active = str(form.get("supervisor_active") or "on").strip().lower()
    supervisor_active = "on" if supervisor_active == "on" else "off"

    judges_json_raw = str(form.get("judges_json") or "[]").strip()

    errors = []

    if not exam_date:
        errors.append("Ken chưa nhập <b>Ngày thi</b>.")

    if not exam_time:
        errors.append("Ken chưa nhập <b>Giờ thi</b>.")

    if not venue:
        errors.append("Ken chưa nhập <b>Địa điểm thi</b>.")

    if not supervisor_name:
        errors.append("Ken chưa nhập <b>Giám sát</b>.")

    try:
        judges = json.loads(judges_json_raw or "[]")

        if not isinstance(judges, list):
            judges = []
    except:
        judges = []

    cleaned_judges = []

    for item in judges:
        if not isinstance(item, dict):
            continue

        name = str(item.get("name") or "").strip()
        active = str(item.get("active") or "on").strip().lower()
        active = "on" if active == "on" else "off"

        if name:
            cleaned_judges.append({
                "name": name,
                "active": active,
            })

    if not cleaned_judges:
        errors.append("Ken chưa nhập <b>Giám khảo 1</b>.")

    return {
        "ok": not errors,
        "errors": errors,
        "data": {
            "exam_date": exam_date,
            "exam_time": exam_time,
            "venue": venue,
            "supervisor_name": supervisor_name,
            "supervisor_active": supervisor_active,
            "judges": cleaned_judges,
        }
    }


def save_exam_info_web(ky_thi, year, quarter, data):
    ky_thi = str(ky_thi or "").strip()
    year = str(year or "").strip()
    quarter = str(quarter or "").strip().upper()

    if not ky_thi:
        return

    payload = {
        "ky_thi": ky_thi,
        "year": year,
        "quarter": quarter,
        "exam_date": data.get("exam_date", ""),
        "exam_time": data.get("exam_time", ""),
        "venue": data.get("venue", ""),
        "supervisor_name": data.get("supervisor_name", ""),
        "supervisor_active": data.get("supervisor_active", "on"),
        "judges_json": json.dumps(data.get("judges", []), ensure_ascii=False),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }

    existing = supabase.table(EXAM_INFO_TABLE) \
        .select("id") \
        .eq("ky_thi", ky_thi) \
        .limit(1) \
        .execute().data or []

    if existing:
        supabase.table(EXAM_INFO_TABLE) \
            .update(payload) \
            .eq("ky_thi", ky_thi) \
            .execute()
    else:
        payload["created_at"] = datetime.now(timezone.utc).isoformat()

        supabase.table(EXAM_INFO_TABLE) \
            .insert(payload) \
            .execute()


@app.post("/exam-list/save-info")
def exam_list_save_info():
    data = request.get_json(silent=True) or {}

    year = str(data.get("year") or "").strip()
    quarter = str(data.get("quarter") or "").strip().upper()

    if not year or not quarter:
        return jsonify({
            "ok": False,
            "message": "Thiếu năm hoặc quý thi."
        }), 400

    dan_quarters = ["L1", "L2", "MN", "MT", "MB", "QG"]

    if quarter in dan_quarters:
        ky_thi = f"{quarter}-{year}"
    else:
        ky_thi = f"{year}-{quarter}"

    form_like_data = {
        "exam_date": str(data.get("exam_date") or "").strip(),
        "exam_time": str(data.get("exam_time") or "").strip(),
        "venue": str(data.get("venue") or "").strip(),
        "supervisor_name": str(data.get("supervisor_name") or "").strip(),
        "supervisor_active": str(data.get("supervisor_active") or "on").strip(),
        "judges_json": json.dumps(data.get("judges") or [], ensure_ascii=False),
    }

    exam_info_check = validate_exam_info_form(form_like_data)

    if not exam_info_check["ok"]:
        return jsonify({
            "ok": False,
            "message": "Chưa đủ thông tin kỳ thi.",
            "errors": exam_info_check["errors"]
        }), 400

    try:
        save_exam_info_web(
            ky_thi,
            year,
            quarter,
            exam_info_check["data"]
        )

        return jsonify({
            "ok": True,
            "message": "Đã lưu thông tin kỳ thi lên Supabase.",
            "ky_thi": ky_thi
        })

    except Exception as e:
        print("[SAVE EXAM INFO API ERROR]", e)

        return jsonify({
            "ok": False,
            "message": f"Lỗi lưu thông tin kỳ thi: {e}"
        }), 500

def is_dan_exam_belt_web(belt):
    """
    True nếu cấp dự thi là Đẳng.
    VD: 1 Đẳng, 2 Đẳng...
    """
    belt = normalize_belt_name_web(belt)
    return "Đẳng" in belt


def build_exam_ma_quy_web(year, quarter, list_type="cap"):
    """
    Thi cấp:
        Q1, Q2, Q3, Q4  -> Q12026, Q22026...

    Thi đẳng:
        L1, L2, MN, MT, MB, QG -> L1-2026, MN-2026...
    """
    year = str(year or "").strip()
    quarter = str(quarter or "").strip().upper()
    list_type = str(list_type or "cap").strip().lower()

    dan_quarters = ["L1", "L2", "MN", "MT", "MB", "QG"]

    if list_type == "dan" or quarter in dan_quarters:
        return f"{quarter}-{year}"

    return f"{quarter}{year}"

def get_exam_list_rows_by_type_web(year, quarter, list_type="cap"):
    """
    list_type = cap  → chỉ lấy thi cấp: Cấp 9 đến Cấp 1
    list_type = dan  → chỉ lấy thi đẳng: 1 Đẳng trở lên
    """
    ma_quy = build_exam_ma_quy_web(year, quarter, list_type)

    year_str = str(year or "").strip()
    quarter_str = str(quarter or "").strip().upper()
    list_type_str = str(list_type or "cap").strip().lower()

    dan_quarters = ["L1", "L2", "MN", "MT", "MB", "QG"]

    if list_type_str == "dan" or quarter_str in dan_quarters:
        ky_thi = f"{quarter_str}-{year_str}"
    else:
        ky_thi = f"{year_str}-{quarter_str}"

    saved_result_map = {}

    try:
        saved_results = supabase.table(KETQUA_TABLE) \
            .select("ma_hv,ket_qua,so_thi,cap_dai_thi") \
            .eq("ky_thi", ky_thi) \
            .execute().data or []

        for item in saved_results:
            ma_hv_saved = str(item.get("ma_hv") or "").strip()

            if ma_hv_saved:
                saved_result_map[ma_hv_saved] = item

    except Exception as e:
        print("[GET SAVED EXAM RESULTS ERROR]", e)
        saved_result_map = {}

    fees = supabase.table(HOCPHI_TABLE) \
        .select("id,ma_hv,ma_quy,thoi_gian,coach_check_status,coach_check_note,coach_checked_by,coach_checked_by_name,coach_checked_at") \
        .eq("ma_quy", ma_quy) \
        .execute().data or []

    fee_map = {}

    for f in fees:
        ma_hv_fee = str(f.get("ma_hv") or "").strip()

        if not ma_hv_fee:
            continue

        old = fee_map.get(ma_hv_fee)

        if not old:
            fee_map[ma_hv_fee] = f
        else:
            old_time = str(old.get("thoi_gian") or "")
            new_time = str(f.get("thoi_gian") or "")

            if new_time > old_time:
                fee_map[ma_hv_fee] = f

    ids = []

    for x in fees:
        ma_hv = str(x.get("ma_hv") or "").strip()

        if ma_hv and ma_hv not in ids:
            ids.append(ma_hv)

    students = []

    if ids:
        students = supabase.table(STUDENT_TABLE) \
            .select("*") \
            .in_("license", ids) \
            .execute().data or []

    rows = []

    for s in students:
        current_belt = normalize_belt_name_web(s.get("belt"))
        license_code = str(s.get("license") or "").strip()
        saved_result = saved_result_map.get(license_code)

        # =========================
        # Nếu đã chốt kết quả rồi:
        # Lấy cấp thi đã lưu trong bảng ketqua để hiển thị lại lịch sử.
        # Không tính lại theo student.belt nữa, vì student.belt đã được nâng cấp sau khi Đạt.
        # =========================
        if saved_result:
            saved_exam_belt = normalize_belt_name_web(saved_result.get("cap_dai_thi"))

            # VD đã lưu Cấp 9 nghĩa là lúc đó học viên đang Cấp 10 và thi lên Cấp 9.
            display_current_belt = get_previous_belt_web(saved_exam_belt)

            cap_du_thi = saved_exam_belt
            current_belt_for_display = display_current_belt

        else:
            cap_du_thi = get_next_belt_web(current_belt)

            # Khóa chắc lỗi Cấp 1 -> 1 Đẳng
            if current_belt == "Cấp 1":
                cap_du_thi = "1 Đẳng"

            current_belt_for_display = current_belt

        s["belt"] = current_belt_for_display
        s["cap_du_thi"] = cap_du_thi
        s["ket_qua_mac_dinh"] = "Đạt"

        # Dữ liệu kết quả đã chốt trong bảng ketqua.
        # Bắt buộc có để exam_list/admin biết dòng nào đã lưu kết quả,
        # từ đó hiện badge "Đã lưu kết quả" thay vì còn combobox Đạt/Không đạt/Vắng.
        s["result_saved"] = bool(saved_result)
        s["saved_ket_qua"] = saved_result.get("ket_qua", "") if saved_result else ""
        s["saved_so_thi"] = saved_result.get("so_thi", "") if saved_result else ""
        s["saved_cap_thi"] = saved_result.get("cap_dai_thi", "") if saved_result else ""

        fee_check = fee_map.get(license_code, {}) or {}

        s["hocphi_id"] = fee_check.get("id", "")
        s["coach_check_status"] = fee_check.get("coach_check_status") or "Chưa KTra"
        s["coach_check_note"] = fee_check.get("coach_check_note") or ""
        s["coach_checked_by"] = fee_check.get("coach_checked_by") or ""
        s["coach_checked_by_name"] = fee_check.get("coach_checked_by_name") or ""
        s["coach_checked_at"] = fee_check.get("coach_checked_at") or ""
        is_dan = is_dan_exam_belt_web(cap_du_thi)

        if list_type == "dan" and is_dan:
            rows.append(s)

        elif list_type == "cap" and not is_dan:
            rows.append(s)

    return rows

@app.get('/exam-list')
def exam_list():
    year = request.args.get('year', str(datetime.now().year))
    quarter = request.args.get('quarter', f'Q{(datetime.now().month - 1)//3 + 1}').strip().upper()

    ky_thi = f"{year}-{quarter}"

    students = get_exam_list_rows_by_type_web(
        year=year,
        quarter=quarter,
        list_type="cap"
    )

    finalize_errors = session.pop("exam_finalize_errors", [])
    exam_info = get_exam_info_web(ky_thi)
    saved_count = len([x for x in students if x.get("result_saved")])
    all_saved = bool(students) and saved_count == len(students)

    return render_template(
        'exam_list.html',
        rows=students,
        year=year,
        quarter=quarter,
        ky_thi=ky_thi,
        finalize_errors=finalize_errors,
        exam_info=exam_info,
        list_type="cap",
        saved_count=saved_count,
        all_saved=all_saved
    )

@app.get('/dan-list')
def dan_list():
    year = request.args.get('year', str(datetime.now().year))
    quarter = request.args.get('quarter', 'L1').strip().upper()

    ky_thi = f"{quarter}-{year}"

    students = get_exam_list_rows_by_type_web(
        year=year,
        quarter=quarter,
        list_type="dan"
    )

    finalize_errors = session.pop("exam_finalize_errors", [])
    exam_info = get_exam_info_web(ky_thi)
    saved_count = len([x for x in students if x.get("result_saved")])
    all_saved = bool(students) and saved_count == len(students)

    return render_template(
        'dan_list.html',
        rows=students,
        year=year,
        quarter=quarter,
        ky_thi=ky_thi,
        finalize_errors=finalize_errors,
        exam_info=exam_info,
        saved_count=saved_count,
        all_saved=all_saved,
        list_type="dan"
    )

def normalize_excel_header_web(value):
    """
    Chuẩn hóa tên cột Excel:
    VD: "Mã HV" -> "ma hv"
    """
    value = remove_accents(str(value or "").strip()).lower()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def find_excel_col_index_web(headers, candidates):
    """
    Tìm vị trí cột theo nhiều tên có thể có.
    headers: list text đã chuẩn hóa
    candidates: list tên cột có thể gặp
    """
    normalized_candidates = [
        normalize_excel_header_web(x)
        for x in candidates
    ]

    for idx, header in enumerate(headers):
        if header in normalized_candidates:
            return idx

    return None


def extract_exam_number_web(raw):
    """
    Lấy số thi từ Excel.
    Nhận:
    - 12
    - 012
    - Cấp 5_12
    - Cap 5 - 12
    """
    raw = str(raw or "").strip()

    if not raw:
        return ""

    # Nếu Excel đọc số dạng 12.0 thì đưa về 12
    if re.fullmatch(r"\d+\.0", raw):
        raw = raw.split(".")[0]

    # Lấy cụm số cuối cùng trong chuỗi
    nums = re.findall(r"\d+", raw)

    if not nums:
        return ""

    return str(int(nums[-1]))


def normalize_exam_period_web(raw):
    """
    Chuẩn hóa kỳ thi từ Excel để so với trang hiện tại.

    Nhận các dạng:
    - Q12026
    - Q1-2026
    - Q1/2026
    - 2026-Q1
    - L1-2026
    - 2026-L1
    """
    raw = str(raw or "").strip().upper()

    if not raw:
        return ""

    raw = raw.replace("_", "-").replace("/", "-").replace(" ", "")

    # Thi cấp: Q12026 / Q1-2026 / Q1.2026
    m = re.search(r"Q([1-4])\D*(20\d{2})", raw)
    if m:
        return f"{m.group(2)}-Q{m.group(1)}"

    # Thi cấp: 2026-Q1 / 2026.Q1
    m = re.search(r"(20\d{2})\D*Q([1-4])", raw)
    if m:
        return f"{m.group(1)}-Q{m.group(2)}"

    # Thi đẳng: L1-2026 / MN-2026...
    dan_codes = "L1|L2|MN|MT|MB|QG"

    m = re.search(rf"({dan_codes})\D*(20\d{{2}})", raw)
    if m:
        return f"{m.group(1)}-{m.group(2)}"

    m = re.search(rf"(20\d{{2}})\D*({dan_codes})", raw)
    if m:
        return f"{m.group(2)}-{m.group(1)}"

    return raw


def current_exam_period_web(year, quarter, list_type):
    year = str(year or "").strip()
    quarter = str(quarter or "").strip().upper()
    list_type = str(list_type or "cap").strip().lower()

    dan_quarters = ["L1", "L2", "MN", "MT", "MB", "QG"]

    if list_type == "dan" or quarter in dan_quarters:
        return f"{quarter}-{year}"

    return f"{year}-{quarter}"




@app.post("/exam-list/import-numbers")
def exam_list_import_numbers():
    """
    Upload Excel để tự load Số thi vào danh sách đang lọc.

    Kiểm tra:
    1. Đúng năm/quý đang mở.
    2. Đúng Mã HV có trong danh sách hiện tại.
    3. Đúng Cấp dự thi nếu Excel có cột Cấp dự thi.
    4. Không trùng mã HV trong Excel.
    5. Không trùng số thi giữa các dòng được áp dụng.
    """
    year = request.form.get("year", str(datetime.now().year)).strip()
    quarter = request.form.get("quarter", "").strip().upper()
    list_type = request.form.get("list_type", "cap").strip().lower()

    back_period = current_exam_period_web(year, quarter, list_type)

    file = request.files.get("excel_file")

    if not file or not file.filename:
        return jsonify({
            "ok": False,
            "message": "Ken chưa chọn file Excel."
        }), 400

    filename = secure_filename(file.filename or "")
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext not in ["xlsx", "xlsm"]:
        return jsonify({
            "ok": False,
            "message": "File phải là Excel .xlsx hoặc .xlsm."
        }), 400

    try:
        current_rows = get_exam_list_rows_by_type_web(
            year=year,
            quarter=quarter,
            list_type=list_type
        )

        current_map = {}

        for r in current_rows:
            license_code = str(r.get("license") or "").strip()

            if not license_code:
                continue

            current_map[license_code] = {
                "license": license_code,
                "name": str(r.get("name") or "").strip(),
                "cap_du_thi": normalize_belt_name_web(r.get("cap_du_thi")),
            }

        if not current_map:
            return jsonify({
                "ok": False,
                "message": "Danh sách hiện tại chưa có học viên để cập nhật số thi."
            }), 400

        wb = load_workbook(file, data_only=True)
        ws = wb.active

        if ws.max_row < 2:
            return jsonify({
                "ok": False,
                "message": "File Excel chưa có dữ liệu."
            }), 400

        raw_headers = [
            ws.cell(row=1, column=col).value
            for col in range(1, ws.max_column + 1)
        ]

        headers = [
            normalize_excel_header_web(x)
            for x in raw_headers
        ]

        col_ma_hv = find_excel_col_index_web(headers, [
            "Mã HV", "Ma HV",
            "Mã hội viên", "Ma hoi vien",
            "Mã học viên", "Ma hoc vien",
            "Mã võ sinh", "Ma vo sinh",
            "Mã thi sinh", "Ma thi sinh",
            "Mã thí sinh", "Ma thi sinh",
            "Mã dự thi", "Ma du thi",
            "License",
            "Student code",
            "Student ID",
            "ID",
            "ma_hv",
            "mahv",
            "license"
        ])

        col_so_thi = find_excel_col_index_web(headers, [
            "Số thi", "So thi",
            "Số báo danh", "So bao danh",
            "Số dự thi", "So du thi",
            "Số thứ tự thi", "So thu tu thi",
            "Mã số thi", "Ma so thi",
            "Mã số dự thi", "Ma so du thi",
            "SBD",
            "Exam number",
            "Exam No",
            "Candidate number",
            "Contest number",
            "Bib number",
            "so_thi",
            "sothi",
            "so_bao_danh",
            "sbd"
        ])

        col_cap_du_thi = find_excel_col_index_web(headers, [
            "Cấp dự thi", "Cap du thi",
            "Cấp thi", "Cap thi",
            "Cấp", "Cap",
            "Đẳng dự thi", "Dang du thi",
            "Đẳng thi", "Dang thi",
            "Belt",
            "Exam belt",
            "cap_du_thi",
            "cap_thi"
        ])

        col_ky_thi = find_excel_col_index_web(headers, [
            "Quý thi", "Quy thi",
            "Kỳ thi", "Ky thi",
            "Quý", "Quy",
            "Lần thi", "Lan thi",
            "Mã quý", "Ma quy",
            "Mã kỳ thi", "Ma ky thi",
            "ma_quy",
            "ky_thi"
        ])

        if col_ky_thi is None or col_ma_hv is None or col_cap_du_thi is None or col_so_thi is None:
            return jsonify({
                "ok": False,
                "message": "Excel bắt buộc phải có đủ các cột: Quý thi, Mã HV/Mã thí sinh, Cấp dự thi, Số thi/Số báo danh.",
                "errors": [
                    "Thiếu cột Quý thi." if col_ky_thi is None else "",
                    "Thiếu cột Mã HV/Mã thí sinh." if col_ma_hv is None else "",
                    "Thiếu cột Cấp dự thi." if col_cap_du_thi is None else "",
                    "Thiếu cột Số thi/Số báo danh." if col_so_thi is None else "",
                ],
                "numbers": {}
            }), 400

        numbers = {}
        errors = []
        outside_students = []
        seen_license = set()
        seen_number = {}

        for row_idx in range(2, ws.max_row + 1):
            values = [
                ws.cell(row=row_idx, column=col).value
                for col in range(1, ws.max_column + 1)
            ]

            ma_hv = str(values[col_ma_hv] or "").strip()
            so_thi = extract_exam_number_web(values[col_so_thi])

            if not ma_hv and not so_thi:
                continue

            excel_period = normalize_exam_period_web(values[col_ky_thi])
            excel_cap = normalize_belt_name_web(values[col_cap_du_thi])

            if not excel_period:
                errors.append(f"Dòng {row_idx}: thiếu Quý thi.")
                continue

            if not ma_hv:
                errors.append(f"Dòng {row_idx}: thiếu Mã HV/Mã thí sinh.")
                continue

            if not excel_cap:
                errors.append(f"{ma_hv}: thiếu Cấp dự thi.")
                continue

            if not so_thi:
                errors.append(f"{ma_hv}: thiếu Số thi/Số báo danh.")
                continue

            if ma_hv in seen_license:
                errors.append(f"{ma_hv}: bị trùng Mã HV trong file Excel.")
                continue

            seen_license.add(ma_hv)

            # =========================
            # KIỂM TRA TRÙNG SỐ THI TRÊN TOÀN BỘ FILE EXCEL
            # Làm trước khi kiểm tra mã có trong hệ thống hay không.
            # Vì VĐV ngoài CLB vẫn có thể bị trùng số thi với học viên của mình.
            # =========================
            excel_number_key = f"{excel_cap}_{so_thi}"

            if excel_number_key in seen_number:
                errors.append(
                    f"Số thi {excel_cap}_{so_thi} bị trùng giữa {seen_number[excel_number_key]} và {ma_hv}."
                )
                continue

            seen_number[excel_number_key] = ma_hv

            current_student = current_map.get(ma_hv)

            if not current_student:
                outside_students.append(
                    f"{ma_hv}: không có trong danh sách đang lọc, có thể là VĐV/CLB khác thi chung."
                )
                continue

            if excel_period != back_period:
                errors.append(
                    f"{ma_hv}: sai Quý thi. Excel là {excel_period}, hệ thống là {back_period}."
                )
                continue

            system_cap = current_student["cap_du_thi"]

            if excel_cap != system_cap:
                errors.append(
                    f"{ma_hv}: sai Cấp dự thi. Excel là {excel_cap}, hệ thống là {system_cap}."
                )
                continue

            numbers[ma_hv] = so_thi

        if errors:
            return jsonify({
                "ok": False,
                "message": "File Excel có lỗi cần xử lý trước.",
                "errors": errors,
                "outside_students": outside_students,
                "numbers": {}
            }), 400

        if not numbers:
            return jsonify({
                "ok": False,
                "message": "Không có dòng nào thuộc danh sách hiện tại để cập nhật số thi.",
                "errors": errors,
                "outside_students": outside_students,
                "numbers": {}
            }), 400

        return jsonify({
            "ok": True,
            "message": f"Đã dò xong Excel. Có thể cập nhật {len(numbers)} số thi.",
            "applied_count": len(numbers),
            "outside_count": len(outside_students),
            "numbers": numbers,
            "outside_students": outside_students,
            "require_confirm": len(outside_students) > 0,
            "errors": []
        })

    except Exception as e:
        print("[IMPORT EXAM NUMBERS ERROR]", e)

        return jsonify({
            "ok": False,
            "message": f"Lỗi đọc Excel: {e}"
        }), 500

@app.post('/exam-list/finalize')
def exam_list_finalize():
    year = request.form.get("year", str(datetime.now().year)).strip()
    quarter = request.form.get("quarter", f'Q{(datetime.now().month - 1)//3 + 1}').strip().upper()
    list_type = request.form.get("list_type", "cap").strip()

    dan_quarters = ["L1", "L2", "MN", "MT", "MB", "QG"]

    if list_type == "dan" or quarter in dan_quarters:
        ky_thi = f"{quarter}-{year}"
    else:
        ky_thi = f"{year}-{quarter}"

    back_endpoint = "dan_list" if list_type == "dan" else "exam_list"

    exam_info_check = validate_exam_info_form(request.form)

    if not exam_info_check["ok"]:
        session["exam_finalize_errors"] = [
            "<b>Chưa đủ thông tin kỳ thi</b><br>" + "<br>".join(exam_info_check["errors"])
        ]

        return redirect(url_for(
            back_endpoint,
            year=year,
            quarter=quarter
        ))

    current_ky_value = parse_ky_thi_web(ky_thi)

    ma_hv_list = request.form.getlist("ma_hv")

    payloads = []
    pass_updates = []
    errors = []
    seen_ma_hv = set()

    for ma_hv in ma_hv_list:
        ma_hv = str(ma_hv or "").strip()

        if not ma_hv:
            continue

        if ma_hv in seen_ma_hv:
            errors.append(format_result_error_web(ma_hv, "", "Mã hội viên bị trùng trong danh sách đang chốt."))
            continue

        seen_ma_hv.add(ma_hv)

        ho_ten = request.form.get(f"name_{ma_hv}", "").strip()
        cap_du_thi = request.form.get(f"cap_du_thi_{ma_hv}", "").strip()
        so_thi_raw = request.form.get(f"so_thi_{ma_hv}", "").strip()
        ket_qua = request.form.get(f"ket_qua_{ma_hv}", "Đạt").strip()

        so_thi = ""
        if so_thi_raw:
            exam_prefix = str(get_app_setting("exam.exam_number_prefix", "Cấp_") or "")

            # Cách dùng trong Setup:
            # "Cấp_"   -> Cấp_01
            # "{cap}_" -> Cấp 5_01
            # ""       -> 01
            exam_prefix = exam_prefix.replace("{cap}", cap_du_thi)

            so_thi = f"{exam_prefix}{so_thi_raw}"

        # 1) Lấy thông tin học viên hiện tại
        student_rows = safe_rows(STUDENT_TABLE, "*", license=ma_hv)
        student = student_rows[0] if student_rows else {}

        if not student:
            errors.append(format_result_error_web(ma_hv, ho_ten, "Không tìm thấy học viên trong bảng student."))
            continue

        current_student_belt = str(student.get("belt") or "").strip()

        # 2) Lấy toàn bộ lịch sử thi cấp của học viên
        all_results = safe_rows(KETQUA_TABLE, "*", ma_hv=ma_hv)

        # 3) Mỗi quý chỉ được chốt 1 lần / 1 mã HV
        same_quarter_results = [
            r for r in all_results
            if str(r.get("ky_thi") or "").strip().upper() == ky_thi.upper()
        ]

        if same_quarter_results:
            errors.append(
                f"""
                <b>{ma_hv} - {ho_ten}</b><br>
                Lý do: Học viên này đã có kết quả trong kỳ <b>{ky_thi}</b>.<br>
                Cách xử lý: Nếu Ken muốn chốt lại, hãy vào tab <b>Kết quả thi cấp</b>,
                click phải dòng kết quả cũ và xóa trước. Sau đó quay lại chốt lại.
                """
            )
            continue

        # 4) Chỉ xét các kết quả trước quý hiện tại
        previous_results = [
            r for r in all_results
            if parse_ky_thi_web(r.get("ky_thi")) < current_ky_value
        ]

        future_or_invalid_results = [
            r for r in all_results
            if parse_ky_thi_web(r.get("ky_thi")) > current_ky_value
        ]

        if future_or_invalid_results:
            errors.append(
                f"""
                <b>{ma_hv} - {ho_ten}</b><br>
                Lý do: Học viên này đã có kết quả ở quý sau <b>{ky_thi}</b>.<br>
                Cách xử lý: Không thể chốt ngược lại kỳ cũ khi đã có dữ liệu kỳ sau.
                Ken cần kiểm tra lại lịch sử trong tab <b>Kết quả thi cấp</b>.
                """
            )
            continue

        # 5) Kiểm tra đúng trình tự cấp thi
        expected_belt = get_expected_exam_belt_web(current_student_belt, previous_results)

        if cap_du_thi != expected_belt:
            errors.append(
                f"""
                <b>{ma_hv} - {ho_ten}</b><br>
                Lý do: Sai trình tự cấp thi.<br>
                Cấp được phép thi hiện tại: <b>{expected_belt}</b><br>
                Cấp đang chốt: <b>{cap_du_thi}</b><br>
                Quy tắc: Nếu quý trước <b>Đạt</b> thì quý sau được thi cấp tiếp theo.
                Nếu <b>Không đạt</b> hoặc <b>Vắng</b> thì quý sau phải thi lại cấp đó.
                """
            )
            continue

        ghi_chu = ""

        if ket_qua == "Vắng":
            ghi_chu = "Vắng thi"
        elif ket_qua == "Không đạt":
            ghi_chu = "Không đạt"

        payloads.append({
            "ky_thi": ky_thi,
            "ma_hv": ma_hv,
            "ho_ten": ho_ten,
            "cap_dai_thi": cap_du_thi,
            "so_thi": so_thi,
            "ket_qua": ket_qua,
            "ghi_chu": ghi_chu,
        })

        # 6) Nếu đạt thì cập nhật cấp mới cho student sau khi insert thành công
        if ket_qua == "Đạt":
            pass_updates.append({
                "ma_hv": ma_hv,
                "new_belt": cap_du_thi,
            })

    # Nếu có lỗi thì không chốt bất kỳ ai, tránh dữ liệu nửa đúng nửa sai
    if errors:
        session["exam_finalize_errors"] = errors

        return redirect(url_for(
            back_endpoint,
            year=year,
            quarter=quarter
        ))

    if not payloads:
        flash("Không có dữ liệu hợp lệ để chốt kết quả.")
        return redirect(url_for(
            back_endpoint,
            year=year,
            quarter=quarter
        ))

    try:
        # Lưu thông tin kỳ thi trước khi chốt kết quả
        save_exam_info_web(ky_thi, year, quarter, exam_info_check["data"])

        # Không xóa kết quả cũ nữa.
        # Vì mỗi mã HV / mỗi quý chỉ được chốt 1 lần.
        supabase.table(KETQUA_TABLE).insert(payloads).execute()

        # Cập nhật cấp mới cho học viên đạt
        for item in pass_updates:
            supabase.table(STUDENT_TABLE) \
                .update({"belt": item["new_belt"]}) \
                .eq("license", item["ma_hv"]) \
                .execute()

        flash(
            f"Đã chốt kết quả kỳ thi {ky_thi}: "
            f"{len(payloads)} võ sinh. "
            f"Đã cập nhật cấp mới cho {len(pass_updates)} võ sinh đạt."
        )

    except Exception as e:
        print("[FINALIZE EXAM ERROR]", e)
        flash(f"Lỗi chốt kết quả: {e}")

        return redirect(url_for(
            back_endpoint,
            year=year,
            quarter=quarter
        ))

    return redirect(url_for("results", year=year, quarter=quarter))

@app.get('/api/student/<license_code>')
def api_student(license_code):
    data = supabase.table(STUDENT_TABLE).select('*').eq('license', license_code).limit(1).execute().data
    return jsonify(data[0] if data else {})

STUDENT_TABLE = "student"
COACH_TABLE = "coaches"
HOCPHI_TABLE = "hocphi"
KETQUA_TABLE = "ketqua"
HOATDONG_TABLE = "hoatdong"

RESTORE_SECURITY_TABLE = "restore_security"
RESTORE_HISTORY_TABLE = "restore_history"

BACKUP_MANIFEST_TABLE = "backup_manifests"
BACKUP_MANIFEST_ITEM_TABLE = "backup_manifest_items"

ACTIVITY_EVENTS_TABLE = "activity_events"
NOTIFICATION_TABLE = "notifications"
PAYMENT_SETTINGS_TABLE = "payment_settings"
APP_SETTINGS_TABLE = "app_settings"
NOTIFICATION_READ_TABLE = "notification_reads"
EXAM_INFO_TABLE = "exam_infos"
BELT_FLOW_WEB = [
    "Cấp 10", "Cấp 9", "Cấp 8", "Cấp 7", "Cấp 6",
    "Cấp 5", "Cấp 4", "Cấp 3", "Cấp 2", "Cấp 1",
    "1 Đẳng", "2 Đẳng", "3 Đẳng", "4 Đẳng", "5 Đẳng",
    "6 Đẳng", "7 Đẳng", "8 Đẳng", "9 Đẳng", "10 Đẳng"
]

def normalize_belt_name_web(belt):
    raw = str(belt or "").strip()

    if not raw:
        return ""

    # Chuẩn hóa Unicode: xử lý lỗi Cấp khác Cấp
    raw = unicodedata.normalize("NFC", raw)

    # Chuẩn hóa khoảng trắng
    raw = re.sub(r"\s+", " ", raw).strip()

    # Bản bỏ dấu để nhận mọi kiểu: Cấp, Cấp, cap, CẤP
    text_no_accent = remove_accents(raw).lower()
    text_no_accent = re.sub(r"\s+", " ", text_no_accent).strip()

    # Chuẩn hóa Cấp: Cấp 1, Cấp 1, cap 1, CAP 1 -> Cấp 1
    cap_match = re.search(r"cap\s*(\d+)", text_no_accent)
    if cap_match:
        return f"Cấp {int(cap_match.group(1))}"

    # Chuẩn hóa Đẳng: 1 đẳng, 1 Đẳng, 1 dang, 1 DANG -> 1 Đẳng
    dang_match = re.search(r"(\d+)\s*dang", text_no_accent)
    if dang_match:
        return f"{int(dang_match.group(1))} Đẳng"

    return raw

def supabase_public_url(bucket, filename):
    filename = str(filename or "").strip().lstrip("/")
    return f"{SUPABASE_URL}/storage/v1/object/public/{bucket}/{filename}"


def supabase_student_photo_public_url(filename):
    filename = str(filename or "").strip().lstrip("/")
    return supabase_public_url(STUDENT_PHOTO_BUCKET, filename)


def upload_club_asset_to_supabase(file_storage, folder="club-info"):
    """
    Upload ảnh logo/HLV/thông tin CLB lên Supabase Storage.
    """
    if not file_storage or not file_storage.filename:
        return ""

    filename = secure_filename(file_storage.filename)
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext not in ALLOWED_PHOTO_EXTENSIONS:
        raise ValueError("Ảnh chỉ nhận file JPG, JPEG, PNG hoặc WEBP.")

    storage_path = f"{folder}/{datetime.now().strftime('%Y%m%d%H%M%S%f')}.{ext}"
    content = file_storage.read()

    content_type_map = {
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "png": "image/png",
        "webp": "image/webp",
    }

    try:
        supabase.storage.from_(CLUB_ASSET_BUCKET).upload(
            storage_path,
            content,
            {
                "content-type": content_type_map.get(ext, "application/octet-stream"),
                "upsert": "true"
            }
        )
    except Exception:
        supabase.storage.from_(CLUB_ASSET_BUCKET).update(
            storage_path,
            content,
            {
                "content-type": content_type_map.get(ext, "application/octet-stream"),
                "upsert": "true"
            }
        )

    return supabase_public_url(CLUB_ASSET_BUCKET, storage_path)

def prepare_student_photo_for_upload(
    file_storage,
    target_width=780,
    target_height=1040,
    min_target_kb=300,
    max_target_kb=480
):
    """
    Chuẩn hóa ảnh học viên theo tỷ lệ 3x4 và nén thích ứng.

    Quy tắc:
    - Ảnh xuất ra dưới 300 KB: giữ nguyên mức chất lượng cao, không nén thêm.
    - Ảnh từ 300 đến 480 KB: giữ nguyên.
    - Ảnh trên 480 KB: tự giảm chất lượng từng bước.
    - Không cố ép ảnh nhỏ phải tăng lên 300 KB.
    - Chỉ giảm độ phân giải thêm khi giảm chất lượng vẫn chưa đủ.
    """

    img = Image.open(file_storage.stream)
    img = ImageOps.exif_transpose(img)
    img = img.convert("RGB")

    # =========================
    # 1. CROP ĐÚNG TỶ LỆ 3x4
    # =========================
    src_w, src_h = img.size
    target_ratio = target_width / target_height
    src_ratio = src_w / src_h

    if src_ratio > target_ratio:
        # Ảnh quá ngang: cắt hai bên
        new_w = int(src_h * target_ratio)
        left = max(0, (src_w - new_w) // 2)

        img = img.crop((
            left,
            0,
            left + new_w,
            src_h
        ))

    elif src_ratio < target_ratio:
        # Ảnh quá dọc: ưu tiên giữ phần đầu và khuôn mặt
        new_h = int(src_w / target_ratio)
        extra_h = src_h - new_h

        top = max(0, int(extra_h * 0.10))

        if top + new_h > src_h:
            top = max(0, src_h - new_h)

        img = img.crop((
            0,
            top,
            src_w,
            top + new_h
        ))

    # =========================
    # 2. CHỈ THU NHỎ, KHÔNG PHÓNG TO
    # =========================
    crop_w, crop_h = img.size

    if crop_w > target_width or crop_h > target_height:
        img = img.resize(
            (target_width, target_height),
            Image.Resampling.LANCZOS
        )

    # Không GaussianBlur vì dễ làm ảnh mặt bị mềm/mờ.

    min_target_bytes = int(min_target_kb * 1024)
    max_target_bytes = int(max_target_kb * 1024)

    def save_jpeg(image, quality):
        output = BytesIO()

        image.save(
            output,
            format="JPEG",
            quality=quality,
            optimize=True,
            progressive=True,
            subsampling="4:2:0"
        )

        return output.getvalue()

    # =========================
    # 3. THỬ CHẤT LƯỢNG CAO TRƯỚC
    # =========================
    content = save_jpeg(img, 95)

    # Ảnh dưới 300 KB hoặc nằm trong 300–480 KB:
    # giữ nguyên bản quality 95, không giảm thêm.
    if len(content) <= max_target_bytes:
        return content

    # =========================
    # 4. ẢNH TRÊN 480 KB:
    # GIẢM CHẤT LƯỢNG TỪ TỪ
    # =========================
    quality_levels = [
        93,
        91,
        89,
        87,
        85,
        83,
        81,
        79
    ]

    best_content = content

    for quality in quality_levels:
        test_content = save_jpeg(img, quality)

        # Luôn giữ bản nhỏ nhất đã thử
        if len(test_content) < len(best_content):
            best_content = test_content

        # Đạt dưới 480 KB thì dừng ngay
        if len(test_content) <= max_target_bytes:
            return test_content

    # =========================
    # 5. VẪN TRÊN 480 KB:
    # GIẢM NHẸ ĐỘ PHÂN GIẢI
    # =========================
    resize_levels = [
        (720, 960),
        (660, 880),
        (600, 800)
    ]

    original_img = img

    for width, height in resize_levels:
        smaller_img = original_img.resize(
            (width, height),
            Image.Resampling.LANCZOS
        )

        for quality in [88, 85, 82, 79]:
            test_content = save_jpeg(smaller_img, quality)

            if len(test_content) < len(best_content):
                best_content = test_content

            if len(test_content) <= max_target_bytes:
                return test_content

    # Trường hợp ảnh cực nhiều chi tiết:
    # trả về bản nhỏ nhất đã tạo được.
    return best_content

def get_student_photo_url(license_code):
    license_code = str(license_code or "").strip()

    if not license_code:
        return ""

    try:
        rows = supabase.table(STUDENT_TABLE) \
            .select("photo_url") \
            .eq("license", license_code) \
            .limit(1) \
            .execute().data or []

        if rows and rows[0].get("photo_url"):
            return rows[0].get("photo_url")

    except Exception as e:
        print("[GET STUDENT PHOTO URL ERROR]", e)

    return ""

def rename_student_photo_file(old_license, new_license):
    old_license = str(old_license or "").strip()
    new_license = str(new_license or "").strip()

    if not old_license or not new_license or old_license == new_license:
        return

    for ext in ["jpg", "jpeg", "png", "webp"]:
        old_path = STUDENT_PHOTO_DIR / f"{old_license}.{ext}"
        new_path = STUDENT_PHOTO_DIR / f"{new_license}.{ext}"

        if old_path.exists():
            if new_path.exists():
                new_path.unlink()
            old_path.rename(new_path)
            break

@app.get("/api/student-detail/<license_code>")
def api_student_detail(license_code):
    license_code = str(license_code or "").strip()

    student_rows = safe_rows(STUDENT_TABLE, "*", license=license_code)
    student = student_rows[0] if student_rows else {}

    results = safe_rows(KETQUA_TABLE, "*", ma_hv=license_code)
    activities = safe_rows(HOATDONG_TABLE, "*", ma_hv=license_code)

    results.sort(key=lambda r: str(r.get("ky_thi") or ""), reverse=True)
    activities.sort(key=lambda r: str(r.get("thoi_gian") or ""), reverse=True)

    return jsonify({
        "student": student,
        "results": results,
        "activities": activities,
        "photo_url": get_student_photo_url(license_code)
    })

@app.post("/api/student-photo-upload/<license_code>")
def api_student_photo_upload(license_code):
    license_code = str(license_code or "").strip()

    if not license_code:
        return jsonify({"ok": False, "message": "Thiếu mã hội viên"}), 400

    if "photo" not in request.files:
        return jsonify({"ok": False, "message": "Chưa chọn hình"}), 400

    file = request.files["photo"]

    if not file or not file.filename:
        return jsonify({"ok": False, "message": "File hình không hợp lệ"}), 400

    filename = secure_filename(file.filename)
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext not in ALLOWED_PHOTO_EXTENSIONS:
        return jsonify({
            "ok": False,
            "message": "Chỉ cho phép ảnh JPG, JPEG, PNG hoặc WEBP"
        }), 400

    try:
        # Luôn xử lý và lưu ảnh học viên thành JPG chuẩn 3x4
        storage_path = f"{license_code}.jpg"
        content = prepare_student_photo_for_upload(file)

        # Xóa toàn bộ ảnh cũ khác đuôi để mỗi học viên chỉ còn 1 ảnh
        old_paths = [
            f"{license_code}.jpg",
            f"{license_code}.jpeg",
            f"{license_code}.png",
            f"{license_code}.webp",
        ]

        try:
            supabase.storage.from_(STUDENT_PHOTO_BUCKET).remove(old_paths)
        except Exception:
            pass

        # Upload mới / ghi đè ảnh cũ
        try:
            supabase.storage.from_(STUDENT_PHOTO_BUCKET).upload(
                storage_path,
                content,
                {
                    "content-type": "image/jpeg",
                    "upsert": "true"
                }
            )
        except Exception:
            supabase.storage.from_(STUDENT_PHOTO_BUCKET).update(
                storage_path,
                content,
                {
                    "content-type": "image/jpeg",
                    "upsert": "true"
                }
            )

        photo_url = supabase_student_photo_public_url(storage_path)

        # Thêm version chống cache để web hiện ảnh mới ngay
        photo_url_with_version = f"{photo_url}?v={int(datetime.now().timestamp())}"

        # Lưu link ảnh vào bảng student
        supabase.table(STUDENT_TABLE) \
            .update({"photo_url": photo_url_with_version}) \
            .eq("license", license_code) \
            .execute()

        return jsonify({
            "ok": True,
            "photo_url": photo_url_with_version
        })

    except Exception as e:
        print("[UPLOAD STUDENT PHOTO SUPABASE ERROR]", repr(e))
        return jsonify({"ok": False, "message": str(e)}), 500

def sync_student_profile_to_related_tables(license_code):
    """
    Đồng bộ thông tin học viên từ bảng student sang các bảng liên quan.
    student là bảng gốc.
    """

    license_code = str(license_code or "").strip()

    if not license_code:
        return

    try:
        rows = supabase.table(STUDENT_TABLE) \
            .select("*") \
            .eq("license", license_code) \
            .limit(1) \
            .execute().data or []

        if not rows:
            return

        s = rows[0]

        name = str(s.get("name") or "").strip()
        birthdate = str(s.get("birthdate") or "").strip()
        gender = str(s.get("gender") or "").strip()
        classroom = str(s.get("classroom") or "").strip()
        timeclass = str(s.get("timeclass") or "").strip()
        clup = str(s.get("clup") or "").strip()
        phone = str(s.get("phonenumber") or "").strip()

        # =========================
        # 1. HỌC PHÍ
        # hocphi đang lưu snapshot: họ tên, ngày sinh, giới tính, lớp, ca
        # =========================
        try:
            supabase.table(HOCPHI_TABLE) \
                .update({
                    "ho_ten": name,
                    "ngay_sinh": birthdate,
                    "gioi_tinh": gender,
                    "lop": classroom,
                    "ca": timeclass,
                }) \
                .eq("ma_hv", license_code) \
                .execute()
        except Exception as e:
            print("[SYNC HOCPHI PROFILE ERROR]", e)

        # =========================
        # 2. KẾT QUẢ THI CẤP
        # ketqua đang lưu ho_ten
        # =========================
        try:
            supabase.table(KETQUA_TABLE) \
                .update({
                    "ho_ten": name,
                }) \
                .eq("ma_hv", license_code) \
                .execute()
        except Exception as e:
            print("[SYNC KETQUA PROFILE ERROR]", e)

        # =========================
        # 3. HOẠT ĐỘNG
        # hoatdong đang lưu ho_ten
        # =========================
        try:
            supabase.table(HOATDONG_TABLE) \
                .update({
                    "ho_ten": name,
                }) \
                .eq("ma_hv", license_code) \
                .execute()
        except Exception as e:
            print("[SYNC HOATDONG PROFILE ERROR]", e)

        # =========================
        # 4. THÔNG BÁO GỬI RIÊNG
        # notifications đang lưu target_name
        # =========================
        try:
            supabase.table(NOTIFICATION_TABLE) \
                .update({
                    "target_name": name,
                }) \
                .eq("target_license", license_code) \
                .execute()
        except Exception as e:
            print("[SYNC NOTIFICATION TARGET NAME ERROR]", e)

    except Exception as e:
        print("[SYNC STUDENT PROFILE ERROR]", e)

@app.post("/api/student-field-update")
def api_student_field_update():
    data = request.get_json(silent=True) or {}

    license_code = str(data.get("license") or "").strip()
    field = str(data.get("field") or "").strip()
    value = str(data.get("value") or "").strip()

    allowed_fields = {
        "name",
        "birthdate",
        "gender",
        "phonenumber",
        "address",
        "telegram_id",
        "license",
        "classroom",
        "timeclass",
        "clup",
        "belt",
        "family",
        "active",
    }

    if not license_code or not field:
        return jsonify({
            "ok": False,
            "message": "Thiếu mã hội viên hoặc trường cần sửa"
        }), 400

    if field not in allowed_fields:
        return jsonify({
            "ok": False,
            "message": "Trường này không được phép sửa"
        }), 400

    try:
        # =========================
        # ĐỔI MÃ HỘI VIÊN TOÀN HỆ THỐNG
        # =========================
        if field == "license":
            old_license = license_code.strip()
            new_license = value.strip()

            if not new_license:
                return jsonify({
                    "ok": False,
                    "message": "Mã hội viên không được trống"
                }), 400

            if new_license == old_license:
                sync_student_profile_to_related_tables(old_license)

                return jsonify({
                    "ok": True,
                    "new_license": new_license,
                    "message": "Mã hội viên không thay đổi. Đã kiểm tra đồng bộ dữ liệu liên quan."
                })

            # Kiểm tra mã mới có trùng không
            duplicated = supabase.table(STUDENT_TABLE) \
                .select("license,name,birthdate") \
                .eq("license", new_license) \
                .limit(1) \
                .execute().data or []

            if duplicated:
                old = duplicated[0]

                return jsonify({
                    "ok": False,
                    "message": (
                        f"Mã HV {new_license} đã tồn tại cho học viên "
                        f"{old.get('name', '')} - {old.get('birthdate', '')}. "
                        f"Không thể đổi sang mã bị trùng."
                    )
                }), 400

            # Lấy học viên hiện tại
            student_rows = supabase.table(STUDENT_TABLE) \
                .select("*") \
                .eq("license", old_license) \
                .limit(1) \
                .execute().data or []

            if not student_rows:
                return jsonify({
                    "ok": False,
                    "message": f"Không tìm thấy mã HV cũ: {old_license}"
                }), 404

            old_student = student_rows[0]

            student_update = {
                "license": new_license
            }

            old_username = str(old_student.get("portal_username") or "").strip()

            # Nếu username portal đang là mã cũ thì đổi sang mã mới
            if not old_username or old_username == old_license:
                student_update["portal_username"] = new_license

            # 1. Đổi mã trong student trước
            supabase.table(STUDENT_TABLE) \
                .update(student_update) \
                .eq("license", old_license) \
                .execute()

            # 2. Đổi mã trong các bảng liên quan
            supabase.table(HOCPHI_TABLE) \
                .update({"ma_hv": new_license}) \
                .eq("ma_hv", old_license) \
                .execute()

            supabase.table(KETQUA_TABLE) \
                .update({"ma_hv": new_license}) \
                .eq("ma_hv", old_license) \
                .execute()

            supabase.table(HOATDONG_TABLE) \
                .update({"ma_hv": new_license}) \
                .eq("ma_hv", old_license) \
                .execute()

            try:
                supabase.table(NOTIFICATION_TABLE) \
                    .update({"target_license": new_license}) \
                    .eq("target_license", old_license) \
                    .execute()
            except Exception as e:
                print("[UPDATE NOTIFICATION TARGET LICENSE ERROR]", e)

            try:
                supabase.table(NOTIFICATION_READ_TABLE) \
                    .update({"student_license": new_license}) \
                    .eq("student_license", old_license) \
                    .execute()
            except Exception as e:
                print("[UPDATE NOTIFICATION READ LICENSE ERROR]", e)

            # 3. Sau khi đổi mã xong, lấy student làm gốc và đồng bộ lại tên/ngày sinh/lớp/ca...
            sync_student_profile_to_related_tables(new_license)

            return jsonify({
                "ok": True,
                "new_license": new_license,
                "message": (
                    f"Đã đổi mã HV từ {old_license} sang {new_license}. "
                    f"Đã đồng bộ lại học phí, thi cấp, hoạt động và thông báo."
                )
            })

        # =========================
        # CẬP NHẬT CÁC TRƯỜNG KHÁC
        # Sau khi sửa student, quét lại các bảng liên quan.
        # =========================
        supabase.table(STUDENT_TABLE) \
            .update({field: value}) \
            .eq("license", license_code) \
            .execute()

        # Nếu sửa các thông tin có thể ảnh hưởng bảng khác thì đồng bộ lại
        fields_need_sync = {
            "name",
            "birthdate",
            "gender",
            "classroom",
            "timeclass",
            "clup",
            "phonenumber",
            "address",
        }

        if field in fields_need_sync:
            sync_student_profile_to_related_tables(license_code)

        return jsonify({
            "ok": True,
            "message": "Đã cập nhật và kiểm tra đồng bộ dữ liệu liên quan."
        })

    except Exception as e:
        print("[STUDENT FIELD UPDATE ERROR]", e)

        return jsonify({
            "ok": False,
            "message": str(e)
        }), 500

def get_next_belt_web(current_belt):
    current_belt = normalize_belt_name_web(current_belt)

    # Khóa riêng các trường hợp dễ lỗi
    if current_belt == "Cấp 1":
        return "1 Đẳng"

    if current_belt == "1 Đẳng":
        return "2 Đẳng"

    if current_belt == "2 Đẳng":
        return "3 Đẳng"

    if current_belt == "3 Đẳng":
        return "4 Đẳng"

    next_map = {
        "Cấp 10": "Cấp 9",
        "Cấp 9": "Cấp 8",
        "Cấp 8": "Cấp 7",
        "Cấp 7": "Cấp 6",
        "Cấp 6": "Cấp 5",
        "Cấp 5": "Cấp 4",
        "Cấp 4": "Cấp 3",
        "Cấp 3": "Cấp 2",
        "Cấp 2": "Cấp 1",

        "Cấp 1": "1 Đẳng",
        "1 Đẳng": "2 Đẳng",
        "2 Đẳng": "3 Đẳng",
        "3 Đẳng": "4 Đẳng",
        "4 Đẳng": "5 Đẳng",
        "5 Đẳng": "6 Đẳng",
        "6 Đẳng": "7 Đẳng",
        "7 Đẳng": "8 Đẳng",
        "8 Đẳng": "9 Đẳng",
        "9 Đẳng": "10 Đẳng",
    }

    return next_map.get(current_belt, current_belt)

def belt_css_class_web(belt_name):
    s = str(belt_name or "").strip().lower()
    s = remove_accents(s)
    s = s.replace(" ", "-")
    return s

def get_previous_belt_web(current_belt):
    current_belt = normalize_belt_name_web(current_belt)

    if current_belt in BELT_FLOW_WEB:
        idx = BELT_FLOW_WEB.index(current_belt)
        if idx - 1 >= 0:
            return BELT_FLOW_WEB[idx - 1]

    return current_belt or "Cấp 10"

def get_belt_index_web(belt):
    belt = str(belt or "").strip()
    if belt in BELT_FLOW_WEB:
        return BELT_FLOW_WEB.index(belt)
    return -1


def parse_ky_thi_web(ky_thi):
    """
    Nhận dạng:
    - Thi cấp: 2026-Q1, 2026-Q2, 2026-Q3, 2026-Q4
    - Thi đẳng: L1-2026, L2-2026, MN-2026, MT-2026, MB-2026, QG-2026

    Trả về số để so sánh thứ tự kỳ thi.
    """
    s = str(ky_thi or "").strip().upper()

    if not s:
        return 0

    # Thi cấp: 2026-Q3
    m = re.match(r"^(20\d{2})-Q([1-4])$", s)
    if m:
        year = int(m.group(1))
        q = int(m.group(2))
        return year * 100 + q

    # Thi đẳng: L1-2026, MN-2026...
    dan_order = {
        "L1": 11,
        "L2": 12,
        "MN": 13,
        "MT": 14,
        "MB": 15,
        "QG": 16,
    }

    m = re.match(r"^(L1|L2|MN|MT|MB|QG)-(20\d{2})$", s)
    if m:
        code = m.group(1)
        year = int(m.group(2))
        return year * 100 + dan_order.get(code, 0)

    # Dự phòng nếu dữ liệu cũ bị đảo: 2026-L1
    m = re.match(r"^(20\d{2})-(L1|L2|MN|MT|MB|QG)$", s)
    if m:
        year = int(m.group(1))
        code = m.group(2)
        return year * 100 + dan_order.get(code, 0)

    return 0


def get_expected_exam_belt_web(student_belt, previous_results):
    """
    Tính cấp dự thi đúng theo hiện trạng mới nhất.

    Quy tắc:
    - Nếu chưa từng thi: cấp dự thi = cấp kế tiếp của cấp hiện tại trong student.
    - Nếu lần gần nhất Vắng hoặc Không đạt: thi lại đúng cấp đó.
    - Nếu lần gần nhất Đạt: cấp dự thi = cấp kế tiếp của cấp hiện tại trong student.
      Vì student.belt đã được cập nhật sau khi chốt Đạt.
    """

    student_belt = normalize_belt_name_web(student_belt)

    if not previous_results:
        return get_next_belt_web(student_belt)

    previous_results = sorted(
        previous_results,
        key=lambda r: parse_ky_thi_web(r.get("ky_thi")),
        reverse=True
    )

    last = previous_results[0]
    last_exam_belt = normalize_belt_name_web(last.get("cap_dai_thi"))
    last_result = str(last.get("ket_qua") or "").strip()

    # Nếu quý trước vắng hoặc không đạt thì quý sau thi lại cấp đó
    if last_result in ["Không đạt", "Vắng"]:
        return last_exam_belt

    # Nếu quý trước đạt thì lấy cấp hiện tại trong student để tính cấp tiếp theo
    if last_result == "Đạt":
        if student_belt in BELT_FLOW_WEB:
            return get_next_belt_web(student_belt)

        return get_next_belt_web(last_exam_belt)

    # Trường hợp dữ liệu lạ thì ưu tiên cấp hiện tại
    if student_belt in BELT_FLOW_WEB:
        return get_next_belt_web(student_belt)

    return get_next_belt_web(last_exam_belt)

def recalc_student_belt_after_delete_web(student_license, deleted_exam_belt=""):
    """
    Tính lại cấp hiện tại của học viên sau khi xóa kết quả.

    Nguyên tắc:
    - Nếu còn kết quả Đạt trước đó => cấp hiện tại = cấp Đạt gần nhất.
    - Nếu không còn kết quả Đạt nào => cấp hiện tại = cấp trước của cấp vừa xóa.
      VD: xóa kết quả Đạt Cấp 7 => học viên quay về Cấp 8.
    """

    student_rows = safe_rows(STUDENT_TABLE, "*", license=student_license)
    student = student_rows[0] if student_rows else {}

    if not student:
        return ""

    current_belt = str(student.get("belt") or "").strip() or "Cấp 10"

    results = safe_rows(KETQUA_TABLE, "*", ma_hv=student_license)

    results = sorted(
        results,
        key=lambda r: parse_ky_thi_web(r.get("ky_thi")),
        reverse=True
    )

    # Nếu còn lịch sử Đạt thì lấy cấp Đạt mới nhất
    for r in results:
        if str(r.get("ket_qua") or "").strip() == "Đạt":
            passed_belt = str(r.get("cap_dai_thi") or "").strip()
            if passed_belt:
                return passed_belt

    # Nếu không còn kết quả Đạt nào, quay về cấp trước khi thi cấp vừa xóa
    deleted_exam_belt = str(deleted_exam_belt or "").strip()

    if deleted_exam_belt:
        return get_previous_belt_web(deleted_exam_belt)

    return current_belt or "Cấp 10"

def format_result_error_web(ma_hv, ho_ten, message):
    return f"{ma_hv} - {ho_ten}: {message}"

def clear_student_temp_leave_if_paid(license_code, month_codes):
    """
    Khi học viên đóng học phí lại, nếu tháng đóng trùng với tháng đang tạm nghỉ
    thì tự hủy trạng thái tạm nghỉ.
    """
    license_code = str(license_code or "").strip()

    if not license_code:
        return

    raw = str(month_codes or "")
    codes = [
        x.strip()
        for x in re.split(r"\s*-\s*", raw)
        if x.strip()
    ]

    if not codes:
        return

    try:
        rows = supabase.table(STUDENT_TABLE) \
            .select("license,temp_leave_month,temp_leave_year") \
            .eq("license", license_code) \
            .limit(1) \
            .execute().data or []

        if not rows:
            return

        student = rows[0]

        temp_month = student.get("temp_leave_month")
        temp_year = student.get("temp_leave_year")

        if not temp_month or not temp_year:
            return

        temp_code = f"{int(temp_month):02d}{int(temp_year)}"

        if temp_code in codes:
            supabase.table(STUDENT_TABLE) \
                .update({
                    "temp_leave_month": None,
                    "temp_leave_year": None,
                    "temp_leave_note": None,
                }) \
                .eq("license", license_code) \
                .execute()

    except Exception as e:
        print("[CLEAR TEMP LEAVE IF PAID ERROR]", e)

def safe_rows(table_name, select_text="*", **filters):
    try:
        q = supabase.table(table_name).select(select_text)

        for k, v in filters.items():
            q = q.eq(k, v)

        return q.execute().data or []

    except Exception as e:
        # Không dùng tiếng Việt ở print để tránh lỗi UnicodeEncodeError trên Windows terminal
        try:
            print(f"[WARN] Cannot read table {table_name}: {repr(e)}")
        except:
            pass

        return []


def get_prev_month_year(month, year):
    month = int(month)
    year = int(year)

    if month == 1:
        return 12, year - 1

    return month - 1, year


def get_month_code_web(month, year):
    return f"{int(month):02d}{int(year)}"


def parse_fee_datetime_web(raw):
    raw = str(raw or "").strip()

    if not raw:
        return None

    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00"))
    except:
        return None


def safe_due_date_web(year, month, day):
    year = int(year)
    month = int(month)
    day = int(day)

    last_day = calendar.monthrange(year, month)[1]
    day = min(day, last_day)

    return date(year, month, day)


def get_compare_date_web(year, month):
    """
    Nếu đang lọc tháng hiện tại: so với ngày hôm nay.
    Nếu lọc tháng cũ: so với ngày cuối tháng đó để xem cuối tháng còn ai quá hạn.
    Nếu lọc tháng tương lai: so với ngày đầu tháng đó.
    """
    today = date.today()
    year = int(year)
    month = int(month)

    if today.year == year and today.month == month:
        return today

    selected_first = date(year, month, 1)

    if selected_first > today.replace(day=1):
        return selected_first

    last_day = calendar.monthrange(year, month)[1]
    return date(year, month, last_day)


@app.get("/unpaid")
def unpaid():
    now = datetime.now()

    year = request.args.get("year", str(now.year))
    month = request.args.get("month", str(now.month))

    try:
        year_int = int(year)
        month_int = int(month)
    except:
        year_int = now.year
        month_int = now.month

    current_code = get_month_code_web(month_int, year_int)

    prev_month, prev_year = get_prev_month_year(month_int, year_int)
    prev_code = get_month_code_web(prev_month, prev_year)

    compare_date = get_compare_date_web(year_int, month_int)

    students = safe_rows(STUDENT_TABLE)
    fees = safe_rows(HOCPHI_TABLE)

    # =========================
    # 1) Ai đã đóng tháng đang lọc thì loại khỏi danh sách chưa đóng
    # =========================
    paid_current_ids = set()

    for f in fees:
        ma_thang = str(f.get("ma_thang") or "")
        ma_hv = str(f.get("ma_hv") or "").strip()

        if ma_hv and current_code in ma_thang:
            paid_current_ids.add(ma_hv)

    # =========================
    # 2) Lấy lần đóng gần nhất của tháng trước cho từng HV
    # =========================
    last_prev_fee_by_student = {}

    for f in fees:
        ma_thang = str(f.get("ma_thang") or "")
        ma_hv = str(f.get("ma_hv") or "").strip()

        if not ma_hv:
            continue

        if prev_code not in ma_thang:
            continue

        paid_dt = parse_fee_datetime_web(f.get("thoi_gian"))

        if not paid_dt:
            continue

        old = last_prev_fee_by_student.get(ma_hv)

        if not old or paid_dt > old["paid_dt"]:
            last_prev_fee_by_student[ma_hv] = {
                "paid_dt": paid_dt,
                "fee": f
            }

    rows = []

    summary = {
        "total": 0,
        "vang_lai": 0,
        "not_due": 0,
        "near_7_4": 0,
        "near_3_0": 0,
        "overdue": 0,
        "temp_leave": 0,
    }

    for s in students:
        active = str(s.get("active") or "").strip().lower()

        # Chỉ xét HV đang hoạt động
        if active not in ["có", "1", "true", "yes"]:
            continue

        license_code = str(s.get("license") or "").strip()

        if not license_code:
            continue

        # Đã đóng tháng đang lọc rồi thì không nằm trong danh sách chưa đóng
        if license_code in paid_current_ids:
            continue

        row = dict(s)

        # =========================
        # TẠM NGHỈ 1 THÁNG
        # Nếu học viên được set tạm nghỉ đúng tháng/năm đang lọc
        # thì ưu tiên hiện trạng thái này.
        # =========================
        temp_leave_month = s.get("temp_leave_month")
        temp_leave_year = s.get("temp_leave_year")

        try:
            is_temp_leave = (
                int(temp_leave_month or 0) == int(month_int)
                and int(temp_leave_year or 0) == int(year_int)
            )
        except:
            is_temp_leave = False

        if is_temp_leave:
            row["fee_status"] = "temp_leave"
            row["fee_status_label"] = "Tạm nghỉ 1 tháng"
            row["fee_due_date"] = "—"
            row["fee_days_left"] = "—"
            row["fee_row_class"] = "fee-row-temp-leave"

            summary["temp_leave"] += 1
            rows.append(row)
            continue

        prev_fee_info = last_prev_fee_by_student.get(license_code)

        # =========================
        # 3) Không có đóng tháng trước => HV vãng lai
        # =========================
        if not prev_fee_info:
            row["fee_status"] = "vang_lai"
            row["fee_status_label"] = "HV vãng lai"
            row["fee_due_date"] = "—"
            row["fee_days_left"] = "—"
            row["fee_row_class"] = "fee-row-vang-lai"

            summary["vang_lai"] += 1
            rows.append(row)
            continue

        # =========================
        # 4) Có đóng tháng trước => lấy ngày đóng tháng trước làm ngày đến hạn tháng này
        # VD đóng 15/6 -> hạn tháng 7 là 15/7
        # =========================
        last_paid_dt = prev_fee_info["paid_dt"]
        due_date = safe_due_date_web(year_int, month_int, last_paid_dt.day)

        days_left = (due_date - compare_date).days

        row["fee_due_date"] = due_date.strftime("%d/%m/%Y")
        row["fee_days_left"] = days_left

        if days_left < 0:
            row["fee_status"] = "overdue"
            row["fee_status_label"] = "Quá hạn"
            row["fee_row_class"] = "fee-row-overdue"
            summary["overdue"] += 1

        elif 0 <= days_left <= 3:
            row["fee_status"] = "near_3_0"
            row["fee_status_label"] = "Gần đến hạn 3-0 ngày"
            row["fee_row_class"] = "fee-row-near-orange"
            summary["near_3_0"] += 1

        elif 4 <= days_left <= 7:
            row["fee_status"] = "near_7_4"
            row["fee_status_label"] = "Gần đến hạn 7-4 ngày"
            row["fee_row_class"] = "fee-row-near-yellow"
            summary["near_7_4"] += 1

        else:
            row["fee_status"] = "not_due"
            row["fee_status_label"] = "Chưa đến hạn"
            row["fee_row_class"] = "fee-row-not-due"
            summary["not_due"] += 1

        rows.append(row)

    summary["total"] = len(rows)

    # Sắp xếp: Quá hạn -> cam -> vàng -> chưa đến hạn -> vãng lai
    status_order = {
        "overdue": 0,
        "near_3_0": 1,
        "near_7_4": 2,
        "not_due": 3,
        "temp_leave": 4,
        "vang_lai": 5,
    }

    rows.sort(key=lambda r: (
        status_order.get(r.get("fee_status"), 9),
        str(r.get("classroom") or ""),
        str(r.get("timeclass") or ""),
        str(r.get("name") or "")
    ))

    return render_template(
        "unpaid.html",
        rows=rows,
        year=str(year_int),
        month=str(month_int),
        current_year=now.year,
        summary=summary,
        compare_date=compare_date.strftime("%d/%m/%Y"),
        prev_month=prev_month,
        prev_year=prev_year
    )

@app.post("/unpaid/temp-leave")
def unpaid_temp_leave():
    license_code = request.form.get("license", "").strip()
    month = request.form.get("month", "").strip()
    year = request.form.get("year", "").strip()

    if not license_code or not month or not year:
        flash("Thiếu thông tin để set tạm nghỉ.", "danger")
        return redirect(url_for("unpaid"))

    try:
        supabase.table(STUDENT_TABLE) \
            .update({
                "temp_leave_month": int(month),
                "temp_leave_year": int(year),
                "temp_leave_note": "Tạm nghỉ 1 tháng",
            }) \
            .eq("license", license_code) \
            .execute()

        flash(f"Đã set tạm nghỉ 1 tháng cho {license_code}.", "success")

    except Exception as e:
        print("[SET TEMP LEAVE ERROR]", e)
        flash(f"Lỗi set tạm nghỉ: {e}", "danger")

    return back_to_current_page("unpaid")


@app.post("/unpaid/temp-leave-cancel")
def unpaid_temp_leave_cancel():
    license_code = request.form.get("license", "").strip()
    month = request.form.get("month", "").strip()
    year = request.form.get("year", "").strip()

    if not license_code:
        flash("Thiếu mã học viên để hủy tạm nghỉ.", "danger")
        return back_to_current_page("unpaid")

    try:
        supabase.table(STUDENT_TABLE) \
            .update({
                "temp_leave_month": None,
                "temp_leave_year": None,
                "temp_leave_note": None,
            }) \
            .eq("license", license_code) \
            .execute()

        flash(f"Đã hủy tạm nghỉ cho {license_code}.", "success")

    except Exception as e:
        print("[CANCEL TEMP LEAVE ERROR]", e)
        flash(f"Lỗi hủy tạm nghỉ: {e}", "danger")

    return back_to_current_page("unpaid")

@app.get("/exam-tracking")
def exam_tracking():
    now = datetime.now()

    year = request.args.get("year", str(now.year)).strip()
    quarter = request.args.get("quarter", f"Q{(now.month - 1)//3 + 1}").strip().upper()

    dan_quarters = ["L1", "L2", "MN", "MT", "MB", "QG"]

    if quarter in dan_quarters:
        ma_quy = f"{quarter}-{year}"
        ky_thi = f"{quarter}-{year}"
        tracking_title = "Theo dõi thi cấp đẳng"
    else:
        ma_quy = f"{quarter}{year}"
        ky_thi = f"{year}-{quarter}"
        tracking_title = "Theo dõi thi cấp đẳng"

    current_ky_value = parse_ky_thi_web(ky_thi)

    rows = safe_rows(HOCPHI_TABLE)
    rows = [
        r for r in rows
        if str(r.get("ma_quy") or "").strip().upper() == ma_quy.upper()
    ]

    rows.sort(key=lambda x: str(x.get("thoi_gian") or ""), reverse=True)

    # =========================
    # LẤY DANH SÁCH MÃ HV
    # =========================
    ids = []

    for r in rows:
        ma_hv = str(r.get("ma_hv") or "").strip()
        if ma_hv and ma_hv not in ids:
            ids.append(ma_hv)

    students_map = {}
    results_map = {}

    if ids:
        students = supabase.table(STUDENT_TABLE) \
            .select("license,name,birthdate,gender,belt") \
            .in_("license", ids) \
            .execute().data or []

        students_map = {
            str(s.get("license") or "").strip(): s
            for s in students
        }

        all_results = supabase.table(KETQUA_TABLE) \
            .select("*") \
            .in_("ma_hv", ids) \
            .execute().data or []

        for result in all_results:
            ma_hv = str(result.get("ma_hv") or "").strip()
            if ma_hv:
                results_map.setdefault(ma_hv, []).append(result)

    # =========================
    # TÍNH TIỀN THEO PHIẾU THU THỰC TẾ
    # =========================
    total_count = len(rows)

    total_exam_fee = sum([
        int(r.get("tong_tien") or 0)
        for r in rows
    ])

    cash_rows = [
        r for r in rows
        if str(r.get("chuyen_khoan") or "").strip().upper() == "TM"
    ]

    bank_rows = [
        r for r in rows
        if str(r.get("chuyen_khoan") or "").strip().upper() == "CK"
    ]

    cash_count = len(cash_rows)
    bank_count = len(bank_rows)

    cash_total = sum([
        int(r.get("tong_tien") or 0)
        for r in cash_rows
    ])

    bank_total = sum([
        int(r.get("tong_tien") or 0)
        for r in bank_rows
    ])

    # =========================
    # THỐNG KÊ CẤP DỰ THI
    # =========================
    belt_order = [
        "Cấp 9", "Cấp 8", "Cấp 7", "Cấp 6", "Cấp 5",
        "Cấp 4", "Cấp 3", "Cấp 2", "Cấp 1",
        "1 Đẳng", "2 Đẳng", "3 Đẳng", "4 Đẳng", "5 Đẳng",
        "6 Đẳng", "7 Đẳng", "8 Đẳng", "9 Đẳng", "10 Đẳng"
    ]

    belt_stats_map = {}

    for r in rows:
        ma_hv = str(r.get("ma_hv") or "").strip()
        student = students_map.get(ma_hv, {})

        current_belt = normalize_belt_name_web(student.get("belt"))
        exam_belt = get_next_belt_web(current_belt)

        if current_belt == "Cấp 1":
            exam_belt = "1 Đẳng"

        r["ho_ten"] = student.get("name") or r.get("ho_ten") or ""
        r["ngay_sinh"] = student.get("birthdate") or r.get("ngay_sinh") or ""
        r["gioi_tinh"] = student.get("gender") or r.get("gioi_tinh") or ""
        r["cap_hien_tai"] = current_belt
        r["cap_du_thi"] = exam_belt

        if exam_belt:
            belt_stats_map[exam_belt] = belt_stats_map.get(exam_belt, 0) + 1

    belt_stats = []

    for belt_name in belt_order:
        count = belt_stats_map.get(belt_name, 0)

        if count > 0:
            belt_stats.append({
                "name": belt_name,
                "count": count,
                "class": belt_css_class_web(belt_name)
            })

    return render_template(
        "exam_tracking.html",
        rows=rows,
        year=year,
        quarter=quarter,
        current_year=now.year,
        total_count=total_count,
        total_exam_fee=total_exam_fee,
        cash_count=cash_count,
        bank_count=bank_count,
        cash_total=cash_total,
        bank_total=bank_total,
        belt_stats=belt_stats,
        tracking_title=tracking_title
    )

@app.get("/exam-tracking/export")
def exam_tracking_export():
    now = datetime.now()

    year = request.args.get("year", str(now.year)).strip()
    quarter = request.args.get("quarter", f"Q{(now.month - 1)//3 + 1}").strip().upper()

    ma_quy = build_exam_ma_quy_web(year, quarter)

    rows = safe_rows(HOCPHI_TABLE)
    rows = [
        r for r in rows
        if str(r.get("ma_quy") or "").strip().upper() == ma_quy.upper()
    ]

    rows.sort(key=lambda x: str(x.get("thoi_gian") or ""), reverse=True)

    ids = []

    for r in rows:
        ma_hv = str(r.get("ma_hv") or "").strip()
        if ma_hv and ma_hv not in ids:
            ids.append(ma_hv)

    students_map = {}

    if ids:
        students = supabase.table(STUDENT_TABLE) \
            .select("license,name,birthdate,gender,belt") \
            .in_("license", ids) \
            .execute().data or []

        students_map = {
            str(s.get("license") or "").strip(): s
            for s in students
        }

    export_rows = []

    for r in rows:
        ma_hv = str(r.get("ma_hv") or "").strip()
        student = students_map.get(ma_hv, {})

        current_belt = normalize_belt_name_web(student.get("belt"))
        exam_belt = get_next_belt_web(current_belt)

        if current_belt == "Cấp 1":
            exam_belt = "1 Đẳng"

        export_rows.append({
            "ma_quy": r.get("ma_quy", ""),
            "ma_hv": ma_hv,
            "ho_ten": student.get("name") or r.get("ho_ten") or "",
            "ngay_sinh": student.get("birthdate") or r.get("ngay_sinh") or "",
            "gioi_tinh": student.get("gender") or r.get("gioi_tinh") or "",
            "cap_hien_tai": current_belt,
            "cap_du_thi": exam_belt,
            "ghi_chu": r.get("ghi_chu", ""),
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "Theo doi thi cap dang"

    headers = [
        "Mã quý",
        "Mã HV",
        "Họ tên",
        "Ngày sinh",
        "Giới tính",
        "Cấp hiện tại",
        "Cấp dự thi",
        "Ghi chú",
    ]

    # =========================
    # STYLE
    # =========================
    title_font = Font(bold=True, size=16, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="0F2A4A")

    sub_font = Font(bold=True, size=11, color="1E293B")

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D4ED8")

    thin_side = Side(style="thin", color="D9E2EF")
    table_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side
    )

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # =========================
    # TITLE
    # =========================
    ws.merge_cells("A1:H1")
    ws["A1"] = "DANH SÁCH THI CẤP ĐẲNG"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = center

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Bộ lọc: Năm {year} - Quý {quarter} | Tổng số: {len(export_rows)} võ sinh"
    ws["A2"].font = sub_font
    ws["A2"].alignment = center

    ws.merge_cells("A3:H3")
    ws["A3"] = f"Xuất lúc: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws["A3"].font = Font(italic=True, color="475569")
    ws["A3"].alignment = center

    # =========================
    # HEADER
    # =========================
    header_row = 5

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = table_border

    # =========================
    # DATA
    # =========================
    data_start_row = header_row + 1

    for row_idx, item in enumerate(export_rows, start=data_start_row):
        values = [
            item["ma_quy"],
            item["ma_hv"],
            item["ho_ten"],
            item["ngay_sinh"],
            item["gioi_tinh"],
            item["cap_hien_tai"],
            item["cap_du_thi"],
            item["ghi_chu"],
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = table_border
            cell.alignment = center if col_idx not in [3, 8] else left

            if col_idx in [6, 7]:
                cell.font = Font(bold=True)

    # =========================
    # AUTO WIDTH THEO ĐỘ DÀI CHỮ
    # =========================
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)

        max_len = len(str(header))

        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))

        width = max_len + 4

        # Giới hạn để không quá nhỏ hoặc quá dài
        if width < 12:
            width = 12
        if width > 36:
            width = 36

        ws.column_dimensions[col_letter].width = width

    # =========================
    # ROW HEIGHT + FREEZE
    # =========================
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[5].height = 24

    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:H{ws.max_row}"

    # Canh giữa tiêu đề sau khi merge
    for row in [1, 2, 3]:
        for col in range(1, 9):
            ws.cell(row=row, column=col).alignment = center

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"theo_doi_thi_cap_dang_{year}_{quarter}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

RESULT_PERIOD_OPTIONS = [
    {"value": "Q1", "label": "Q1"},
    {"value": "Q2", "label": "Q2"},
    {"value": "Q3", "label": "Q3"},
    {"value": "Q4", "label": "Q4"},
    {"value": "L1", "label": "L1"},
    {"value": "L2", "label": "L2"},
    {"value": "KVMN", "label": "KVMN"},
    {"value": "KVMT", "label": "KVMT"},
    {"value": "KVMB", "label": "KVMB"},
    {"value": "QG", "label": "QG"},
]

RESULT_CAP_OPTIONS = [
    "Cấp 9", "Cấp 8", "Cấp 7", "Cấp 6", "Cấp 5",
    "Cấp 4", "Cấp 3", "Cấp 2", "Cấp 1"
]

RESULT_DAN_OPTIONS = [
    "1 Đẳng", "2 Đẳng", "3 Đẳng", "4 Đẳng", "5 Đẳng",
    "6 Đẳng", "7 Đẳng", "8 Đẳng", "9 Đẳng", "10 Đẳng"
]

RESULT_CAP_THI_OPTIONS = RESULT_CAP_OPTIONS + RESULT_DAN_OPTIONS

def normalize_result_period_code_web(value):
    """
    Chuẩn hóa mã Quý / Lần cho trang kết quả.
    Hệ thống cũ đang lưu thi đẳng: MN, MT, MB.
    Giao diện Ken muốn hiện: KVMN, KVMT, KVMB.
    """
    s = str(value or "").strip().upper()
    s = s.replace(" ", "")

    period_map = {
        "MN": "KVMN",
        "MT": "KVMT",
        "MB": "KVMB",
        "KVMN": "KVMN",
        "KVMT": "KVMT",
        "KVMB": "KVMB",
    }

    return period_map.get(s, s)


def parse_result_ky_thi_web(ky_thi):
    """
    Nhận:
    - 2026-Q1, 2026-Q2...
    - Q1-2026, Q12026
    - L1-2026, L2-2026
    - MN-2026, MT-2026, MB-2026, QG-2026
    - KVMN-2026, KVMT-2026, KVMB-2026
    Trả về: year, period
    """
    s = str(ky_thi or "").strip().upper()

    if not s:
        return "", ""

    s = s.replace("_", "-").replace("/", "-").replace(" ", "")

    # Dạng 2026-Q1
    m = re.match(r"^(20\d{2})-(Q[1-4])$", s)
    if m:
        return m.group(1), normalize_result_period_code_web(m.group(2))

    # Dạng Q1-2026 hoặc Q12026
    m = re.match(r"^(Q[1-4])\D*(20\d{2})$", s)
    if m:
        return m.group(2), normalize_result_period_code_web(m.group(1))

    # Dạng L1-2026, MN-2026, KVMN-2026...
    m = re.match(r"^(L1|L2|MN|MT|MB|KVMN|KVMT|KVMB|QG)\D*(20\d{2})$", s)
    if m:
        return m.group(2), normalize_result_period_code_web(m.group(1))

    # Dạng 2026-L1, 2026-MN...
    m = re.match(r"^(20\d{2})\D*(L1|L2|MN|MT|MB|KVMN|KVMT|KVMB|QG)$", s)
    if m:
        return m.group(1), normalize_result_period_code_web(m.group(2))

    return "", ""


def result_sort_key_web(row):
    """
    Sort kết quả theo thứ tự kỳ thi mới nhất trước.
    """
    ky_thi = str(row.get("ky_thi") or "").strip()
    year, period = parse_result_ky_thi_web(ky_thi)

    period_order = {
        "Q1": 1,
        "Q2": 2,
        "Q3": 3,
        "Q4": 4,
        "L1": 11,
        "L2": 12,
        "KVMN": 13,
        "KVMT": 14,
        "KVMB": 15,
        "QG": 16,
    }

    try:
        year_int = int(year)
    except:
        year_int = 0

    return (
        year_int,
        period_order.get(period, 0),
        str(row.get("so_thi") or "")
    )

@app.get("/results")
def results():
    now = datetime.now()

    year = request.args.get("year", "").strip()
    quarter = request.args.get("quarter", "").strip().upper()
    quarter = normalize_result_period_code_web(quarter)

    name = request.args.get("name", "").strip()
    cap_thi = request.args.get("cap_thi", "").strip()
    ket_qua = request.args.get("ket_qua", "").strip()

    rows = safe_rows(KETQUA_TABLE)

    # Sắp xếp mới nhất lên trên, hỗ trợ cả thi cấp và thi đẳng
    rows.sort(key=result_sort_key_web, reverse=True)

    # Lấy danh sách năm đúng cho cả 2026-Q2 và L1-2026 / MN-2026
    year_options = sorted(
        {
            parse_result_ky_thi_web(r.get("ky_thi"))[0]
            for r in rows
            if parse_result_ky_thi_web(r.get("ky_thi"))[0]
        },
        reverse=True
    )

    # Nếu chưa có dữ liệu năm thì vẫn cho hiện quanh năm hiện tại
    if not year_options:
        year_options = [str(y) for y in range(now.year - 2, now.year + 6)]

    # Nếu chưa chọn năm thì mặc định năm hiện tại
    if not year:
        year = str(now.year)

    if quarter in ["Q1", "Q2", "Q3", "Q4"]:
        cap_thi_options = RESULT_CAP_OPTIONS
    elif quarter in ["L1", "L2", "KVMN", "KVMT", "KVMB", "QG"]:
        cap_thi_options = RESULT_DAN_OPTIONS
    else:
        cap_thi_options = RESULT_CAP_THI_OPTIONS

    period_options = RESULT_PERIOD_OPTIONS

    filtered = []

    for r in rows:
        ky_thi = str(r.get("ky_thi") or "").strip()
        r_year, r_period = parse_result_ky_thi_web(ky_thi)

        if year and r_year != year:
            continue

        if quarter and r_period != quarter:
            continue

        if name:
            text = f"{r.get('ma_hv', '')} {r.get('ho_ten', '')}".lower()
            if name.lower() not in text:
                continue

        if cap_thi:
            row_cap = normalize_belt_name_web(r.get("cap_dai_thi"))
            filter_cap = normalize_belt_name_web(cap_thi)

            if row_cap != filter_cap:
                continue

        if ket_qua and str(r.get("ket_qua") or "") != ket_qua:
            continue

        # Thêm nhãn hiển thị cho kỳ thi nếu cần dùng sau này
        r["period_code"] = r_period
        filtered.append(r)

    delete_errors = session.pop("result_delete_errors", [])

    return render_template(
        "results.html",
        rows=filtered,
        year=year,
        quarter=quarter,
        name=name,
        cap_thi=cap_thi,
        ket_qua=ket_qua,
        current_year=now.year,
        year_options=year_options,
        cap_thi_options=cap_thi_options,
        period_options=period_options,
        delete_errors=delete_errors
    )

@app.post("/results/delete")
def results_delete():
    ky_thi = request.form.get("ky_thi", "").strip()
    ma_hv = request.form.get("ma_hv", "").strip()
    so_thi = request.form.get("so_thi", "").strip()

    back_year = request.args.get("year", "").strip()
    back_quarter = request.args.get("quarter", "").strip()

    if not ky_thi or not ma_hv:
        session["result_delete_errors"] = [
            """
            <b>Thiếu thông tin xóa kết quả</b><br>
            Lý do: Không nhận được kỳ thi hoặc mã hội viên.<br>
            Cách xử lý: Ken tải lại trang rồi click phải xóa lại.
            """
        ]
        return redirect(url_for("results", year=back_year, quarter=back_quarter))

    try:
        all_results = safe_rows(KETQUA_TABLE, "*", ma_hv=ma_hv)

        if not all_results:
            session["result_delete_errors"] = [
                f"""
                <b>{ma_hv}</b><br>
                Lý do: Không tìm thấy lịch sử thi cấp của học viên này.<br>
                Cách xử lý: Ken kiểm tra lại bảng kết quả thi cấp.
                """
            ]
            return redirect(url_for("results", year=back_year, quarter=back_quarter))

        target_result = None

        for r in all_results:
            same_ky = str(r.get("ky_thi") or "").strip() == ky_thi
            same_so_thi = True

            if so_thi:
                same_so_thi = str(r.get("so_thi") or "").strip() == so_thi

            if same_ky and same_so_thi:
                target_result = r
                break

        if not target_result:
            session["result_delete_errors"] = [
                f"""
                <b>{ma_hv}</b><br>
                Lý do: Không tìm thấy dòng kết quả cần xóa trong kỳ <b>{ky_thi}</b>.<br>
                Cách xử lý: Ken tải lại trang hoặc kiểm tra lại bảng Supabase.
                """
            ]
            return redirect(url_for("results", year=back_year, quarter=back_quarter))

        target_ky_value = parse_ky_thi_web(target_result.get("ky_thi"))

        newer_results = [
            r for r in all_results
            if parse_ky_thi_web(r.get("ky_thi")) > target_ky_value
        ]

        if newer_results:
            newer_results = sorted(
                newer_results,
                key=lambda r: parse_ky_thi_web(r.get("ky_thi")),
                reverse=True
            )

            newest = newer_results[0]

            session["result_delete_errors"] = [
                f"""
                <b>{ma_hv} - {target_result.get('ho_ten', '')}</b><br>
                Lý do: Không thể xóa kết quả kỳ <b>{target_result.get('ky_thi')}</b>
                vì học viên này đã có kết quả mới hơn.<br><br>

                Kết quả mới nhất hiện tại:<br>
                Kỳ thi: <b>{newest.get('ky_thi')}</b><br>
                Cấp thi: <b>{newest.get('cap_dai_thi')}</b><br>
                Số thi: <b>{newest.get('so_thi')}</b><br>
                Kết quả: <b>{newest.get('ket_qua')}</b><br><br>

                Cách xử lý: Ken phải xóa từ kết quả mới nhất về cũ nhất.
                Ví dụ đã thi Cấp 7 → Cấp 6 → Cấp 5 thì phải xóa <b>Cấp 5</b> trước.
                """
            ]

            return redirect(url_for("results", year=back_year, quarter=back_quarter))

        q = supabase.table(KETQUA_TABLE) \
            .delete() \
            .eq("ky_thi", ky_thi) \
            .eq("ma_hv", ma_hv)

        if so_thi:
            q = q.eq("so_thi", so_thi)

        q.execute()

        deleted_exam_belt = str(target_result.get("cap_dai_thi") or "").strip()
        new_belt = recalc_student_belt_after_delete_web(ma_hv, deleted_exam_belt)

        if new_belt:
            supabase.table(STUDENT_TABLE) \
                .update({"belt": new_belt}) \
                .eq("license", ma_hv) \
                .execute()

        flash(
            f"Đã xóa kết quả {ma_hv} - {ky_thi}. "
            f"Đã cập nhật cấp hiện tại về {new_belt}."
        )

    except Exception as e:
        print("[DELETE RESULT ERROR]", e)

        session["result_delete_errors"] = [
            f"""
            <b>Lỗi xóa kết quả</b><br>
            Lý do hệ thống: {e}<br>
            Cách xử lý: Ken kiểm tra lại kết nối Supabase hoặc dữ liệu trong bảng ketqua.
            """
        ]

    return redirect(url_for("results", year=back_year, quarter=back_quarter))

def parse_web_date(raw):
    """
    Nhận yyyy-mm-dd từ input type=date.
    Trả về datetime hoặc None.
    """
    raw = str(raw or "").strip()

    try:
        return datetime.strptime(raw, "%Y-%m-%d")
    except:
        return None


def format_activity_date_range(start_raw, end_raw):
    start = parse_web_date(start_raw)
    end = parse_web_date(end_raw)

    if not start and not end:
        return ""

    if start and not end:
        return start.strftime("%d/%m/%Y")

    if end and not start:
        return end.strftime("%d/%m/%Y")

    # Nếu cùng ngày
    if start.date() == end.date():
        return start.strftime("%d/%m/%Y")

    # Cùng tháng, cùng năm: 14 - 16/07/2026
    if start.month == end.month and start.year == end.year:
        return f"{start.strftime('%d')} - {end.strftime('%d/%m/%Y')}"

    # Khác tháng hoặc khác năm: 28/06 - 04/07/2026
    if start.year == end.year:
        return f"{start.strftime('%d/%m')} - {end.strftime('%d/%m/%Y')}"

    # Khác năm: 28/12/2025 - 04/01/2026
    return f"{start.strftime('%d/%m/%Y')} - {end.strftime('%d/%m/%Y')}"

def normalize_activity_text_web(text):
    text = str(text or "").strip().lower()
    text = remove_accents(text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def parse_activity_start_date_web(raw):
    """
    Lấy ngày bắt đầu từ chuỗi thời gian hoạt động.

    Nhận các dạng:
    - 06 - 12/07/2026
    - 12-15/07/2026
    - 13 - 15/07/2026
    - 19 - 25/06/2026
    - 14/06 - 17/07/2026
    - 27/06/2026
    """
    raw = str(raw or "").strip()

    if not raw:
        return date(9999, 12, 31)

    s = raw.replace("–", "-").replace("—", "-")
    s = re.sub(r"\s+", " ", s).strip()

    # Dạng: 06 - 12/07/2026 hoặc 12-15/07/2026
    m = re.search(r"(\d{1,2})\s*-\s*(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        start_day = int(m.group(1))
        month = int(m.group(3))
        year = int(m.group(4))

        try:
            return date(year, month, start_day)
        except:
            return date(9999, 12, 31)

    # Dạng: 14/06 - 17/07/2026
    m = re.search(r"(\d{1,2})/(\d{1,2})\s*-\s*(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        start_day = int(m.group(1))
        start_month = int(m.group(2))
        year = int(m.group(5))

        try:
            return date(year, start_month, start_day)
        except:
            return date(9999, 12, 31)

    # Dạng: 27/06/2026
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        day = int(m.group(1))
        month = int(m.group(2))
        year = int(m.group(3))

        try:
            return date(year, month, day)
        except:
            return date(9999, 12, 31)

    return date(9999, 12, 31)


def activity_medal_rank_web(result_text):
    """
    Thứ tự trong cùng giải:
    HCV -> HCB -> HCĐ -> khác
    """
    text = normalize_activity_text_web(result_text)

    if (
        "huy chuong vang" in text
        or "hcv" in text
        or "hang nhat" in text
        or "giai nhat" in text
    ):
        return 0

    if (
        "huy chuong bac" in text
        or "hcb" in text
        or "hang nhi" in text
        or "giai nhi" in text
        or "a quan" in text
    ):
        return 1

    if (
        "huy chuong dong" in text
        or "hcd" in text
        or "hcđ" in str(result_text or "").lower()
        or "hang ba" in text
        or "giai ba" in text
    ):
        return 2

    return 9


def activity_group_key_web(activity_name):
    """
    Gom theo tên giải, bỏ bớt khoảng trắng và dấu để sort ổn định.
    """
    return normalize_activity_text_web(activity_name)


@app.get("/activities")
def activities():
    rows = safe_rows(HOATDONG_TABLE)

    # Chuẩn bị dữ liệu phụ cho template tô màu từng giải
    for r in rows:
        start_date = parse_activity_start_date_web(r.get("thoi_gian"))
        medal_rank = activity_medal_rank_web(r.get("ket_qua"))
        group_key = activity_group_key_web(r.get("hoat_dong"))

        r["_activity_start_date"] = start_date
        r["_activity_group_key"] = group_key
        r["_activity_medal_rank"] = medal_rank

    # Sort:
    # 1. Ngày bắt đầu sớm trước
    # 2. Tên giải
    # 3. HCV -> HCB -> HCĐ
    # 4. Họ tên
    rows.sort(key=lambda r: (
        r.get("_activity_start_date") or date(9999, 12, 31),
        r.get("_activity_group_key") or "",
        r.get("_activity_medal_rank", 9),
        normalize_activity_text_web(r.get("ho_ten"))
    ))

    # Tô màu luân phiên theo từng giải
    last_group = None
    group_index = -1

    for r in rows:
        current_group = r.get("_activity_group_key") or ""

        if current_group != last_group:
            group_index += 1
            last_group = current_group

        r["_activity_group_class"] = f"activity-group-{group_index % 6}"

        medal_rank = r.get("_activity_medal_rank", 9)

        if medal_rank == 0:
            r["_activity_medal_class"] = "activity-medal-gold"
        elif medal_rank == 1:
            r["_activity_medal_class"] = "activity-medal-silver"
        elif medal_rank == 2:
            r["_activity_medal_class"] = "activity-medal-bronze"
        else:
            r["_activity_medal_class"] = ""

    students = safe_rows(STUDENT_TABLE)
    students.sort(key=lambda s: str(s.get("name") or ""))

    events = get_activity_events_web()

    return render_template(
        "activities.html",
        rows=rows,
        students=students,
        events=events
    )

@app.post("/activities/add")
def activities_add():
    f = request.form

    ma_hv = f.get("ma_hv", "").strip()

    student_rows = safe_rows(STUDENT_TABLE, "*", license=ma_hv)
    student = student_rows[0] if student_rows else {}

    if not student:
        flash("Không tìm thấy hội viên.")
        return redirect(url_for("activities"))

    start_date = f.get("start_date", "").strip()
    end_date = f.get("end_date", "").strip()

    thoi_gian_text = format_activity_date_range(start_date, end_date)

    payload = {
        "ma_hv": ma_hv,
        "ho_ten": student.get("name", ""),
        "hoat_dong": f.get("hoat_dong", "").strip(),
        "thoi_gian": thoi_gian_text,
        "dia_diem": f.get("dia_diem", "").strip(),
        "noi_dung": f.get("noi_dung", "").strip(),
        "ket_qua": f.get("ket_qua", "").strip(),
    }

    try:
        supabase.table(HOATDONG_TABLE).insert(payload).execute()
        flash(f"Đã thêm hoạt động cho {student.get('name', '')}")
    except Exception as e:
        print("[ADD ACTIVITY ERROR]", e)
        flash(f"Lỗi thêm hoạt động: {e}")

    return back_to_current_page("activities")


@app.post("/activities/update/<activity_id>")
def activities_update(activity_id):
    f = request.form

    ma_hv = f.get("ma_hv", "").strip()

    student_rows = safe_rows(STUDENT_TABLE, "*", license=ma_hv)
    student = student_rows[0] if student_rows else {}

    if not student:
        flash("Không tìm thấy hội viên để cập nhật hoạt động.", "danger")
        return back_to_current_page("activities")

    payload = {
        "ma_hv": ma_hv,
        "ho_ten": student.get("name", ""),
        "hoat_dong": f.get("hoat_dong", "").strip(),
        "thoi_gian": f.get("thoi_gian", "").strip(),
        "dia_diem": f.get("dia_diem", "").strip(),
        "noi_dung": f.get("noi_dung", "").strip(),
        "ket_qua": f.get("ket_qua", "").strip(),
    }

    try:
        supabase.table(HOATDONG_TABLE) \
            .update(payload) \
            .eq("id", activity_id) \
            .execute()

        flash(f"Đã cập nhật hoạt động cho {student.get('name', '')}", "success")

    except Exception as e:
        print("[UPDATE ACTIVITY ERROR]", e)
        flash(f"Lỗi cập nhật hoạt động: {e}", "danger")

    return back_to_current_page("activities")


@app.post("/activities/delete/<activity_id>")
def activities_delete(activity_id):
    try:
        supabase.table(HOATDONG_TABLE) \
            .delete() \
            .eq("id", activity_id) \
            .execute()

        flash("Đã xóa hoạt động.", "success")

    except Exception as e:
        print("[DELETE ACTIVITY ERROR]", e)
        flash(f"Lỗi xóa hoạt động: {e}", "danger")

    return back_to_current_page("activities")


def format_activity_event_time_web(start_date, end_date):
    """
    Nhận YYYY-MM-DD, trả về:
    - 06 - 12/07/2026 nếu cùng tháng/năm
    - 14/06 - 17/07/2026 nếu khác tháng
    - 27/06/2026 nếu cùng ngày
    """
    start_date = str(start_date or "").strip()
    end_date = str(end_date or "").strip()

    def parse_iso_date(v):
        try:
            return datetime.strptime(v, "%Y-%m-%d").date()
        except:
            return None

    s = parse_iso_date(start_date)
    e = parse_iso_date(end_date)

    if not s and not e:
        return ""

    if s and not e:
        return s.strftime("%d/%m/%Y")

    if e and not s:
        return e.strftime("%d/%m/%Y")

    if s == e:
        return s.strftime("%d/%m/%Y")

    if s.year == e.year and s.month == e.month:
        return f"{s.day:02d} - {e.day:02d}/{e.month:02d}/{e.year}"

    return f"{s.day:02d}/{s.month:02d} - {e.day:02d}/{e.month:02d}/{e.year}"


def get_activity_events_web():
    try:
        events = supabase.table(ACTIVITY_EVENTS_TABLE) \
            .select("*") \
            .order("start_date") \
            .execute().data or []

        return events

    except Exception as e:
        print("[GET ACTIVITY EVENTS ERROR]", e)
        return []


@app.post("/activity-events/add")
def activity_events_add():
    data = request.get_json(silent=True) or {}

    event_name = str(data.get("event_name") or "").strip()
    start_date = str(data.get("start_date") or "").strip()
    end_date = str(data.get("end_date") or "").strip()
    location = str(data.get("location") or "").strip()

    if not event_name:
        return jsonify({
            "ok": False,
            "message": "Ken chưa nhập tên sự kiện."
        }), 400

    if not start_date:
        return jsonify({
            "ok": False,
            "message": "Ken chưa chọn ngày bắt đầu."
        }), 400

    if not end_date:
        end_date = start_date

    payload = {
        "event_name": event_name,
        "start_date": start_date,
        "end_date": end_date,
        "location": location,
    }

    try:
        result = supabase.table(ACTIVITY_EVENTS_TABLE) \
            .insert(payload) \
            .execute().data or []

        event = result[0] if result else payload

        return jsonify({
            "ok": True,
            "message": "Đã lưu sự kiện.",
            "event": event
        })

    except Exception as e:
        print("[ADD ACTIVITY EVENT ERROR]", e)

        return jsonify({
            "ok": False,
            "message": f"Lỗi lưu sự kiện: {e}"
        }), 500


@app.get("/activity-events/list")
def activity_events_list():
    try:
        events = supabase.table(ACTIVITY_EVENTS_TABLE) \
            .select("*") \
            .order("start_date") \
            .execute().data or []

        return jsonify({
            "ok": True,
            "events": events
        })

    except Exception as e:
        print("[LIST ACTIVITY EVENTS ERROR]", e)

        return jsonify({
            "ok": False,
            "message": str(e),
            "events": []
        }), 500


@app.post("/activities/bulk-add")
def activities_bulk_add():
    data = request.get_json(silent=True) or {}

    event_name = str(data.get("event_name") or "").strip()
    start_date = str(data.get("start_date") or "").strip()
    end_date = str(data.get("end_date") or "").strip()
    location = str(data.get("location") or "").strip()
    items = data.get("items") or []

    if not event_name:
        return jsonify({
            "ok": False,
            "message": "Thiếu tên sự kiện."
        }), 400

    if not start_date:
        return jsonify({
            "ok": False,
            "message": "Thiếu ngày bắt đầu."
        }), 400

    if not end_date:
        end_date = start_date

    if not isinstance(items, list) or not items:
        return jsonify({
            "ok": False,
            "message": "Ken chưa thêm nội dung và hội viên."
        }), 400

    all_licenses = []

    for item in items:
        students = item.get("students") or []

        for license_code in students:
            license_code = str(license_code or "").strip()

            if license_code and license_code not in all_licenses:
                all_licenses.append(license_code)

    if not all_licenses:
        return jsonify({
            "ok": False,
            "message": "Ken chưa chọn hội viên nào."
        }), 400

    try:
        student_rows = supabase.table(STUDENT_TABLE) \
            .select("license,name") \
            .in_("license", all_licenses) \
            .execute().data or []

        student_map = {
            str(s.get("license") or "").strip(): s
            for s in student_rows
        }

        thoi_gian = format_activity_event_time_web(start_date, end_date)

        payloads = []

        for item in items:
            noi_dung = str(item.get("noi_dung") or "").strip()
            ket_qua = str(item.get("ket_qua") or "").strip()
            students = item.get("students") or []

            if not noi_dung:
                continue

            if not ket_qua:
                continue

            for license_code in students:
                license_code = str(license_code or "").strip()
                student = student_map.get(license_code)

                if not student:
                    continue

                payloads.append({
                    "ma_hv": license_code,
                    "ho_ten": student.get("name") or "",
                    "hoat_dong": event_name,
                    "thoi_gian": thoi_gian,
                    "dia_diem": location,
                    "noi_dung": noi_dung,
                    "ket_qua": ket_qua,
                })

        if not payloads:
            return jsonify({
                "ok": False,
                "message": "Không có dữ liệu hợp lệ để lưu."
            }), 400

        supabase.table(HOATDONG_TABLE).insert(payloads).execute()

        return jsonify({
            "ok": True,
            "message": f"Đã lưu {len(payloads)} hoạt động.",
            "count": len(payloads)
        })

    except Exception as e:
        print("[BULK ADD ACTIVITIES ERROR]", e)

        return jsonify({
            "ok": False,
            "message": f"Lỗi lưu hoạt động: {e}"
        }), 500
        
def clean_option_list(values):
    cleaned = []
    seen = set()

    for value in values or []:
        value = str(value or "").strip()

        if not value:
            continue

        key = value.lower()

        if key in seen:
            continue

        seen.add(key)
        cleaned.append(value)

    return cleaned


def get_class_options_settings(settings=None):
    settings = settings or load_app_settings()
    defaults = DEFAULT_APP_SETTINGS["class_options"]

    class_options = settings.setdefault("class_options", {})

    for key, default_values in defaults.items():
        current_values = class_options.get(key)

        if not isinstance(current_values, list):
            current_values = default_values.copy()

        class_options[key] = clean_option_list(current_values)

        if not class_options[key]:
            class_options[key] = default_values.copy()

    return class_options


OPTION_META = {
    "classrooms": "Lớp",
    "timeclasses": "Ca",
    "clubs": "CLB",
}

def get_list_count(form, name):
    try:
        return int(form.get(name, "0") or 0)
    except:
        return 0


def read_club_info_from_form(form, files, old_club_info):
    old_club_info = old_club_info or {}

    head_old = old_club_info.get("head_coach", {}) or {}

    head_photo_url = str(form.get("head_photo_url") or head_old.get("photo_url") or "").strip()
    head_photo_file = files.get("head_photo_file")

    if head_photo_file and head_photo_file.filename:
        head_photo_url = upload_club_asset_to_supabase(head_photo_file, "club-info/coaches")

    head_coach = {
        "name": str(form.get("head_name") or "").strip(),
        "role": str(form.get("head_role") or "HLV Trưởng").strip(),
        "phone": str(form.get("head_phone") or "").strip(),
        "qualification": str(form.get("head_qualification") or "").strip(),
        "photo_url": head_photo_url,
        "description": str(form.get("head_description") or "").strip(),
    }

    registrar_old = old_club_info.get("registrar", {}) or {}

    registrar_photo_url = str(
        form.get("registrar_photo_url") or registrar_old.get("photo_url") or ""
    ).strip()

    registrar_photo_file = files.get("registrar_photo_file")

    if registrar_photo_file and registrar_photo_file.filename:
        registrar_photo_url = upload_club_asset_to_supabase(
            registrar_photo_file,
            "club-info/registrar"
        )

    registrar = {
        "name": str(form.get("registrar_name") or "").strip(),
        "role": str(form.get("registrar_role") or "Người ghi danh").strip(),
        "phone": str(form.get("registrar_phone") or "").strip(),
        "photo_url": registrar_photo_url,
        "description": str(form.get("registrar_description") or "").strip(),
    }

    coaches = []
    coach_count = get_list_count(form, "coach_count")

    group_map = {
        "246": {
            "group_title": "HLV lớp 2-4-6",
            "group_subtitle": "Lớp buổi chiều",
        },
        "357": {
            "group_title": "HLV lớp 3-5-7",
            "group_subtitle": "Lớp chính trong tuần",
        },
        "weekend": {
            "group_title": "HLV lớp sáng & cuối tuần",
            "group_subtitle": "Thứ 7, Chủ nhật và lớp sáng",
        },
    }

    for i in range(coach_count):
        if str(form.get(f"coach_delete_{i}") or "") == "1":
            continue

        name = str(form.get(f"coach_name_{i}") or "").strip()
        role = str(form.get(f"coach_role_{i}") or "").strip()
        group = str(form.get(f"coach_group_{i}") or "246").strip()
        phone = str(form.get(f"coach_phone_{i}") or "").strip()
        qualification = str(form.get(f"coach_qualification_{i}") or "").strip()

        if not name and not role and not phone and not qualification:
            continue

        photo_url = str(form.get(f"coach_photo_url_{i}") or "").strip()
        photo_file = files.get(f"coach_photo_file_{i}")

        if photo_file and photo_file.filename:
            photo_url = upload_club_asset_to_supabase(photo_file, "club-info/coaches")

        group_info = group_map.get(group, group_map["246"])

        coaches.append({
            "group": group,
            "group_title": group_info["group_title"],
            "group_subtitle": group_info["group_subtitle"],
            "name": name,
            "role": role,
            "phone": phone,
            "qualification": qualification,
            "photo_url": photo_url,
        })

    regular_schedules = []
    regular_count = get_list_count(form, "regular_schedule_count")

    for i in range(regular_count):
        if str(form.get(f"regular_schedule_delete_{i}") or "") == "1":
            continue

        title = str(form.get(f"regular_schedule_title_{i}") or "").strip()
        time_text = str(form.get(f"regular_schedule_time_{i}") or "").strip()

        if title or time_text:
            regular_schedules.append({
                "title": title,
                "time": time_text,
            })

    summer_schedules = []
    summer_count = get_list_count(form, "summer_schedule_count")

    for i in range(summer_count):
        if str(form.get(f"summer_schedule_delete_{i}") or "") == "1":
            continue

        title = str(form.get(f"summer_schedule_title_{i}") or "").strip()
        time_text = str(form.get(f"summer_schedule_time_{i}") or "").strip()

        if title or time_text:
            summer_schedules.append({
                "title": title,
                "time": time_text,
            })

    exam_notes = []
    note_count = get_list_count(form, "exam_note_count")

    for i in range(note_count):
        if str(form.get(f"exam_note_delete_{i}") or "") == "1":
            continue

        note = str(form.get(f"exam_note_{i}") or "").strip()

        if note:
            exam_notes.append(note)

    return {
        "intro_title": str(form.get("intro_title") or "").strip(),
        "intro_subtitle": str(form.get("intro_subtitle") or "").strip(),
        "intro_content": str(form.get("intro_content") or "").strip(),

        "head_coach": head_coach,
        "registrar": registrar,
        "coaches": coaches,

        "regular_schedules": regular_schedules,
        "summer_schedules": summer_schedules,

        "fees_info": {
            "monthly_fee": str(form.get("monthly_fee") or "").strip(),
            "three_month_discount": str(form.get("three_month_discount") or "").strip(),
            "six_month_bonus": str(form.get("six_month_bonus") or "").strip(),
            "family_discount": str(form.get("family_discount") or "").strip(),
            "uniform_fee": str(form.get("uniform_fee") or "").strip(),
            "exam_fee": str(form.get("club_exam_fee") or "").strip(),
            "exam_notes": exam_notes,
        }
    }


@app.post("/setup/class-options/add")
def setup_class_options_add():
    option_type = request.form.get("option_type", "").strip()
    value = request.form.get("value", "").strip()

    if option_type not in OPTION_META:
        flash("Nhóm thông tin không hợp lệ.", "danger")
        return redirect(url_for("setup", tab="classes"))

    if not value:
        flash(f"{OPTION_META[option_type]} không được để trống.", "danger")
        return redirect(url_for("setup", tab="classes"))

    settings = load_app_settings()
    class_options = get_class_options_settings(settings)

    current = clean_option_list(class_options.get(option_type, []))

    if value.lower() in [x.lower() for x in current]:
        flash(f"{OPTION_META[option_type]} này đã có trong danh sách.", "danger")
        return redirect(url_for("setup", tab="classes"))

    current.append(value)

    class_options[option_type] = current
    settings["class_options"] = class_options

    save_app_settings(settings)

    flash(f"Đã thêm {OPTION_META[option_type]}: {value}", "success")
    return redirect(url_for("setup", tab="classes"))


@app.post("/setup/class-options/delete")
def setup_class_options_delete():
    option_type = request.form.get("option_type", "").strip()
    value = request.form.get("value", "").strip()

    if option_type not in OPTION_META:
        flash("Nhóm thông tin không hợp lệ.", "danger")
        return redirect(url_for("setup", tab="classes"))

    settings = load_app_settings()
    class_options = get_class_options_settings(settings)

    current = clean_option_list(class_options.get(option_type, []))
    class_options[option_type] = [x for x in current if x != value]

    settings["class_options"] = class_options

    save_app_settings(settings)

    flash(f"Đã xóa {OPTION_META[option_type]}: {value}", "success")
    return redirect(url_for("setup", tab="classes"))

@app.get("/information")
def information():
    return render_template("information.html", settings=load_app_settings())


@app.route("/setup", methods=["GET", "POST"])
def setup():
    settings = load_app_settings()
    active_tab = request.args.get("tab", "header")

    if request.method == "POST":
        active_tab = request.form.get("active_tab", "header").strip() or "header"

        try:
            if active_tab == "header":
                settings["header"]["club_small_title"] = request.form.get(
                    "club_small_title",
                    settings["header"]["club_small_title"]
                ).strip() or DEFAULT_APP_SETTINGS["header"]["club_small_title"]

                settings["header"]["club_name"] = request.form.get(
                    "club_name",
                    settings["header"]["club_name"]
                ).strip() or DEFAULT_APP_SETTINGS["header"]["club_name"]

                logo_file = request.files.get("logo_file")

                if logo_file and logo_file.filename:
                    filename = secure_filename(logo_file.filename)
                    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

                    if ext not in ALLOWED_PHOTO_EXTENSIONS:
                        flash("Logo chỉ nhận file jpg, jpeg, png, webp.", "danger")
                        return redirect(url_for("setup", tab="header"))

                    settings["header"]["logo_url"] = upload_club_asset_to_supabase(
                        logo_file,
                        "club-info/header"
                    )

            elif active_tab == "fees":
                settings["fees"]["tuition_fee"] = money_to_int_web(
                    request.form.get("tuition_fee", settings["fees"].get("tuition_fee", 500000))
                )

                settings["fees"]["exam_fee"] = money_to_int_web(
                    request.form.get("exam_fee", settings["fees"].get("exam_fee", 300000))
                )

                dan_fees = {}

                for dan_name in ["1 Đẳng", "2 Đẳng", "3 Đẳng", "4 Đẳng", "5 Đẳng", "6 Đẳng"]:
                    field_name = f"dan_fee_{dan_name.replace(' ', '_')}"
                    dan_fees[dan_name] = money_to_int_web(
                        request.form.get(
                            field_name,
                            settings["fees"].get("dan_fees", {}).get(dan_name, 0)
                        )
                    )

                settings["fees"]["dan_fees"] = dan_fees

            elif active_tab == "exam":
                settings["exam"]["exam_number_prefix"] = request.form.get(
                    "exam_number_prefix",
                    settings["exam"]["exam_number_prefix"]
                )

            elif active_tab == "club_info":
                settings["club_info"] = read_club_info_from_form(
                    request.form,
                    request.files,
                    settings.get("club_info", DEFAULT_APP_SETTINGS.get("club_info", {}))
                )

            elif active_tab == "payment":
                save_payment_settings(request.form)
                flash("Đã lưu thông tin thanh toán lên Supabase")
            save_app_settings(settings)
            flash("Đã lưu cài đặt.", "success")



        except Exception as e:
            print("[SETUP SAVE ERROR]", e)
            flash(f"Lỗi lưu cài đặt: {e}", "danger")

        return redirect(url_for("setup", tab=active_tab))

    get_class_options_settings(settings)

    return render_template(
        "setup.html",
        settings=settings,
        active_tab=active_tab,
        payment_settings=get_payment_settings()
    )

def ensure_student_portal_account(student):
    """
    Tạo tài khoản mặc định cho học viên nếu chưa có:
    username = Mã HV
    password = SĐT
    """
    if not student:
        return {}

    license_code = str(student.get("license") or "").strip()
    phone = str(student.get("phonenumber") or "").strip()

    if not license_code:
        return student

    update_data = {}

    if not str(student.get("portal_username") or "").strip():
        update_data["portal_username"] = license_code

    if not str(student.get("portal_password_hash") or "").strip() and phone:
        update_data["portal_password_hash"] = generate_password_hash(phone)

    if update_data:
        try:
            supabase.table(STUDENT_TABLE) \
                .update(update_data) \
                .eq("license", license_code) \
                .execute()

            student.update(update_data)
        except Exception as e:
            print("[PORTAL ACCOUNT INIT ERROR]", e)

    return student


def get_logged_student():
    license_code = str(session.get("student_license") or "").strip()

    if not license_code:
        return None

    rows = safe_rows(STUDENT_TABLE, "*", license=license_code)

    if not rows:
        session.pop("student_license", None)
        return None

    return rows[0]


def require_student_login():
    student = get_logged_student()

    if not student:
        return None

    return student


def get_student_notifications(student_license):
    student_license = str(student_license or "").strip()

    all_notifications = safe_rows(NOTIFICATION_TABLE)

    rows = []

    for n in all_notifications:
        target_type = str(n.get("target_type") or "all").strip()
        target_license = str(n.get("target_license") or "").strip()

        if target_type == "all" or target_license == student_license:
            rows.append(n)

    rows.sort(key=lambda r: str(r.get("created_at") or ""), reverse=True)

    reads = safe_rows(NOTIFICATION_READ_TABLE, "*", student_license=student_license)
    read_ids = {
        str(r.get("notification_id"))
        for r in reads
    }

    for n in rows:
        n["is_read"] = str(n.get("id")) in read_ids

    unread_count = len([n for n in rows if not n.get("is_read")])

    return rows, unread_count


def mark_notification_read(notification_id, student_license):
    notification_id = str(notification_id or "").strip()
    student_license = str(student_license or "").strip()

    if not notification_id or not student_license:
        return

    try:
        existing = supabase.table(NOTIFICATION_READ_TABLE) \
            .select("id") \
            .eq("notification_id", notification_id) \
            .eq("student_license", student_license) \
            .limit(1) \
            .execute().data or []

        if existing:
            return

        supabase.table(NOTIFICATION_READ_TABLE).insert({
            "notification_id": notification_id,
            "student_license": student_license
        }).execute()

    except Exception as e:
        print("[MARK NOTIFICATION READ ERROR]", e)

NOTIFICATION_CONTACT_FOOTER = (
    "Mọi thắc mắc xin liên hệ SĐT/Zalo: "
    "0963830315 - Tâm (Người ghi danh)."
)


def append_notification_footer(content):
    content = str(content or "").strip()

    if not content:
        return ""

    # Tránh bị chèn lặp nếu Ken đã nhập sẵn dòng này
    if "0963830315" in content and "Tâm" in content:
        return content

    return f"{content}\n\n{NOTIFICATION_CONTACT_FOOTER}"

def cleanup_old_notifications():
    """
    Tự dọn thông báo cũ để Supabase không phình dữ liệu.

    Quy tắc:
    - Thông báo đã đọc: xóa sau 15 ngày.
    - Thông báo chưa đọc: xóa sau 30 ngày.
    """

    now = datetime.now(timezone.utc)

    read_cutoff = (now - timedelta(days=15)).isoformat()
    unread_cutoff = (now - timedelta(days=30)).isoformat()

    try:
        # =========================
        # 1. Lấy các thông báo đã đọc
        # =========================
        read_rows = supabase.table(NOTIFICATION_READ_TABLE) \
            .select("notification_id") \
            .execute().data or []

        read_ids = []

        for r in read_rows:
            nid = r.get("notification_id")

            if nid is not None and nid not in read_ids:
                read_ids.append(nid)

        # =========================
        # 2. Xóa thông báo đã đọc quá 15 ngày
        # =========================
        if read_ids:
            supabase.table(NOTIFICATION_TABLE) \
                .delete() \
                .in_("id", read_ids) \
                .lt("created_at", read_cutoff) \
                .execute()

        # =========================
        # 3. Xóa thông báo chưa đọc quá 30 ngày
        # =========================
        old_notifications = supabase.table(NOTIFICATION_TABLE) \
            .select("id,created_at") \
            .lt("created_at", unread_cutoff) \
            .execute().data or []

        old_unread_ids = []

        for n in old_notifications:
            nid = n.get("id")

            if nid is None:
                continue

            # Nếu không nằm trong read_ids thì xem là chưa đọc
            if nid not in read_ids:
                old_unread_ids.append(nid)

        if old_unread_ids:
            supabase.table(NOTIFICATION_TABLE) \
                .delete() \
                .in_("id", old_unread_ids) \
                .execute()

        # =========================
        # 4. Xóa read logs mồ côi nếu notification đã mất
        # =========================
        existing_notifications = supabase.table(NOTIFICATION_TABLE) \
            .select("id") \
            .execute().data or []

        existing_ids = {
            str(n.get("id"))
            for n in existing_notifications
            if n.get("id") is not None
        }

        all_read_rows = supabase.table(NOTIFICATION_READ_TABLE) \
            .select("id,notification_id") \
            .execute().data or []

        orphan_read_ids = []

        for r in all_read_rows:
            notification_id = str(r.get("notification_id") or "")

            if notification_id and notification_id not in existing_ids:
                orphan_read_ids.append(r.get("id"))

        if orphan_read_ids:
            supabase.table(NOTIFICATION_READ_TABLE) \
                .delete() \
                .in_("id", orphan_read_ids) \
                .execute()

    except Exception as e:
        print("[CLEANUP NOTIFICATIONS ERROR]", e)

def personalize_notification_content(content, student_name="", is_all=False):
    content = str(content or "").strip()
    student_name = str(student_name or "").strip()

    if not content:
        return ""

    lines = content.splitlines()

    greeting_all = "Kính gửi quý Phụ huynh và võ sinh!"
    greeting_student = f"Kính gửi quý Phụ huynh và võ sinh: {student_name}" if student_name else "Kính gửi quý Phụ huynh và võ sinh:"

    greeting_regex = re.compile(r"^Kính gửi quý Phụ huynh và võ sinh[:!].*", re.IGNORECASE)

    if is_all:
        new_greeting = greeting_all
    else:
        new_greeting = greeting_student

    if lines:
        first_line = lines[0].strip()

        if greeting_regex.match(first_line):
            lines[0] = new_greeting
        else:
            lines.insert(0, new_greeting)
            lines.insert(1, "")
    else:
        lines = [new_greeting, ""]

    return "\n".join(lines).strip()

def build_notification_filter_ma_quy(target_mode, year, exam_quarter="", dan_round=""):
    """
    Tạo mã kỳ thi để lọc thông báo theo học phí đã đăng ký thi.
    - Thi cấp: Q12026, Q22026...
    - Thi đẳng: L1-2026, L2-2026, MN-2026...
    """
    target_mode = str(target_mode or "").strip()
    year = str(year or datetime.now().year).strip()
    exam_quarter = str(exam_quarter or "").strip().upper()
    dan_round = str(dan_round or "").strip().upper()

    if target_mode in ["exam_registered", "exam_not_registered"]:
        if not exam_quarter:
            return ""
        return f"{exam_quarter}{year}"

    if target_mode in ["dan_registered", "dan_not_registered"]:
        if not dan_round:
            return ""
        return f"{dan_round}-{year}"

    return ""


def get_students_by_notification_target(target_mode, students, year, exam_quarter="", dan_round=""):
    """
    Trả về danh sách học viên theo tùy chọn gửi thông báo:
    - exam_registered: đã đăng ký thi cấp kỳ chọn.
    - exam_not_registered: chưa đăng ký thi cấp kỳ chọn.
    - dan_registered: đã đăng ký thi đẳng kỳ/lần chọn.
    - dan_not_registered: chưa đăng ký thi đẳng kỳ/lần chọn.
    """
    target_mode = str(target_mode or "").strip()

    ma_quy = build_notification_filter_ma_quy(
        target_mode=target_mode,
        year=year,
        exam_quarter=exam_quarter,
        dan_round=dan_round
    )

    if not ma_quy:
        return [], ""

    fees = safe_rows(HOCPHI_TABLE, "ma_hv,ma_quy")

    registered_ids = {
        str(f.get("ma_hv") or "").strip()
        for f in fees
        if str(f.get("ma_quy") or "").strip().upper() == ma_quy.upper()
    }

    active_students = [
        s for s in students
        if is_active(s.get("active")) and str(s.get("license") or "").strip()
    ]

    if target_mode in ["exam_registered", "dan_registered"]:
        filtered_students = [
            s for s in active_students
            if str(s.get("license") or "").strip() in registered_ids
        ]

    elif target_mode in ["exam_not_registered", "dan_not_registered"]:
        filtered_students = [
            s for s in active_students
            if str(s.get("license") or "").strip() not in registered_ids
        ]

    else:
        filtered_students = []

    return filtered_students, ma_quy


def get_notification_target_label(target_mode, year, exam_quarter="", dan_round=""):
    target_mode = str(target_mode or "").strip()
    year = str(year or datetime.now().year).strip()
    exam_quarter = str(exam_quarter or "").strip().upper()
    dan_round = str(dan_round or "").strip().upper()

    exam_label_map = {
        "Q1": "Quý 1",
        "Q2": "Quý 2",
        "Q3": "Quý 3",
        "Q4": "Quý 4",
    }

    dan_label_map = {
        "L1": "Lần 1",
        "L2": "Lần 2",
        "MN": "KV miền Nam",
        "MT": "KV miền Trung",
        "MB": "KV miền Bắc",
        "QG": "Quốc Gia",
    }

    if target_mode == "exam_registered":
        return f"Thi cấp - Đã đăng ký - {exam_label_map.get(exam_quarter, exam_quarter)} - {year}"

    if target_mode == "exam_not_registered":
        return f"Thi cấp - Chưa đăng ký - {exam_label_map.get(exam_quarter, exam_quarter)} - {year}"

    if target_mode == "dan_registered":
        return f"Thi đẳng - Đã đăng ký - {dan_label_map.get(dan_round, dan_round)} - {year}"

    if target_mode == "dan_not_registered":
        return f"Thi đẳng - Chưa đăng ký - {dan_label_map.get(dan_round, dan_round)} - {year}"

    return "Tùy chọn"

@app.route("/notifications", methods=["GET", "POST"])
def notifications():
    cleanup_old_notifications()

    students = safe_rows(STUDENT_TABLE)
    students.sort(key=lambda s: str(s.get("name") or ""))

    now_year = datetime.now().year
    year_options = [now_year - 1, now_year, now_year + 1]

    if request.method == "POST":
        target_mode = request.form.get("target_mode", "all").strip()
        target_licenses = request.form.getlist("target_licenses")
        title = request.form.get("title", "").strip()
        content = request.form.get("content", "").strip()

        notify_year = request.form.get("notify_year", str(now_year)).strip() or str(now_year)
        exam_quarter = request.form.get("exam_quarter", "").strip().upper()
        dan_round = request.form.get("dan_round", "").strip().upper()

        target_licenses = [
            str(x or "").strip()
            for x in target_licenses
            if str(x or "").strip()
        ]

        if not title:
            flash("Ken chưa nhập tiêu đề thông báo.", "danger")
            return redirect(url_for("notifications"))

        if not content:
            flash("Ken chưa nhập nội dung thông báo.", "danger")
            return redirect(url_for("notifications"))

        try:
            # =========================
            # GỬI CHO TẤT CẢ
            # =========================
            if target_mode == "all":
                final_content = personalize_notification_content(
                    content,
                    student_name="",
                    is_all=True
                )

                payload = {
                    "target_type": "all",
                    "target_license": None,
                    "target_name": "Tất cả học viên",
                    "title": title,
                    "content": append_notification_footer(final_content),
                }

                supabase.table(NOTIFICATION_TABLE).insert(payload).execute()
                flash("Đã gửi thông báo cho tất cả học viên.", "success")
                return redirect(url_for("notifications"))

            student_map = {
                str(s.get("license") or "").strip(): s
                for s in students
            }

            payloads = []

            # =========================
            # GỬI THEO NHÓM THI CẤP / THI ĐẲNG
            # =========================
            group_target_modes = [
                "exam_registered",
                "exam_not_registered",
                "dan_registered",
                "dan_not_registered",
            ]

            if target_mode in group_target_modes:
                if target_mode in ["exam_registered", "exam_not_registered"] and not exam_quarter:
                    flash("Ken chưa chọn quý thi cấp.", "danger")
                    return redirect(url_for("notifications"))

                if target_mode in ["dan_registered", "dan_not_registered"] and not dan_round:
                    flash("Ken chưa chọn lần/khu vực thi đẳng.", "danger")
                    return redirect(url_for("notifications"))

                target_students, ma_quy = get_students_by_notification_target(
                    target_mode=target_mode,
                    students=students,
                    year=notify_year,
                    exam_quarter=exam_quarter,
                    dan_round=dan_round
                )

                target_label = get_notification_target_label(
                    target_mode=target_mode,
                    year=notify_year,
                    exam_quarter=exam_quarter,
                    dan_round=dan_round
                )

                if not target_students:
                    flash(f"Không có học viên phù hợp với nhóm: {target_label}.", "danger")
                    return redirect(url_for("notifications"))

                for student in target_students:
                    license_code = str(student.get("license") or "").strip()
                    student_name = student.get("name", "")

                    final_content = personalize_notification_content(
                        content,
                        student_name=student_name,
                        is_all=False
                    )

                    payloads.append({
                        "target_type": target_mode,
                        "target_license": license_code,
                        "target_name": student_name,
                        "title": title,
                        "content": append_notification_footer(final_content),
                    })

                supabase.table(NOTIFICATION_TABLE).insert(payloads).execute()

                flash(
                    f"Đã gửi thông báo cho {len(payloads)} học viên thuộc nhóm: {target_label}.",
                    "success"
                )
                return redirect(url_for("notifications"))

            # =========================
            # GỬI TÙY CHỌN NHIỀU HỌC VIÊN
            # =========================
            if not target_licenses:
                flash("Ken chưa chọn học viên nhận thông báo.", "danger")
                return redirect(url_for("notifications"))

            for license_code in target_licenses:
                student = student_map.get(license_code)

                if not student:
                    continue

                student_name = student.get("name", "")

                final_content = personalize_notification_content(
                    content,
                    student_name=student_name,
                    is_all=False
                )

                payloads.append({
                    "target_type": "student",
                    "target_license": license_code,
                    "target_name": student_name,
                    "title": title,
                    "content": append_notification_footer(final_content),
                })

            if not payloads:
                flash("Không tìm thấy học viên hợp lệ để gửi thông báo.", "danger")
                return redirect(url_for("notifications"))

            supabase.table(NOTIFICATION_TABLE).insert(payloads).execute()

            flash(f"Đã gửi thông báo cho {len(payloads)} học viên.", "success")
            return redirect(url_for("notifications"))

        except Exception as e:
            print("[ADD NOTIFICATION ERROR]", e)
            flash(f"Lỗi gửi thông báo: {e}", "danger")
            return redirect(url_for("notifications"))

    notifications_rows = safe_rows(NOTIFICATION_TABLE)
    notifications_rows.sort(key=lambda r: str(r.get("created_at") or ""), reverse=True)

    return render_template(
        "notifications.html",
        rows=students,
        notifications=notifications_rows,
        current_year=now_year,
        year_options=year_options
    )

@app.post("/notifications/delete/<notification_id>")
def notification_delete(notification_id):
    """
    Xóa thủ công một thông báo khỏi Supabase.

    Sau khi xóa:
    - Học viên thuộc thông báo đó không còn thấy trong danh sách.
    - Link chi tiết cũ cũng không mở được nữa.
    - Xóa luôn lịch sử đã đọc liên quan để tránh dữ liệu mồ côi.
    """
    notification_id = str(notification_id or "").strip()

    if not notification_id:
        flash("Thiếu mã thông báo cần xóa.", "danger")
        return back_to_current_page("notifications")

    try:
        rows = (
            supabase.table(NOTIFICATION_TABLE)
            .select("id,title,target_type,target_license,target_name")
            .eq("id", notification_id)
            .limit(1)
            .execute()
            .data
            or []
        )

        if not rows:
            flash(
                "Thông báo không còn tồn tại hoặc đã được xóa trước đó.",
                "warning"
            )
            return back_to_current_page("notifications")

        notification = rows[0]

        # Xóa lịch sử đã đọc trước để tránh vướng khóa ngoại,
        # đồng thời không để lại dữ liệu mồ côi.
        try:
            (
                supabase.table(NOTIFICATION_READ_TABLE)
                .delete()
                .eq("notification_id", notification_id)
                .execute()
            )
        except Exception as read_error:
            print("[DELETE NOTIFICATION READ LOG ERROR]", read_error)

        # Xóa thông báo gốc.
        # Trang học viên đọc trực tiếp từ bảng notifications,
        # nên sau khi xóa học viên sẽ không còn nhìn thấy.
        (
            supabase.table(NOTIFICATION_TABLE)
            .delete()
            .eq("id", notification_id)
            .execute()
        )

        title = str(
            notification.get("title") or "Thông báo"
        ).strip()

        target_name = str(
            notification.get("target_name")
            or notification.get("target_license")
            or "Tất cả học viên"
        ).strip()

        flash(
            f"Đã xóa thông báo “{title}” của {target_name}. "
            f"Học viên sẽ không còn nhìn thấy thông báo này.",
            "success"
        )

    except Exception as e:
        print("[DELETE NOTIFICATION ERROR]", e)
        flash(f"Lỗi xóa thông báo: {e}", "danger")

    return back_to_current_page("notifications")
    
@app.route("/student-login", methods=["GET", "POST"])
def student_login():
    if request.method == "POST":
        action = request.form.get("action", "login").strip()

        # =========================
        # QUÊN MẬT KHẨU HỌC VIÊN
        # Reset về mặc định:
        # username = Mã HV hiện tại
        # password = SĐT đăng ký
        # =========================
        if action == "forgot_password":
            username = request.form.get("reset_username", "").strip()
            birthdate_raw = request.form.get("reset_birthdate", "").strip()
            phone_raw = request.form.get("reset_phone", "").strip()

            birthdate = normalize_birthdate_web(birthdate_raw)
            phone = re.sub(r"\D", "", phone_raw)

            if not username or not birthdate or not phone:
                flash("Vui lòng nhập đủ Mã HV/Username, ngày sinh và SĐT đăng ký.", "danger")
                return redirect(url_for("student_login"))

            students = safe_rows(STUDENT_TABLE)

            matched = None

            for s in students:
                license_code = str(s.get("license") or "").strip()
                portal_username = str(s.get("portal_username") or "").strip()
                student_birthdate = str(s.get("birthdate") or "").strip()
                student_phone = re.sub(r"\D", "", str(s.get("phonenumber") or ""))

                same_account = username in [license_code, portal_username]
                same_birthdate = birthdate == student_birthdate
                same_phone = phone == student_phone

                if same_account and same_birthdate and same_phone:
                    matched = s
                    break

            if not matched:
                flash("Thông tin xác minh chưa đúng. Ken kiểm tra lại Mã HV, ngày sinh hoặc SĐT đăng ký.", "danger")
                return redirect(url_for("student_login"))

            license_code = str(matched.get("license") or "").strip()
            student_phone = re.sub(r"\D", "", str(matched.get("phonenumber") or ""))

            if not license_code or not student_phone:
                flash("Tài khoản thiếu Mã HV hoặc SĐT nên chưa thể reset mật khẩu.", "danger")
                return redirect(url_for("student_login"))

            supabase.table(STUDENT_TABLE) \
                .update({
                    "portal_username": license_code,
                    "portal_password_hash": generate_password_hash(student_phone)
                }) \
                .eq("license", license_code) \
                .execute()

            flash("Đã cài lại tài khoản về mặc định. Tên đăng nhập = Mã HV, mật khẩu = SĐT.", "success")
            return redirect(url_for("student_login"))

        # =========================
        # ĐĂNG NHẬP BÌNH THƯỜNG
        # =========================
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        if not username or not password:
            flash("Vui lòng nhập tên đăng nhập và mật khẩu.", "danger")
            return redirect(url_for("student_login"))

        students = safe_rows(STUDENT_TABLE)

        matched = None

        for s in students:
            ensure_student_portal_account(s)

            portal_username = str(s.get("portal_username") or "").strip()
            license_code = str(s.get("license") or "").strip()
            phone = str(s.get("phonenumber") or "").strip()

            # Cho phép đăng nhập bằng username, mã HV hoặc SĐT
            if username in [portal_username, license_code, phone]:
                matched = s
                break

        if not matched:
            flash("Không tìm thấy tài khoản học viên.", "danger")
            return redirect(url_for("student_login"))

        password_hash = str(matched.get("portal_password_hash") or "").strip()

        # Nếu học viên cũ chưa có hash, dùng SĐT làm mật khẩu lần đầu
        if not password_hash:
            matched = ensure_student_portal_account(matched)
            password_hash = str(matched.get("portal_password_hash") or "").strip()

        if not check_password_hash(password_hash, password):
            flash("Mật khẩu không đúng.", "danger")
            return redirect(url_for("student_login"))

        session["student_logged_in"] = True
        session["student_license"] = matched.get("license")
        session["student_name"] = matched.get("name")
        session["show_student_welcome_popup"] = True

        flash("Đăng nhập thành công.", "success")
        return redirect(url_for("student_portal_notifications"))

    return render_template("student_login.html")


@app.get("/student-logout")
def student_logout():
    session.pop("student_logged_in", None)
    session.pop("student_license", None)
    session.pop("student_name", None)
    session.pop("show_student_welcome_popup", None)

    flash("Đã đăng xuất.", "success")
    return redirect(url_for("student_login"))

# =========================
# COACH PORTAL - HLV
# =========================

def normalize_phone_web(value):
    return re.sub(r"\D", "", str(value or ""))


def is_active_text(value):
    text = remove_accents(str(value or "")).strip().lower()
    return text in ["co", "yes", "true", "1", "active", "dang hoat dong"]


def get_logged_coach():
    coach_code = str(session.get("coach_code") or "").strip()

    if not coach_code:
        return None

    try:
        rows = supabase.table(COACH_TABLE) \
            .select("*") \
            .eq("coach_code", coach_code) \
            .limit(1) \
            .execute().data or []

        if not rows:
            session.pop("coach_logged_in", None)
            session.pop("coach_code", None)
            return None

        coach = rows[0]

        if not is_active_text(coach.get("active")):
            session.pop("coach_logged_in", None)
            session.pop("coach_code", None)
            return None

        return coach

    except Exception as e:
        print("[GET LOGGED COACH ERROR]", e)
        return None


def require_coach_login():
    if not session.get("coach_logged_in"):
        return None

    return get_logged_coach()


@app.route("/coach-login", methods=["GET", "POST"])
def coach_login():
    if request.method == "POST":
        coach_code = request.form.get("coach_code", "").strip()
        phone = normalize_phone_web(request.form.get("phone", ""))

        if not coach_code or not phone:
            flash("Vui lòng nhập Mã HLV và SĐT.", "danger")
            return redirect(url_for("coach_login"))

        try:
            rows = supabase.table(COACH_TABLE) \
                .select("*") \
                .eq("coach_code", coach_code) \
                .limit(1) \
                .execute().data or []

            if not rows:
                flash("Không tìm thấy Mã HLV.", "danger")
                return redirect(url_for("coach_login"))

            coach = rows[0]

            if not is_active_text(coach.get("active")):
                flash("Tài khoản HLV này đang bị khóa.", "danger")
                return redirect(url_for("coach_login"))

            db_phone = normalize_phone_web(coach.get("phone"))

            if not db_phone or db_phone != phone:
                flash("SĐT đăng nhập chưa đúng.", "danger")
                return redirect(url_for("coach_login"))

            # Không dùng session.clear(), vì sẽ làm out Admin và Student
            session["coach_logged_in"] = True
            session["coach_code"] = coach.get("coach_code")
            session["coach_name"] = coach.get("name")

            flash("Đăng nhập HLV thành công.", "success")
            return redirect(url_for("coach_exam"))

        except Exception as e:
            print("[COACH LOGIN ERROR]", e)
            flash(f"Lỗi đăng nhập HLV: {e}", "danger")
            return redirect(url_for("coach_login"))

    return render_template("coach_login.html")


@app.get("/coach-logout")
def coach_logout():
    session.pop("coach_logged_in", None)
    session.pop("coach_code", None)
    session.pop("coach_name", None)

    flash("Đã đăng xuất HLV.", "success")
    return redirect(url_for("coach_login"))


@app.get("/coach/exam")
def coach_exam():
    coach = require_coach_login()

    if not coach:
        return redirect(url_for("coach_login"))

    now = datetime.now()

    year = request.args.get("year", str(now.year)).strip()
    quarter = request.args.get(
        "quarter",
        f"Q{(now.month - 1) // 3 + 1}"
    ).strip().upper()

    if quarter not in ["Q1", "Q2", "Q3", "Q4"]:
        quarter = f"Q{(now.month - 1) // 3 + 1}"

    rows = get_exam_list_rows_by_type_web(
        year=year,
        quarter=quarter,
        list_type="cap"
    )

    rows.sort(key=lambda r: (
        get_belt_index_web(r.get("cap_du_thi")),
        str(r.get("gender") or ""),
        str(r.get("name") or "")
    ))

    status_summary = {
        "total": len(rows),
        "Đúng": 0,
        "Sai": 0,
        "Ktra lại": 0,
        "Chưa KTra": 0,
    }

    for r in rows:
        status = str(r.get("coach_check_status") or "").strip()

        if not status:
            status = "Chưa KTra"

        if status not in status_summary:
            status_summary[status] = 0

        status_summary[status] += 1

    years = [str(y) for y in range(now.year - 2, now.year + 3)]
    quarters = ["Q1", "Q2", "Q3", "Q4"]

    return render_template(
        "coach_exam.html",
        coach=coach,
        rows=rows,
        year=year,
        quarter=quarter,
        years=years,
        quarters=quarters,
        status_summary=status_summary
    )

@app.get("/coach/dan")
def coach_dan():
    coach = require_coach_login()

    if not coach:
        return redirect(url_for("coach_login"))

    now = datetime.now()

    year = request.args.get("year", str(now.year)).strip()

    # Nhận cả 2 tên để tránh lỗi HTML cũ / mới:
    # - quarter: đang dùng trong coach_dan.html
    # - dan_round: dự phòng nếu sau này đổi tên input
    dan_round = (
        request.args.get("quarter")
        or request.args.get("dan_round")
        or "L1"
    )
    dan_round = str(dan_round or "L1").strip().upper()

    dan_round_options = [
        {"value": "L1", "label": "Lần 1"},
        {"value": "L2", "label": "Lần 2"},
        {"value": "MN", "label": "KV miền Nam"},
        {"value": "MT", "label": "KV miền Trung"},
        {"value": "MB", "label": "KV miền Bắc"},
        {"value": "QG", "label": "Quốc Gia"},
    ]

    dan_round_values = [
        item["value"]
        for item in dan_round_options
    ]

    if dan_round not in dan_round_values:
        dan_round = "L1"

    quarter_label = next(
        (
            item["label"]
            for item in dan_round_options
            if item["value"] == dan_round
        ),
        dan_round
    )

    rows = get_exam_list_rows_by_type_web(
        year=year,
        quarter=dan_round,
        list_type="dan"
    )

    rows.sort(key=lambda r: (
        get_belt_index_web(r.get("cap_du_thi")),
        str(r.get("gender") or ""),
        str(r.get("name") or "")
    ))

    status_summary = {
        "total": len(rows),
        "Đúng": 0,
        "Sai": 0,
        "Ktra lại": 0,
        "Chưa KTra": 0,
    }

    for r in rows:
        status = str(r.get("coach_check_status") or "").strip()

        if not status:
            status = "Chưa KTra"

        if status not in status_summary:
            status_summary[status] = 0

        status_summary[status] += 1

    years = [str(y) for y in range(now.year - 2, now.year + 3)]

    return render_template(
        "coach_dan.html",
        coach=coach,
        rows=rows,
        year=year,

        # Biến dùng cho coach_dan.html hiện tại
        quarter=dan_round,
        quarters=dan_round_options,
        quarter_label=quarter_label,

        # Biến dự phòng nếu sau này HTML dùng dan_round
        dan_round=dan_round,
        dan_round_options=dan_round_options,

        years=years,
        status_summary=status_summary
    )

@app.post("/coach/exam/update-status")
def coach_exam_update_status():
    coach = require_coach_login()

    if not coach:
        return jsonify({
            "ok": False,
            "message": "HLV chưa đăng nhập."
        }), 401

    data = request.get_json(silent=True) or {}

    hocphi_id = str(data.get("hocphi_id") or "").strip()
    status = str(data.get("status") or "").strip()
    note = str(data.get("note") or "").strip()

    allowed_status = ["Chưa KTra", "Đúng", "Sai", "Ktra lại"]

    if not hocphi_id:
        return jsonify({
            "ok": False,
            "message": "Thiếu ID học phí."
        }), 400

    if status not in allowed_status:
        return jsonify({
            "ok": False,
            "message": "Trạng thái không hợp lệ."
        }), 400

    try:
        payload = {
            "coach_check_status": status,
            "coach_check_note": note,
            "coach_checked_by": str(coach.get("coach_code") or ""),
            "coach_checked_by_name": str(coach.get("name") or ""),
            "coach_checked_at": datetime.now(timezone.utc).isoformat(),
        }

        supabase.table(HOCPHI_TABLE) \
            .update(payload) \
            .eq("id", hocphi_id) \
            .execute()

        return jsonify({
            "ok": True,
            "message": "Đã lưu kiểm tra HLV.",
            "status": status,
            "checked_by_name": payload["coach_checked_by_name"],
            "checked_at": payload["coach_checked_at"],
        })

    except Exception as e:
        print("[COACH EXAM UPDATE STATUS ERROR]", e)

        return jsonify({
            "ok": False,
            "message": str(e)
        }), 500

@app.get("/student-portal")
def student_portal_home():
    return redirect(url_for("student_portal_notifications"))


@app.get("/student-portal/notifications")
def student_portal_notifications():
    cleanup_old_notifications()

    student = require_student_login()
    if not student:
        return redirect(url_for("student_login"))

    notifications_rows, unread_count = get_student_notifications(student.get("license"))

    return render_template(
        "student_portal_notifications.html",
        student=student,
        notifications=notifications_rows,
        unread_count=unread_count
    )


@app.get("/student-portal/notifications/<notification_id>")
def student_portal_notification_detail(notification_id):
    cleanup_old_notifications()

    student = require_student_login()
    if not student:
        return redirect(url_for("student_login"))

    rows = safe_rows(NOTIFICATION_TABLE, "*", id=notification_id)
    notification = rows[0] if rows else None

    if not notification:
        flash("Không tìm thấy thông báo.", "danger")
        return redirect(url_for("student_portal_notifications"))

    notification["content"] = append_notification_footer(notification.get("content", ""))

    target_type = str(notification.get("target_type") or "all")
    target_license = str(notification.get("target_license") or "")

    if target_type != "all" and target_license != str(student.get("license")):
        flash("Thông báo này không thuộc tài khoản của bạn.", "danger")
        return redirect(url_for("student_portal_notifications"))

    mark_notification_read(notification_id, student.get("license"))

    notifications_rows, unread_count = get_student_notifications(student.get("license"))

    return render_template(
        "student_portal_notification_detail.html",
        student=student,
        notification=notification,
        unread_count=unread_count
    )


@app.get("/student-portal/info")
def student_portal_info():
    student = require_student_login()
    if not student:
        return redirect(url_for("student_login"))

    notifications_rows, unread_count = get_student_notifications(student.get("license"))

    return render_template(
        "student_portal_info.html",
        student=student,
        unread_count=unread_count,
        photo_url=get_student_photo_url(student.get("license"))
    )

@app.get("/student-portal/club-info")
def student_portal_club_info():
    student = require_student_login()
    if not student:
        return redirect(url_for("student_login"))

    license_code = student.get("license")
    notifications_rows, unread_count = get_student_notifications(license_code)

    settings = load_app_settings()
    club_info = settings.get("club_info", DEFAULT_APP_SETTINGS.get("club_info", {}))

    return render_template(
        "student_portal_club_info.html",
        student=student,
        unread_count=unread_count,
        club_info=club_info
    )

@app.get("/student-portal/fees")
def student_portal_fees():
    student = require_student_login()
    if not student:
        return redirect(url_for("student_login"))

    license_code = student.get("license")

    fees = safe_rows(HOCPHI_TABLE, "*", ma_hv=license_code)
    fees.sort(key=lambda r: str(r.get("thoi_gian") or ""), reverse=True)

    def is_exam_fee_row(f):
        title = str(f.get("thang_dong_phi") or "").lower()
        ma_quy = str(f.get("ma_quy") or "").strip()

        return bool(
            ma_quy
            or "thi cấp" in title
            or "thi đẳng" in title
        )


    def format_month_code_for_portal(raw):
        raw = str(raw or "").strip()

        if not raw:
            return "Chưa có"

        # Nếu có nhiều tháng: 062026 - 072026
        first_code = raw.split("-")[0].strip()

        if re.fullmatch(r"\d{6}", first_code):
            return f"{first_code[:2]}/{first_code[2:]}"

        return raw


    def format_exam_code_for_portal(f):
        ma_quy = str(f.get("ma_quy") or "").strip()
        title = str(f.get("thang_dong_phi") or "").strip()

        if title:
            # VD: Thi cấp quý 2-2026
            return title.replace("Thi cấp quý ", "Q").replace("-2026", "/2026")

        if ma_quy:
            return format_ma_quy_label_web(ma_quy)

        return "Chưa có"


    tuition_fees = [
        f for f in fees
        if not is_exam_fee_row(f)
    ]

    exam_fees = [
        f for f in fees
        if is_exam_fee_row(f)
    ]

    latest_tuition_text = "Chưa có"
    latest_exam_text = "Chưa có"

    if tuition_fees:
        latest_tuition_text = format_month_code_for_portal(
            tuition_fees[0].get("ma_thang") or tuition_fees[0].get("thang_dong_phi")
        )

    if exam_fees:
        latest_exam_text = format_exam_code_for_portal(exam_fees[0])

    tuition_count = len(tuition_fees)
    exam_count = len(exam_fees)
    # =========================
    # NHẮC HẠN ĐÓNG HỌC PHÍ
    # Quy tắc:
    # - Nếu đã đóng tháng hiện tại: báo đã đóng.
    # - Nếu có đóng tháng trước: lấy ngày đóng tháng trước làm hạn tháng này.
    # - Nếu không có tháng trước: xem là võ sinh vãng lai / chưa có dữ liệu liên tiếp.
    # =========================
    today = date.today()
    current_code = f"{today.month:02d}{today.year}"

    if today.month == 1:
        prev_month = 12
        prev_year = today.year - 1
    else:
        prev_month = today.month - 1
        prev_year = today.year

    prev_code = f"{prev_month:02d}{prev_year}"

    fee_due_info = {
        "status": "casual",
        "status_class": "fee-due-casual",
        "title": "Chưa có dữ liệu tháng trước",
        "main": "Võ sinh tạm ngưng tập",
        "sub": "Nhanh chóng đóng phí để tiếp tục tập luyện nhé.",
        "paid_date": "—",
        "due_date": "—",
        "days_text": "—",
    }

    current_month_fees = []
    prev_month_fees = []

    for f in fees:
        ma_thang = str(f.get("ma_thang") or "")

        if current_code in ma_thang:
            current_month_fees.append(f)

        if prev_code in ma_thang:
            prev_month_fees.append(f)

    def parse_fee_dt(raw):
        try:
            return datetime.fromisoformat(str(raw or "").replace("Z", "+00:00"))
        except:
            return None

    if current_month_fees:
        latest_current = current_month_fees[0]
        paid_dt = parse_fee_dt(latest_current.get("thoi_gian"))

        fee_due_info = {
            "status": "paid",
            "status_class": "fee-due-green",
            "title": "Đã đóng tháng này",
            "main": "Học phí tháng hiện tại đã hoàn tất",
            "sub": "Cảm ơn phụ huynh và võ sinh đã đồng hành cùng CLB.",
            "paid_date": paid_dt.strftime("%d/%m/%Y") if paid_dt else "—",
            "due_date": "Đã đóng",
            "days_text": "Đã hoàn tất",
        }

    elif prev_month_fees:
        latest_prev = prev_month_fees[0]
        paid_dt = parse_fee_dt(latest_prev.get("thoi_gian"))

        if paid_dt:
            last_day = calendar.monthrange(today.year, today.month)[1]
            due_day = min(paid_dt.day, last_day)
            due_date = date(today.year, today.month, due_day)

            days_left = (due_date - today).days

            if days_left >= 8:
                status_class = "fee-due-green"
                title = "Chưa đến hạn"
                days_text = f"Còn {days_left} ngày"
                sub = "Học phí vẫn còn hạn."

            elif 4 <= days_left <= 7:
                status_class = "fee-due-yellow"
                title = "Gần đến hạn"
                days_text = f"Còn {days_left} ngày"
                sub = "Sắp đến hạn đóng học phí, phụ huynh/võ sinh lưu ý giúp CLB."

            elif 0 <= days_left <= 3:
                status_class = "fee-due-orange"
                title = "Sắp đến hạn"
                days_text = f"Còn {days_left} ngày"
                sub = "Phụ huynh/võ sinh vui lòng chuẩn bị đóng phí để không gián đoạn tập luyện."

            else:
                status_class = "fee-due-red"
                title = "Quá hạn"
                days_text = f"Quá hạn {abs(days_left)} ngày"
                sub = "Phụ huynh/võ sinh nhanh chóng đóng phí để tiếp tục tập luyện nhé."

            fee_due_info = {
                "status": "normal",
                "status_class": status_class,
                "title": title,
                "main": days_text,
                "sub": sub,
                "paid_date": paid_dt.strftime("%d/%m/%Y"),
                "due_date": due_date.strftime("%d/%m/%Y"),
                "days_text": days_text,
            }

    notifications_rows, unread_count = get_student_notifications(license_code)

    payment_info = get_payment_settings()

    payment_transfer_note = build_transfer_note(
        payment_info.get("transfer_note", "{student_name}"),
        student
    )

    payment_qr_url = build_vietqr_url(payment_info, student)

    return render_template(
        "student_portal_fees.html",
        student=student,
        fees=fees,
        payment_info=payment_info,
        payment_qr_url=payment_qr_url,
        payment_transfer_note=payment_transfer_note,
        fee_due_info=fee_due_info,
        latest_tuition_text=latest_tuition_text,
        latest_exam_text=latest_exam_text,
        tuition_count=tuition_count,
        exam_count=exam_count
    )

def format_ma_quy_label_web(ma_quy):
    """
    Q22026 -> Quý 2 - 2026
    Q2-2026 -> Quý 2 - 2026
    2026-Q2 -> Quý 2 - 2026
    """
    s = str(ma_quy or "").strip().upper()

    if not s:
        return "Sẽ bổ sung"

    m = re.search(r"Q([1-4])\D*(20\d{2})", s)
    if m:
        return f"Quý {m.group(1)} - {m.group(2)}"

    m = re.search(r"(20\d{2})\D*Q([1-4])", s)
    if m:
        return f"Quý {m.group(2)} - {m.group(1)}"

    return s


def ma_quy_to_ky_thi_web(ma_quy):
    """
    Q22026 -> 2026-Q2
    Q2-2026 -> 2026-Q2
    2026-Q2 -> 2026-Q2
    """
    s = str(ma_quy or "").strip().upper()

    if not s:
        return ""

    m = re.search(r"Q([1-4])\D*(20\d{2})", s)
    if m:
        return f"{m.group(2)}-Q{m.group(1)}"

    m = re.search(r"(20\d{2})\D*Q([1-4])", s)
    if m:
        return f"{m.group(1)}-Q{m.group(2)}"

    return s


@app.get("/student-portal/exams")
def student_portal_exams():
    student = require_student_login()
    if not student:
        return redirect(url_for("student_login"))

    license_code = student.get("license")

    # =========================
    # KẾT QUẢ ĐÃ THI
    # =========================
    results = safe_rows(KETQUA_TABLE, "*", ma_hv=license_code)
    results.sort(key=lambda r: parse_ky_thi_web(r.get("ky_thi")), reverse=True)

    for r in results:
        ky = str(r.get("ky_thi") or "").strip()
        info = get_exam_info_web(ky)

        judges = info.get("judges", []) if info else []
        judge_names = []

        for j in judges:
            if not isinstance(j, dict):
                continue

            name = str(j.get("name") or "").strip()
            active = str(j.get("active") or "on").strip().lower()

            if name:
                status = "ON" if active == "on" else "OFF"
                judge_names.append(f"{name} ({status})")

        r["exam_date_text"] = info.get("exam_date") or "Sẽ bổ sung"
        r["exam_time_text"] = info.get("exam_time") or "Sẽ bổ sung"
        r["venue_text"] = info.get("venue") or "Sẽ bổ sung"

        supervisor_name = info.get("supervisor_name") or ""
        supervisor_active = str(info.get("supervisor_active") or "on").lower()
        supervisor_status = "ON" if supervisor_active == "on" else "OFF"

        r["supervisor_text"] = (
            f"{supervisor_name} ({supervisor_status})"
            if supervisor_name else "Sẽ bổ sung"
        )

        r["judges_text"] = " | ".join(judge_names) if judge_names else "Sẽ bổ sung"

    result_ky_set = {
        str(r.get("ky_thi") or "").strip().upper()
        for r in results
        if str(r.get("ky_thi") or "").strip()
    }

    # =========================
    # TÍNH CẤP DỰ THI TIẾP THEO CHO HỌC VIÊN
    # Logic đúng:
    # - Nếu kết quả mới nhất Đạt: cấp tiếp theo = cấp kế tiếp của cấp đã đạt
    # - Nếu Không đạt hoặc Vắng: thi lại đúng cấp đó
    # - Nếu chưa có kết quả: lấy cấp hiện tại trong student để tính cấp tiếp theo
    # =========================
    current_belt = normalize_belt_name_web(student.get("belt"))
    next_belt = get_next_belt_web(current_belt)

    latest_result = results[0] if results else None

    if latest_result:
        latest_exam_belt = normalize_belt_name_web(latest_result.get("cap_dai_thi"))
        latest_exam_result = str(latest_result.get("ket_qua") or "").strip()

        if latest_exam_result == "Đạt":
            next_belt = get_next_belt_web(latest_exam_belt)

        elif latest_exam_result in ["Không đạt", "Vắng"]:
            next_belt = latest_exam_belt

    fee_rows = safe_rows(HOCPHI_TABLE, "*", ma_hv=license_code)

    exam_fee_rows = []

    for f in fee_rows:
        ma_quy = str(f.get("ma_quy") or "").strip()

        if not ma_quy:
            continue

        ky_thi = ma_quy_to_ky_thi_web(ma_quy)

        if not ky_thi:
            continue

        exam_fee_rows.append({
            "raw": f,
            "ma_quy": ma_quy,
            "ky_thi": ky_thi,
            "label": format_ma_quy_label_web(ma_quy),
            "order": parse_ky_thi_web(ky_thi),
            "thoi_gian": str(f.get("thoi_gian") or "")
        })

    # Lấy kỳ thi đã đăng ký mới nhất
    exam_fee_rows.sort(
        key=lambda x: (
            int(x.get("order") or 0),
            str(x.get("thoi_gian") or "")
        ),
        reverse=True
    )

    selected_exam_row = exam_fee_rows[0] if exam_fee_rows else None

    not_registered_text = "Chưa đăng ký kỳ mới"

    if selected_exam_row:
        next_ky_thi = selected_exam_row["ky_thi"]
        next_quarter = selected_exam_row["label"]

        # Kiểm tra kỳ đăng ký mới nhất đã có kết quả chưa
        matched_result = None

        for r in results:
            if str(r.get("ky_thi") or "").strip().upper() == next_ky_thi.upper():
                matched_result = r
                break

        if matched_result:
            # Nếu kỳ đã có kết quả rồi thì không xem là kỳ sắp thi nữa.
            # Chuyển sang hiển thị cấp tiếp theo sau kết quả.
            next_status = "Chưa đăng ký kỳ mới"
            next_quarter = not_registered_text
            next_ky_thi = ""
            next_belt_display = next_belt
        else:
            # Nếu đã đóng phí kỳ mới nhưng chưa có kết quả thì đây là kỳ sắp thi.
            next_status = "Đã đăng ký thi"
            next_belt_display = next_belt

    else:
        next_ky_thi = ""
        next_quarter = not_registered_text
        next_status = not_registered_text
        next_belt_display = next_belt

    next_exam_info = get_exam_info_web(next_ky_thi) if next_ky_thi else {}

    next_exam_info = get_exam_info_web(next_ky_thi) if next_ky_thi else {}

    next_judges = next_exam_info.get("judges", []) if next_exam_info else []
    next_judge_names = []

    for j in next_judges:
        if not isinstance(j, dict):
            continue

        name = str(j.get("name") or "").strip()
        active = str(j.get("active") or "on").lower().strip()

        if name:
            status = "ON" if active == "on" else "OFF"
            next_judge_names.append(f"{name} ({status})")

    next_supervisor_name = next_exam_info.get("supervisor_name") or ""
    next_supervisor_active = str(next_exam_info.get("supervisor_active") or "on").lower().strip()
    next_supervisor_status = "ON" if next_supervisor_active == "on" else "OFF"

    next_exam = {
        "cap_du_thi": next_belt_display,
        "status": next_status,
        "quarter": next_quarter,
        "exam_date": next_exam_info.get("exam_date") or not_registered_text,
        "exam_time": next_exam_info.get("exam_time") or not_registered_text,
        "venue": next_exam_info.get("venue") or not_registered_text,
        "supervisor": (
            f"{next_supervisor_name} ({next_supervisor_status})"
            if next_supervisor_name else not_registered_text
        ),
        "judges": " | ".join(next_judge_names) if next_judge_names else not_registered_text,
    }

    notifications_rows, unread_count = get_student_notifications(license_code)

    return render_template(
        "student_portal_exams.html",
        student=student,
        results=results,
        next_exam=next_exam,
        unread_count=unread_count
    )

@app.get("/student-portal/activities")
def student_portal_activities():
    student = require_student_login()
    if not student:
        return redirect(url_for("student_login"))

    license_code = student.get("license")

    activities = safe_rows(HOATDONG_TABLE, "*", ma_hv=license_code)
    activities.sort(key=lambda r: str(r.get("thoi_gian") or ""), reverse=True)

    notifications_rows, unread_count = get_student_notifications(license_code)

    return render_template(
        "student_portal_activities.html",
        student=student,
        activities=activities,
        unread_count=unread_count
    )


@app.route("/student-portal/settings", methods=["GET", "POST"])
def student_portal_settings():
    student = require_student_login()
    if not student:
        return redirect(url_for("student_login"))

    license_code = student.get("license")

    if request.method == "POST":
        new_username = request.form.get("portal_username", "").strip()
        old_password = request.form.get("old_password", "").strip()
        new_password = request.form.get("new_password", "").strip()
        confirm_password = request.form.get("confirm_password", "").strip()

        update_data = {}

        if new_username:
            duplicated = supabase.table(STUDENT_TABLE) \
                .select("license,portal_username") \
                .eq("portal_username", new_username) \
                .neq("license", license_code) \
                .limit(1) \
                .execute().data or []

            if duplicated:
                flash("Tên đăng nhập này đã có người dùng.", "danger")
                return redirect(url_for("student_portal_settings"))

            update_data["portal_username"] = new_username

        if old_password or new_password or confirm_password:
            password_hash = str(student.get("portal_password_hash") or "").strip()

            if not check_password_hash(password_hash, old_password):
                flash("Mật khẩu cũ không đúng.", "danger")
                return redirect(url_for("student_portal_settings"))

            if not new_password:
                flash("Ken chưa nhập mật khẩu mới.", "danger")
                return redirect(url_for("student_portal_settings"))

            if new_password != confirm_password:
                flash("Xác nhận mật khẩu mới không khớp.", "danger")
                return redirect(url_for("student_portal_settings"))

            update_data["portal_password_hash"] = generate_password_hash(new_password)

        if update_data:
            supabase.table(STUDENT_TABLE) \
                .update(update_data) \
                .eq("license", license_code) \
                .execute()

            flash("Đã cập nhật tài khoản.", "success")

        return redirect(url_for("student_portal_settings"))

    notifications_rows, unread_count = get_student_notifications(license_code)

    return render_template(
        "student_portal_settings.html",
        student=student,
        unread_count=unread_count
    )


if __name__ == '__main__':
    app.run(
        host='0.0.0.0',
        port=5050,
        debug=False,
        use_reloader=False
    )
